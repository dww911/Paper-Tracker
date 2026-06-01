import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None
try:
    import requests
except ImportError:
    requests = None
try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


ROOT_DIR = Path(__file__).resolve().parents[1]
APP_SETTINGS_PATH = ROOT_DIR / "app_settings.json"
SKILL_DIR = ROOT_DIR / ".agents" / "skills" / "ptychography-paper-tracker"
SCRIPTS_DIR = SKILL_DIR / "scripts"
OUTPUT_DIR = ROOT_DIR / ".agents" / "skills"
PROFILES_PATH = SKILL_DIR / "research_profiles.json"
JOURNAL_METRICS_PATH = SKILL_DIR / "journal_metrics.csv"
PAPER_NOTES_DIR = OUTPUT_DIR / "paper_notes"
WEEKLY_REPORTS_DIR = OUTPUT_DIR / "weekly_reports"
EXPORTS_DIR = OUTPUT_DIR / "exports"
REVIEW_REPORTS_DIR = OUTPUT_DIR / "review_reports"
ROADMAP_REPORTS_DIR = OUTPUT_DIR / "roadmap_reports"
ANNUAL_REPORTS_DIR = OUTPUT_DIR / "annual_reports"
MY_NOTES_DIR = OUTPUT_DIR / "my_notes"
IDEA_NOTES_DIR = OUTPUT_DIR / "idea_notes"
PDFS_DIR = OUTPUT_DIR / "pdfs"
PROGRESS_REPORTS_DIR = OUTPUT_DIR / "progress_reports"
SEMINAR_REPORTS_DIR = OUTPUT_DIR / "seminar_reports"
INTRO_DRAFTS_DIR = OUTPUT_DIR / "intro_drafts"
TRACKER_SCRIPT = SCRIPTS_DIR / "ptychography_tracker.py"
WEB_RUN_LOG = ROOT_DIR / "web_run.log"
WEB_RUN_LOCK = ROOT_DIR / "web_run.lock"
ENV_PATH = ROOT_DIR / ".env"

if load_dotenv is not None:
    load_dotenv(ENV_PATH)
    load_dotenv(SKILL_DIR / ".env")

sys.path.insert(0, str(SCRIPTS_DIR))

from radar_db import (  # noqa: E402
    DB_PATH,
    connect,
    get_run,
    get_run_candidates,
    init_db,
    latest_run,
    paper_stable_id,
    save_reading_note_for_paper,
    sync_profiles,
    top_papers,
    update_paper_abstract_zh,
)
from abstract_translate import ensure_paper_chinese_abstract, is_missing_zh  # noqa: E402
from paper_ai_parse import (  # noqa: E402
    ai_parse_configured,
    call_llm,
    parse_structured_fields,
    note_needs_generation,
    parse_paper_with_ai,
    parsed_to_reading_note_payload,
)
from prompt_templates import ensure_prompt_template_records, prompt_status, recent_prompt_runs, render_template  # noqa: E402
from rag_indexer import (  # noqa: E402
    build_answer_prompt,
    format_context as format_rag_context,
    index_all as rag_index_all,
    index_paper as rag_index_paper,
    index_pdf_text as rag_index_pdf_text,
    rag_status,
    save_query as save_rag_query,
    search as rag_search,
    source_list as rag_source_list,
)
from agent_runner import (  # noqa: E402
    create_plan as create_agent_plan,
    execute_run as execute_agent_run,
    get_run as get_agent_run,
    list_agents,
    recent_runs as recent_agent_runs,
)
from idea_lab_service import (  # noqa: E402
    branch_meta,
    extract_fetch_keywords,
    group_evidence,
    polish_idea_structured,
    search_external_evidence,
    search_local_evidence,
)


app = FastAPI(title="Research Radar")
app.mount("/static", StaticFiles(directory=Path(__file__).parent / "static"), name="static")
templates = Jinja2Templates(directory=Path(__file__).parent / "templates")


def load_profiles_doc() -> dict:
    with open(PROFILES_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_profiles_doc(doc: dict) -> None:
    with open(PROFILES_PATH, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
    try:
        sync_profiles(doc)
    except (OSError, sqlite3.Error):
        pass


def rows(sql: str, params: tuple = ()) -> list[dict]:
    init_db()
    with connect() as conn:
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def row(sql: str, params: tuple = ()) -> Optional[dict]:
    init_db()
    with connect() as conn:
        result = conn.execute(sql, params).fetchone()
        return dict(result) if result else None


def execute(sql: str, params: tuple = ()) -> None:
    init_db()
    with connect() as conn:
        conn.execute(sql, params)
        conn.commit()


def execute_many(sql: str, params_list: list[tuple]) -> None:
    init_db()
    with connect() as conn:
        conn.executemany(sql, params_list)


def split_lines(value: str) -> list[str]:
    return [item.strip() for item in value.replace("\r", "").split("\n") if item.strip()]


def slugify_profile_id(value: str) -> str:
    value = re.sub(r"\s+", "_", str(value or "").strip().lower())
    value = re.sub(r"[^a-z0-9_]+", "", value)
    return value.strip("_") or "custom_profile"


DEFAULT_OUTPUT_FIELDS = [
    "论文名字",
    "网址",
    "期刊",
    "影响因子",
    "发布时间",
    "摘要中文翻译",
    "研究背景",
    "论文创新点",
    "实验结果",
    "总结",
    "未来展望",
    "可创新点",
]


def build_profile_payload(
    profile_id: str,
    name: str,
    display_name: str,
    description: str,
    include_keywords: str,
    must_have_any: str,
    exclude_keywords: str,
    research_focus: str,
    google_scholar_query: str,
    arxiv_categories: str,
    min_score: int,
    enable_wechat: bool,
    use_arxiv: bool,
    use_semantic: bool,
    use_google: bool,
    ingest_min_score: int = 1,
    ingest_below_must_have: bool = True,
    old: Optional[dict] = None,
) -> dict:
    old = old or {}
    focus_items = split_lines(research_focus)
    default_sources = {
        "arxiv": use_arxiv,
        "semantic_scholar": use_semantic,
        "google_scholar": use_google,
    }
    sources = [key for key, enabled in default_sources.items() if enabled] or ["arxiv"]
    return {
        **old,
        "name": name or profile_id,
        "display_name": display_name or name or profile_id,
        "description": description,
        "include_keywords": split_lines(include_keywords),
        "must_have_any": split_lines(must_have_any),
        "exclude_keywords": split_lines(exclude_keywords),
        "research_focus": focus_items,
        "google_scholar_query": google_scholar_query,
        "arxiv_categories": split_lines(arxiv_categories) or old.get("arxiv_categories") or [
            "physics.app-ph",
            "cond-mat.mtrl-sci",
        ],
        "enable_wechat": enable_wechat,
        "sources": sources,
        "default_sources": default_sources,
        "score_rules": {
            **old.get("score_rules", {}),
            "title_keyword": old.get("score_rules", {}).get("title_keyword", 3),
            "abstract_keyword": old.get("score_rules", {}).get("abstract_keyword", 1),
            "research_focus": old.get("score_rules", {}).get("research_focus", 2),
            "exclude_keyword": old.get("score_rules", {}).get("exclude_keyword", -5),
            "must_have_bonus": old.get("score_rules", {}).get("must_have_bonus", 2),
            "min_score": min_score,
        },
        "score_threshold": min_score,
        "parse_template": old.get("parse_template", "custom_research"),
        "parse_focus": old.get("parse_focus")
        or (
            [f"重点关注：{', '.join(focus_items)}"]
            if focus_items
            else ["结合当前研究方向给出具体、可落地的研究启发。"]
        ),
        "output_fields": old.get("output_fields") or DEFAULT_OUTPUT_FIELDS,
        "ingest_min_score": int(ingest_min_score),
        "ingest_below_must_have": bool(ingest_below_must_have),
    }


def parse_authors(value: str) -> str:
    try:
        parsed = json.loads(value or "[]")
        if isinstance(parsed, list):
            return ", ".join(parsed)
    except json.JSONDecodeError:
        pass
    return value or ""


templates.env.filters["authors"] = parse_authors


def today_string() -> str:
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d")


def week_string() -> str:
    from datetime import datetime
    year, week, _ = datetime.now().isocalendar()
    return f"{year}-W{week:02d}"


def current_year() -> int:
    from datetime import datetime
    return datetime.now().year


def safe_slug(value: str, limit: int = 80) -> str:
    value = re.sub(r"\s+", "_", str(value or "").strip())
    value = re.sub(r"[^A-Za-z0-9_\-\u4e00-\u9fff]+", "", value)
    return (value[:limit].strip("_") or "paper")


def score_to_rating(score) -> int:
    try:
        score = int(float(score or 0))
    except (TypeError, ValueError):
        score = 0
    if score >= 24:
        return 5
    if score >= 18:
        return 4
    if score >= 12:
        return 3
    if score >= 6:
        return 2
    return 1


def rating_label(rating) -> str:
    rating = max(1, min(int(rating or 1), 5))
    return {
        5: "强烈建议精读",
        4: "推荐阅读",
        3: "值得关注",
        2: "可归档备用",
        1: "低优先级",
    }[rating]


def star_text(rating) -> str:
    rating = max(1, min(int(rating or 1), 5))
    return "★" * rating + "☆" * (5 - rating)


def gold_stars(rating) -> str:
    rating = max(1, min(int(rating or 5), 5))
    return "★" * rating


templates.env.filters["stars"] = star_text
templates.env.filters["gold_stars"] = gold_stars
templates.env.filters["rating_label"] = rating_label


RATING_SQL = """
COALESCE(
    user_rating,
    system_rating,
    CASE
        WHEN COALESCE(final_score, relevance_score, 0) >= 24 THEN 5
        WHEN COALESCE(final_score, relevance_score, 0) >= 18 THEN 4
        WHEN COALESCE(final_score, relevance_score, 0) >= 12 THEN 3
        WHEN COALESCE(final_score, relevance_score, 0) >= 6 THEN 2
        ELSE 1
    END
)
"""


def paper_display_url(paper: dict) -> str:
    for key in ("display_url", "url", "arxiv_url", "pdf_url", "doi_url", "scholar_url", "publisher_url"):
        value = (paper.get(key) or "").strip()
        if value:
            return value
    doi = (paper.get("doi") or "").strip()
    if doi:
        return f"https://doi.org/{doi}"
    return ""


templates.env.filters["paper_url"] = paper_display_url

try:
    from journal_rank_enhancer import (  # noqa: E402
        export_unmatched_journals,
        journal_rank_summary,
        match_journal_rank,
        metrics_file_stats,
        paper_metrics_line,
    )
    from radar_db import rematch_journal_ranks  # noqa: E402
except ImportError:
    export_unmatched_journals = None
    journal_rank_summary = None
    match_journal_rank = None
    metrics_file_stats = None
    paper_metrics_line = None
    rematch_journal_ranks = None

templates.env.filters["journal_rank"] = journal_rank_summary or (
    lambda paper: paper.get("impact_factor") or "未匹配"
)
templates.env.filters["paper_metrics"] = paper_metrics_line or (
    lambda paper: paper.get("impact_factor") or "未匹配"
)


def _apply_journal_rank_filters(clauses: list, params: list, filters: dict) -> None:
    if filters.get("jcr_quartile"):
        clauses.append("UPPER(jcr_quartile) = ?")
        params.append(filters["jcr_quartile"].upper())
    if filters.get("cas_quartile"):
        clauses.append("cas_quartile LIKE ?")
        params.append(f"%{filters['cas_quartile']}%")
    if filters.get("cas_top") == "yes":
        clauses.append("LOWER(COALESCE(cas_top, '')) IN ('yes', 'y', '1', '是', 'true')")
    elif filters.get("cas_top") == "no":
        clauses.append("(cas_top IS NULL OR cas_top = '' OR LOWER(cas_top) IN ('no', 'n', '0', '否', 'false'))")
    if filters.get("cas_warning") == "yes":
        clauses.append("LOWER(COALESCE(cas_warning, '')) IN ('yes', 'y', '1', '是', 'true', 'warning')")
    elif filters.get("cas_warning") == "no":
        clauses.append("(cas_warning IS NULL OR cas_warning = '' OR LOWER(cas_warning) IN ('no', 'n', '0', '否', 'false'))")
    if filters.get("core_tag"):
        clauses.append("core_tags LIKE ?")
        params.append(f"%{filters['core_tag']}%")
    if filters.get("journal_matched") == "yes":
        clauses.append("journal_matched = 1")
    elif filters.get("journal_matched") == "no":
        clauses.append("(journal_matched IS NULL OR journal_matched = 0)")


def build_papers_query(filters: dict) -> tuple[str, tuple]:
    clauses = []
    params: list = []
    if filters.get("profile") and filters.get("profile") != "__all__":
        clauses.append("profile_id = ?")
        params.append(filters["profile"])
    if filters.get("year"):
        clauses.append("year = ?")
        params.append(filters["year"])
    if filters.get("level"):
        clauses.append("recommendation_level LIKE ?")
        params.append(f"{filters['level']}%")
    if filters.get("rating"):
        clauses.append(f"{RATING_SQL} = ?")
        params.append(int(filters["rating"]))
    if filters.get("abstract") == "complete":
        clauses.append("abstract_is_complete = 1")
    elif filters.get("abstract") == "incomplete":
        clauses.append("abstract_is_complete = 0")
    if filters.get("has_if") == "yes":
        clauses.append("impact_factor IS NOT NULL AND impact_factor != ''")
    elif filters.get("has_if") == "no":
        clauses.append("(impact_factor IS NULL OR impact_factor = '')")
    if filters.get("status"):
        clauses.append("reading_status = ?")
        params.append(filters["status"])
    if filters.get("date") == "today":
        clauses.append("date(COALESCE(created_at, updated_at)) = date('now')")
    if filters.get("if_min"):
        clauses.append("CAST(NULLIF(impact_factor, '') AS REAL) >= ?")
        params.append(float(filters["if_min"]))
    if filters.get("if_max"):
        clauses.append("CAST(NULLIF(impact_factor, '') AS REAL) <= ?")
        params.append(float(filters["if_max"]))
    if filters.get("citation_min"):
        clauses.append("COALESCE(citation_count, 0) >= ?")
        params.append(int(filters["citation_min"]))
    if filters.get("tier") == "full":
        clauses.append("(ingestion_tier IS NULL OR ingestion_tier = 'full')")
    elif filters.get("tier") == "low":
        clauses.append("ingestion_tier = 'low'")
    _apply_journal_rank_filters(clauses, params, filters)
    if filters.get("q"):
        clauses.append("(title LIKE ? OR journal LIKE ? OR abstract_original LIKE ?)")
        keyword = f"%{filters['q']}%"
        params.extend([keyword, keyword, keyword])
    branch = filters.get("branch")
    if branch:
        profile_id = filters.get("profile") or ""
        paper_ids = roadmap_branch_paper_ids(profile_id, branch) if profile_id else []
        if paper_ids:
            placeholders = ",".join("?" * len(paper_ids))
            clauses.append(f"stable_id IN ({placeholders})")
            params.extend(paper_ids)
        else:
            kw = roadmap_branch_keywords(branch)
            if kw:
                token_clauses = []
                for token in [t.strip() for t in kw.split(",") if t.strip()][:4]:
                    token_clauses.append("(title LIKE ? OR abstract_original LIKE ? OR abstract_zh LIKE ?)")
                    like = f"%{token}%"
                    params.extend([like, like, like])
                if token_clauses:
                    clauses.append("(" + " OR ".join(token_clauses) + ")")
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where, tuple(params)


PAPER_FILTER_KEYS = (
    "profile", "year", "level", "rating", "abstract", "has_if", "status", "q", "date",
    "if_min", "if_max", "citation_min", "tier", "jcr_quartile", "cas_quartile",
    "cas_top", "cas_warning", "core_tag", "journal_matched", "branch",
)


def parse_paper_filters_from_request(request: Request, default_profile: str = "") -> dict:
    qp = request.query_params
    profile = qp.get("profile")
    if profile is None or profile == "":
        profile = default_profile
    filters = {key: (qp.get(key) or "") for key in PAPER_FILTER_KEYS}
    filters["profile"] = profile
    return filters


def count_papers_for_filters(filters: dict) -> int:
    where, params = build_papers_query(filters)
    row_data = row(f"SELECT COUNT(*) AS c FROM papers {where}", params)
    return int((row_data or {}).get("c") or 0)


def describe_export_scope(filters: dict, profiles_doc: dict | None = None) -> list[str]:
    profiles_doc = profiles_doc or load_profiles_doc()
    lines = []
    profile_id = filters.get("profile") or ""
    if profile_id == "__all__":
        lines.append("研究方向：全部方向")
    elif profile_id:
        prof = profiles_doc.get("profiles", {}).get(profile_id, {})
        lines.append(f"研究方向：{prof.get('display_name') or prof.get('name') or profile_id}")
    if filters.get("year"):
        lines.append(f"年份：{filters['year']}")
    if filters.get("rating"):
        lines.append(f"星级：{filters['rating']} 星及以上")
    if filters.get("abstract") == "complete":
        lines.append("摘要状态：完整")
    elif filters.get("abstract") == "incomplete":
        lines.append("摘要状态：未补全")
    if filters.get("if_min") or filters.get("if_max"):
        lines.append(f"IF 范围：{filters.get('if_min') or '—'} ~ {filters.get('if_max') or '—'}")
    if filters.get("cas_quartile"):
        lines.append(f"中科院分区：{filters['cas_quartile']}")
    if filters.get("jcr_quartile"):
        lines.append(f"JCR 分区：{filters['jcr_quartile']}")
    if filters.get("q"):
        lines.append(f"关键词：{filters['q']}")
    if filters.get("branch"):
        meta = branch_meta(filters["branch"])
        lines.append(f"研究分支：{meta.get('title') or filters['branch']}")
    return lines


def export_filters_query_string(filters: dict) -> str:
    from urllib.parse import urlencode
    return urlencode({k: v for k, v in filters.items() if v not in ("", None)})


def query_papers(filters: dict, limit: int = 300) -> list[dict]:
    where, params = build_papers_query(filters)
    return rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        {where}
        ORDER BY display_rating DESC, final_score DESC, updated_at DESC
        LIMIT ?
        """,
        (*params, limit),
    )


def papers_ingested_recent(days: int = 7, profile: str = "", limit: int = 20) -> list[dict]:
    clauses = ["date(COALESCE(created_at, updated_at)) >= date('now', ?)"]
    params: list = [f"-{int(days)} days"]
    if profile:
        clauses.append("profile_id = ?")
        params.append(profile)
    where = "WHERE " + " AND ".join(clauses)
    params.append(limit)
    return rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        {where}
        ORDER BY display_rating DESC, final_score DESC, created_at DESC
        LIMIT ?
        """,
        tuple(params),
    )


def daily_report_candidates(profile: str = "", limit: int = 3) -> list[dict]:
    clauses = [
        "(date(COALESCE(created_at, updated_at)) = date('now') OR is_recommended = 1)"
    ]
    params: list = []
    if profile:
        clauses.append("profile_id = ?")
        params.append(profile)
    where = "WHERE " + " AND ".join(clauses)
    params.append(limit)
    candidates = rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        {where}
        ORDER BY is_recommended DESC, display_rating DESC, final_score DESC, updated_at DESC
        LIMIT ?
        """,
        tuple(params),
    )
    if candidates:
        return candidates
    return top_papers_for_profile(profile, limit)


def top_papers_for_profile(profile: str = "", limit: int = 3) -> list[dict]:
    if not profile or profile == "__all__":
        return []
    recommended = rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        WHERE profile_id = ?
          AND (
            COALESCE(is_recommended, 0) = 1
            OR recommendation_level LIKE 'A%'
            OR {RATING_SQL} >= 4
          )
        ORDER BY
            COALESCE(is_recommended, 0) DESC,
            CASE
                WHEN recommendation_level LIKE 'A+%' THEN 1
                WHEN recommendation_level LIKE 'A %' THEN 2
                WHEN recommendation_level LIKE 'B %' THEN 3
                ELSE 4
            END,
            display_rating DESC,
            COALESCE(final_score, relevance_score, 0) DESC,
            updated_at DESC
        LIMIT ?
        """,
        (profile, limit),
    )
    if recommended:
        return recommended
    return rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        WHERE profile_id = ?
          AND COALESCE(is_relevant, 1) = 1
          AND COALESCE(ingestion_tier, 'full') != 'low'
        ORDER BY
            COALESCE(relevance_score, 0) DESC,
            COALESCE(final_score, 0) DESC,
            display_rating DESC,
            updated_at DESC
        LIMIT ?
        """,
        (profile, limit),
    )


def my_notes_path_for_paper(paper: dict) -> Path:
    profile = safe_slug(paper.get("profile_id") or "default")
    year = paper.get("year") or paper.get("publication_year") or "unknown"
    slug = safe_slug(paper.get("title") or paper.get("stable_id"))
    return MY_NOTES_DIR / profile / f"{year}_{slug}.md"


def load_my_notes(paper: dict) -> str:
    path = paper.get("my_notes_path")
    note_path = Path(path) if path else my_notes_path_for_paper(paper)
    if note_path.exists():
        return note_path.read_text(encoding="utf-8")
    return ""


def save_my_notes(paper: dict, content: str) -> Path:
    path = my_notes_path_for_paper(paper)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    execute(
        "UPDATE papers SET my_notes_path = ?, updated_at = datetime('now') WHERE stable_id = ?",
        (str(path), paper["stable_id"]),
    )
    return path


def default_my_notes_template(paper: dict, note: Optional[dict] = None) -> str:
    note = note or {}
    links = []
    for label, key in [
        ("原文", "display_url"),
        ("PDF", "pdf_url"),
        ("DOI", "doi_url"),
        ("arXiv", "arxiv_url"),
    ]:
        url = paper.get(key) or (paper_display_url(paper) if key == "display_url" else "")
        if url:
            links.append(f"- [{label}]({url})")
    link_block = "\n".join(links) if links else "- （暂无链接）"
    return "\n".join([
        f"# {paper.get('title') or '论文笔记'}",
        "",
        f"- 期刊：{paper.get('journal') or ''}",
        f"- 年份：{paper.get('year') or ''}",
        f"- DOI：{paper.get('doi') or ''}",
        "",
        "## 链接",
        link_block,
        "",
        "## 我的阅读目标",
        "",
        "## 关键方法",
        note.get("core_method") or "",
        "",
        "## 启发与可跟进方向",
        note.get("possible_ideas") or note.get("inspiration") or "",
        "",
    ])


def run_doctor_checks() -> list[dict]:
    checks = []
    try:
        init_db()
        with connect() as conn:
            conn.execute("SELECT 1")
        checks.append({"name": "SQLite 数据库", "ok": True, "detail": str(DB_PATH)})
    except (OSError, sqlite3.Error) as exc:
        checks.append({"name": "SQLite 数据库", "ok": False, "detail": str(exc)})

    db_dir = Path(DB_PATH).parent
    db_writable = db_dir.exists() and os.access(str(db_dir), os.W_OK)
    checks.append({
        "name": "数据库目录可写",
        "ok": db_writable,
        "detail": "可写" if db_writable else "请检查目录权限",
    })

    latest = latest_run()
    if latest:
        checks.append({
            "name": "最近运行",
            "ok": True,
            "detail": f"{latest.get('run_time')} · {latest.get('mode')} · 入库 {latest.get('ingested_count') or latest.get('new_papers') or 0} 篇",
        })
    else:
        checks.append({"name": "最近运行", "ok": False, "detail": "尚无运行记录，请从「获取更多文献」启动 daily 模式"})

    checks.append({
        "name": "Kimi / AI 解析",
        "ok": ai_parse_configured(),
        "detail": "已配置" if ai_parse_configured() else "未配置 KIMI_API_KEY / OPENAI / ANTHROPIC",
    })
    checks.append({
        "name": "SerpApi（high_quality / Scholar）",
        "ok": bool(os.environ.get("SERPAPI_API_KEY")),
        "detail": mask_key(os.environ.get("SERPAPI_API_KEY", "")),
    })
    checks.append({
        "name": "微信 Server 酱",
        "ok": bool(os.environ.get("SCT_KEY")),
        "detail": mask_key(os.environ.get("SCT_KEY", "")),
    })
    return checks


def build_kimi_review_body(topic: str, review_type: str, language: str, papers: list[dict]) -> str:
    if not ai_parse_configured():
        return ""
    summaries = []
    for idx, paper in enumerate(papers[:25], 1):
        summaries.append(
            f"{idx}. {paper.get('title')} ({paper.get('year')}, {paper.get('journal')}, "
            f"等级 {paper.get('recommendation_level')})\n"
            f"摘要：{(paper.get('abstract_zh') or paper.get('abstract_original') or '')[:500]}"
        )
    fallback = f"""请基于以下文献列表，用{language}撰写一份「{review_type}」风格的综述正文初稿（资料包级别，非投稿终稿）。
主题：{topic}
要求：包含研究背景、发展脉络、方法分类、代表性工作、挑战与未来方向；引用文献时保留序号。
文献列表：
{chr(10).join(summaries)}
"""
    prompt = render_template(
        "review_writing_prompt",
        {
            "language": language,
            "review_type": review_type,
            "topic": topic,
            "paper_context": "\n".join(summaries),
        },
        fallback,
    )
    return call_llm(prompt) or ""


def build_roadmap_markdown(profile: dict, papers: list[dict], kimi_body: str = "", profile_id: str = "") -> str:
    profile_name = profile.get("display_name") or profile.get("name") or profile.get("id")
    pid = profile_id or profile.get("id") or ""
    lines = [
        f"# {profile_name} 研究方向地图",
        "",
        f"- 生成日期：{today_string()}",
        f"- 研究方向：{profile_name}",
        "",
    ]
    structured = False
    if pid:
        roadmap = build_roadmap_view_model(pid)
        if roadmap.get("stages"):
            structured = True
            lines.extend(["## 时间轴阶段", ""])
            for stage in roadmap["stages"]:
                prog = stage.get("progress") or {}
                lines.append(f"### {stage.get('stage_label') or ''} · {stage.get('title')}")
                lines.append(stage.get("description") or "")
                lines.append(f"- 代表论文 {prog.get('total', 0)} 篇 · 必读 {prog.get('must_read', 0)}/{prog.get('must', 0)}")
                for paper in (stage.get("papers") or [])[:5]:
                    rating = paper.get("display_rating") or paper.get("system_rating") or 3
                    lines.append(f"  - {star_text(rating)} [{paper.get('title')}]({paper_display_url(paper) or '#'})")
                lines.append("")
        if roadmap.get("branches"):
            structured = True
            lines.extend(["## 分支路线", ""])
            for branch in roadmap["branches"]:
                prog = branch.get("progress") or {}
                lines.append(f"### {branch.get('title')}")
                lines.append(f"- 覆盖 {prog.get('total', 0)} 篇 · 已读 {prog.get('read', 0)} 篇 · 覆盖度 {prog.get('pct', 0)}%")
                for child in branch.get("children") or []:
                    cprog = child.get("progress") or {}
                    lines.append(f"  - {child.get('title')}：{cprog.get('total', 0)} 篇，已读 {cprog.get('read', 0)} 篇")
                lines.append("")
        if roadmap.get("routes"):
            structured = True
            lines.extend(["## 精读路线", ""])
            for route in roadmap["routes"]:
                meta = ROADMAP_ROUTE_META.get(route.get("node_key") or "", {})
                prog = route.get("progress") or {}
                lines.append(f"### {route.get('title')}")
                lines.append(route.get("description") or "")
                if meta.get("audience"):
                    lines.append(f"- 适合：{meta['audience']} · 预计 {meta.get('eta_days', '按需')}")
                lines.append(f"- 进度 {prog.get('read', 0)}/{prog.get('total', 0)}")
                for idx, paper in enumerate(route.get("papers") or [], 1):
                    lines.append(f"  {idx}. [{paper.get('title')}]({paper_display_url(paper) or '#'})")
                lines.append("")
    if not structured and papers:
        lines.extend(["## 时间线（按年份）", ""])
        by_year: dict[str, list[dict]] = {}
        for paper in papers:
            by_year.setdefault(str(paper.get("year") or "未知"), []).append(paper)
        for year in sorted(by_year.keys(), reverse=True):
            lines.append(f"### {year}")
            for paper in by_year[year][:8]:
                rating = paper.get("display_rating") or paper.get("system_rating") or 3
                lines.append(
                    f"- {star_text(rating)} [{paper.get('title')}]({paper_display_url(paper) or '#'})"
                )
    if kimi_body:
        lines.extend(["", "## AI 脉络分析", kimi_body])
    else:
        lines.extend([
            "",
            "## 热点关键词",
            ", ".join((profile.get("research_focus") or profile.get("include_keywords") or [])[:8]) or "暂无",
            "",
            "## 追踪建议",
            "保持定期文献获取；对高星论文生成笔记并标记阅读状态。",
        ])
    return "\n".join(lines)


ROADMAP_STAGE_DEFS = [
    {
        "node_key": "stage_foundation",
        "title": "基础概念与相位恢复期",
        "stage_label": "基础期",
        "start_year": 1900,
        "end_year": 2005,
        "description": "理解 ptychography、相位恢复和早期实验/算法概念的来源。",
        "keywords": "ptychography,phase retrieval,coherent diffraction",
    },
    {
        "node_key": "stage_pie_epie",
        "title": "PIE / ePIE 算法发展期",
        "stage_label": "算法期",
        "start_year": 2006,
        "end_year": 2012,
        "description": "迭代相位恢复逐步走向可用，PIE/ePIE 等核心算法形成。",
        "keywords": "PIE,ePIE,iterative,probe recovery",
    },
    {
        "node_key": "stage_electron_4dstem",
        "title": "Electron Ptychography 与 4D-STEM 扩展期",
        "stage_label": "扩展期",
        "start_year": 2013,
        "end_year": 2018,
        "description": "ptychography 与电子显微、4D-STEM 数据采集和重建流程结合。",
        "keywords": "electron ptychography,4D-STEM,STEM",
    },
    {
        "node_key": "stage_advanced_applications",
        "title": "低剂量、多切片与实际应用期",
        "stage_label": "应用期",
        "start_year": 2019,
        "end_year": 2022,
        "description": "多切片、低剂量、厚样品和真实材料应用成为重点。",
        "keywords": "low-dose,multislice,materials,semiconductor",
    },
    {
        "node_key": "stage_ai_frontier",
        "title": "AI、自动化与前沿深化期",
        "stage_label": "前沿期",
        "start_year": 2023,
        "end_year": 2100,
        "description": "AI 辅助、自动化流程、前沿材料应用和高通量分析持续发展。",
        "keywords": "AI,deep learning,automation,foundation model",
    },
]


ROADMAP_BRANCH_DEFS = [
    ("branch_phase_retrieval", "Phase retrieval 基础", "相位恢复、成像反演和算法基础。", "phase retrieval,ptychography,coherent diffraction"),
    ("branch_pie_epie", "PIE / ePIE", "核心迭代算法、探针恢复和重建稳定性。", "PIE,ePIE,iterative,probe"),
    ("branch_electron", "Electron Ptychography", "电子显微 ptychography 的主干方向。", "electron ptychography,electron microscopy,TEM,STEM"),
    ("branch_4dstem", "4D-STEM acquisition", "4D-STEM 数据采集、扫描策略和探测器数据。", "4D-STEM,4D STEM,detector,scan"),
    ("branch_low_dose", "Low-dose reconstruction", "低剂量、噪声鲁棒性和剂量效率。", "low dose,low-dose,dose,noise"),
    ("branch_multislice", "Multislice / thick samples", "多切片、厚样品和多重散射建模。", "multislice,thick,multiple scattering"),
    ("branch_wdd", "WDD / initialization", "WDD、SSB、初始化和直接方法。", "WDD,SSB,single sideband,initialization,direct"),
    ("branch_applications", "Materials / semiconductor applications", "材料、半导体、缺陷和应变等应用。", "semiconductor,materials,strain,defect,device"),
    ("branch_ai", "AI / automation", "深度学习、自动化重建和智能分析。", "AI,deep learning,machine learning,automation,neural"),
]


ROADMAP_ROUTE_DEFS = [
    ("route_intro", "入门理解路线", "适合刚进入方向，建立全局认识与基础概念。", ["stage_foundation", "stage_pie_epie", "stage_electron_4dstem"]),
    ("route_algorithm", "核心算法路线", "适合算法复现、方法改进与初始化策略研究。", ["branch_phase_retrieval", "branch_pie_epie", "branch_wdd", "branch_multislice"]),
    ("route_experiment", "实验与数据路线", "适合理解 4D-STEM 采集、低剂量与实验数据处理。", ["branch_electron", "branch_4dstem", "branch_low_dose"]),
    ("route_frontier", "应用前沿路线", "适合追踪 AI 辅助、材料应用与前沿深化。", ["stage_ai_frontier", "branch_ai", "branch_applications"]),
    ("route_custom", "我的课题定制路线", "结合当前 profile、高星和未读状态生成的下一步阅读路线。", []),
]

ROADMAP_ROUTE_META = {
    "route_intro": {"audience": "新手建立方向框架", "eta_days": "2-3 天", "goal": "建立方向整体框架，理解基础概念与主要应用"},
    "route_algorithm": {"audience": "核心算法复现与方法改进", "eta_days": "3-5 天", "goal": "掌握 PIE/ePIE、初始化与多切片等核心算法"},
    "route_experiment": {"audience": "实验数据与 4D-STEM 场景", "eta_days": "3-4 天", "goal": "理解采集策略、低剂量约束与数据处理流程"},
    "route_frontier": {"audience": "追踪前沿与 AI 辅助方向", "eta_days": "2-3 天", "goal": "把握 AI 辅助重建与材料/半导体应用前沿"},
    "route_custom": {"audience": "结合当前未读高星论文定制", "eta_days": "按需", "goal": "按个人课题优先级推进精读"},
}


def roadmap_year_stage(year) -> str:
    try:
        y = int(year or 0)
    except (TypeError, ValueError):
        y = 0
    for stage in ROADMAP_STAGE_DEFS:
        if int(stage["start_year"]) <= y <= int(stage["end_year"]):
            return stage["node_key"]
    return "stage_ai_frontier"


def roadmap_text_for_paper(paper: dict) -> str:
    keys = ["title", "abstract_original", "abstract_zh", "journal", "core_method", "paper_contribution", "possible_ideas"]
    return " ".join(str(paper.get(k) or "") for k in keys).lower().replace("-", " ")


def roadmap_branch_matches(paper: dict) -> list[str]:
    text = roadmap_text_for_paper(paper)
    matches = []
    for key, _title, _desc, keywords in ROADMAP_BRANCH_DEFS:
        tokens = [kw.strip().lower().replace("-", " ") for kw in keywords.split(",") if kw.strip()]
        if any(token in text for token in tokens):
            matches.append(key)
    if not matches:
        matches.append("branch_phase_retrieval")
    return matches[:3]


def roadmap_branch_keywords(branch_key: str) -> str:
    for key, _title, _desc, keywords in ROADMAP_BRANCH_DEFS:
        if key == branch_key:
            return keywords
    node = row("SELECT keywords FROM roadmap_nodes WHERE node_key = ? LIMIT 1", (branch_key,))
    return (node or {}).get("keywords") or ""


def roadmap_branch_paper_ids(profile_id: str, branch_key: str) -> list[str]:
    if not profile_id or not branch_key:
        return []
    result = rows(
        "SELECT paper_stable_id FROM roadmap_papers WHERE profile = ? AND node_key = ?",
        (profile_id, branch_key),
    )
    return [r["paper_stable_id"] for r in result]


def roadmap_branch_trend(profile_id: str, branch_key: str) -> str:
    from datetime import datetime as dt
    current = dt.now().year
    papers = rows(
        """
        SELECT p.year FROM roadmap_papers rp
        JOIN papers p ON p.stable_id = rp.paper_stable_id
        WHERE rp.profile = ? AND rp.node_key = ? AND p.year IS NOT NULL
        """,
        (profile_id, branch_key),
    )
    if not papers:
        return "偏少"
    recent = sum(1 for p in papers if int(p.get("year") or 0) >= current - 2)
    prior = sum(1 for p in papers if current - 5 <= int(p.get("year") or 0) < current - 2)
    if recent > prior + 1:
        return "增长明显"
    if recent >= prior:
        return "稳定"
    return "偏少"


def roadmap_enrich_branch_node(profile_id: str, node: dict, papers: list[dict]) -> dict:
    progress = roadmap_progress_for_papers(papers)
    high_star = sum(1 for p in papers if int(p.get("display_rating") or p.get("system_rating") or 0) >= 4)
    keywords = node.get("keywords") or roadmap_branch_keywords(node.get("node_key") or "")
    enriched_papers = []
    for paper in papers:
        enriched_papers.append({
            **paper,
            "read_why": paper.get("rating_reason") or paper.get("paper_contribution") or paper.get("note") or "与分支主题高度相关",
        })
    return {
        **node,
        "papers": enriched_papers,
        "progress": progress,
        "coverage_pct": progress["pct"],
        "high_star_count": high_star,
        "trend_label": roadmap_branch_trend(profile_id, node.get("node_key") or ""),
        "keywords_display": keywords,
    }


def roadmap_candidate_papers(profile_id: str, limit: int = 160) -> list[dict]:
    return rows(
        f"""
        SELECT p.*, rn.core_method, rn.paper_contribution, rn.possible_ideas,
               {RATING_SQL} AS display_rating
        FROM papers p
        LEFT JOIN reading_notes rn ON rn.paper_id = p.stable_id
        WHERE p.profile_id = ?
        ORDER BY
            p.is_milestone DESC,
            p.included_in_review DESC,
            display_rating DESC,
            COALESCE(p.citation_count, 0) DESC,
            COALESCE(p.final_score, p.relevance_score, 0) DESC,
            COALESCE(p.publication_year, p.year, 0) DESC
        LIMIT ?
        """,
        (profile_id, limit),
    )


def roadmap_importance_for_paper(paper: dict, rank: int) -> tuple[str, int]:
    rating = int(paper.get("display_rating") or paper.get("system_rating") or 1)
    if paper.get("is_milestone") or rank <= 2 or rating >= 5:
        return "milestone", 1
    if paper.get("included_in_review") or rating >= 4:
        return "representative", 0
    return "recommended", 0


def save_roadmap_node(conn, profile_id: str, node: dict) -> None:
    now = today_string()
    conn.execute(
        """
        INSERT INTO roadmap_nodes (
            profile, node_key, title, description, node_type, stage_label,
            start_year, end_year, parent_node_key, importance_level, reading_order,
            keywords, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(profile, node_key) DO UPDATE SET
            title=excluded.title,
            description=excluded.description,
            node_type=excluded.node_type,
            stage_label=excluded.stage_label,
            start_year=excluded.start_year,
            end_year=excluded.end_year,
            parent_node_key=excluded.parent_node_key,
            importance_level=excluded.importance_level,
            reading_order=excluded.reading_order,
            keywords=excluded.keywords,
            updated_at=excluded.updated_at
        """,
        (
            profile_id,
            node["node_key"],
            node["title"],
            node.get("description", ""),
            node["node_type"],
            node.get("stage_label", ""),
            node.get("start_year"),
            node.get("end_year"),
            node.get("parent_node_key", ""),
            node.get("importance_level", "medium"),
            int(node.get("reading_order") or 0),
            node.get("keywords", ""),
            now,
            now,
        ),
    )


def save_roadmap_edge(conn, profile_id: str, from_key: str, to_key: str, edge_type: str, description: str = "") -> None:
    conn.execute(
        """
        INSERT INTO roadmap_edges (profile, from_node_key, to_node_key, edge_type, description)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(profile, from_node_key, to_node_key, edge_type) DO UPDATE SET description=excluded.description
        """,
        (profile_id, from_key, to_key, edge_type, description),
    )


def save_roadmap_paper(conn, profile_id: str, node_key: str, paper_id: str, role: str, must_read: int, rank: int, note: str = "") -> None:
    conn.execute(
        """
        INSERT INTO roadmap_papers (profile, node_key, paper_stable_id, paper_role, is_must_read, reading_rank, note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(profile, node_key, paper_stable_id) DO UPDATE SET
            paper_role=excluded.paper_role,
            is_must_read=excluded.is_must_read,
            reading_rank=excluded.reading_rank,
            note=excluded.note
        """,
        (profile_id, node_key, paper_id, role, must_read, rank, note),
    )


def generate_roadmap_map(profile_id: str) -> dict:
    init_db()
    doc = load_profiles_doc()
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    display_name = profile_obj.get("display_name") or profile_obj.get("name") or profile_id
    focus_hint = ", ".join((profile_obj.get("research_focus") or profile_obj.get("include_keywords") or [])[:4])
    candidates = roadmap_candidate_papers(profile_id)
    with connect() as conn:
        conn.execute("DELETE FROM roadmap_edges WHERE profile = ?", (profile_id,))
        conn.execute("DELETE FROM roadmap_papers WHERE profile = ?", (profile_id,))
        conn.execute("DELETE FROM roadmap_nodes WHERE profile = ?", (profile_id,))
        root = {
            "node_key": "root",
            "title": f"{display_name} 研究方向",
            "description": f"当前 profile 的研究主干（{focus_hint or profile_id}），用于连接阶段、分支和精读路线。",
            "node_type": "root",
            "importance_level": "high",
            "reading_order": 0,
            "keywords": focus_hint or "research",
        }
        save_roadmap_node(conn, profile_id, root)
        for idx, stage in enumerate(ROADMAP_STAGE_DEFS, 1):
            save_roadmap_node(conn, profile_id, {**stage, "node_type": "stage", "importance_level": "high", "reading_order": idx})
            save_roadmap_edge(conn, profile_id, "root", stage["node_key"], "evolves_to", "方向发展阶段")
            if idx > 1:
                save_roadmap_edge(conn, profile_id, ROADMAP_STAGE_DEFS[idx - 2]["node_key"], stage["node_key"], "evolves_to", "阶段演进")
        for idx, (key, title, desc, keywords) in enumerate(ROADMAP_BRANCH_DEFS, 1):
            parent = "branch_electron" if key in {"branch_4dstem", "branch_low_dose", "branch_multislice", "branch_wdd"} else "root"
            save_roadmap_node(
                conn,
                profile_id,
                {
                    "node_key": key,
                    "title": title,
                    "description": desc,
                    "node_type": "branch",
                    "parent_node_key": parent,
                    "importance_level": "high" if key in {"branch_electron", "branch_pie_epie"} else "medium",
                    "reading_order": idx,
                    "keywords": keywords,
                },
            )
            save_roadmap_edge(conn, profile_id, parent, key, "branches_to", "研究分支")
        for idx, (key, title, desc, source_nodes) in enumerate(ROADMAP_ROUTE_DEFS, 1):
            save_roadmap_node(
                conn,
                profile_id,
                {
                    "node_key": key,
                    "title": title,
                    "description": desc,
                    "node_type": "route",
                    "parent_node_key": "root",
                    "importance_level": "high" if key == "route_custom" else "medium",
                    "reading_order": idx,
                    "keywords": ",".join(source_nodes),
                },
            )
            save_roadmap_edge(conn, profile_id, "root", key, "reading_path", desc)
        grouped: dict[str, list[dict]] = {}
        branch_grouped: dict[str, list[dict]] = {}
        for paper in candidates:
            grouped.setdefault(roadmap_year_stage(paper.get("year") or paper.get("publication_year")), []).append(paper)
            for branch in roadmap_branch_matches(paper):
                branch_grouped.setdefault(branch, []).append(paper)
        for node_key, papers in grouped.items():
            for rank, paper in enumerate(papers[:6], 1):
                role, must = roadmap_importance_for_paper(paper, rank)
                save_roadmap_paper(conn, profile_id, node_key, paper["stable_id"], role, must, rank, "阶段代表论文")
        for node_key, papers in branch_grouped.items():
            for rank, paper in enumerate(papers[:8], 1):
                role, must = roadmap_importance_for_paper(paper, rank)
                save_roadmap_paper(conn, profile_id, node_key, paper["stable_id"], role, must, rank, "分支代表论文")
        used_route_papers: set[str] = set()
        route_sources = {key: source_nodes for key, _title, _desc, source_nodes in ROADMAP_ROUTE_DEFS}
        for route_key, source_nodes in route_sources.items():
            route_pool: list[dict] = []
            if route_key == "route_custom":
                route_pool = [p for p in candidates if (p.get("reading_status") or "unread") in {"", "unread", "todo", None}][:10]
            else:
                seen_ids = set()
                for node_key in source_nodes:
                    for paper in (grouped.get(node_key) or branch_grouped.get(node_key) or []):
                        if paper["stable_id"] not in seen_ids:
                            route_pool.append(paper)
                            seen_ids.add(paper["stable_id"])
            for rank, paper in enumerate(route_pool[:8], 1):
                used_route_papers.add(paper["stable_id"])
                role, must = roadmap_importance_for_paper(paper, rank)
                save_roadmap_paper(conn, profile_id, route_key, paper["stable_id"], role, 1 if rank <= 5 or must else 0, rank, "路线推荐论文")
        conn.commit()
    return {"candidates": len(candidates), "routes": len(ROADMAP_ROUTE_DEFS), "route_papers": len(used_route_papers)}


def roadmap_nodes(profile_id: str, node_type: str = "") -> list[dict]:
    sql = "SELECT * FROM roadmap_nodes WHERE profile = ?"
    params: list = [profile_id]
    if node_type:
        sql += " AND node_type = ?"
        params.append(node_type)
    sql += " ORDER BY reading_order, start_year, id"
    return rows(sql, tuple(params))


def roadmap_papers_for_node(profile_id: str, node_key: str, limit: int = 20) -> list[dict]:
    return rows(
        f"""
        SELECT rp.*, p.*, {RATING_SQL} AS display_rating
        FROM roadmap_papers rp
        JOIN papers p ON p.stable_id = rp.paper_stable_id
        WHERE rp.profile = ? AND rp.node_key = ?
        ORDER BY rp.reading_rank, rp.is_must_read DESC, display_rating DESC
        LIMIT ?
        """,
        (profile_id, node_key, limit),
    )


def roadmap_progress_for_papers(papers: list[dict]) -> dict:
    total = len(papers)
    read = sum(1 for p in papers if p.get("reading_status") == "read")
    reading = sum(1 for p in papers if p.get("reading_status") == "reading")
    must = sum(1 for p in papers if p.get("is_must_read"))
    must_read = sum(1 for p in papers if p.get("is_must_read") and p.get("reading_status") == "read")
    return {
        "total": total,
        "read": read,
        "reading": reading,
        "must": must,
        "must_read": must_read,
        "pct": int(100 * read / total) if total else 0,
    }


def build_roadmap_view_model(profile_id: str, selected_node: str = "") -> dict:
    stages = []
    for node in roadmap_nodes(profile_id, "stage"):
        papers = roadmap_papers_for_node(profile_id, node["node_key"], 20)
        enriched = []
        for paper in papers:
            enriched.append({
                **paper,
                "read_why": paper.get("rating_reason") or paper.get("paper_contribution") or paper.get("note") or "阶段代表论文",
            })
        stages.append({**node, "papers": enriched, "progress": roadmap_progress_for_papers(papers)})
    branch_nodes = roadmap_nodes(profile_id, "branch")
    branch_by_key = {node["node_key"]: node for node in branch_nodes}
    branches = []
    flat_branches = []
    for node in branch_nodes:
        parent = (node.get("parent_node_key") or "").strip()
        if parent and parent not in ("", "root"):
            continue
        papers = roadmap_papers_for_node(profile_id, node["node_key"], 8)
        children = []
        for child_key, child in branch_by_key.items():
            if (child.get("parent_node_key") or "") != node["node_key"]:
                continue
            child_papers = roadmap_papers_for_node(profile_id, child_key, 8)
            child_enriched = roadmap_enrich_branch_node(profile_id, child, child_papers)
            children.append(child_enriched)
            flat_branches.append(child_enriched)
        branch_enriched = roadmap_enrich_branch_node(profile_id, node, papers)
        branch_enriched["children"] = children
        branches.append(branch_enriched)
        flat_branches.append(branch_enriched)
    routes = []
    for node in roadmap_nodes(profile_id, "route"):
        papers = roadmap_papers_for_node(profile_id, node["node_key"], 12)
        meta = ROADMAP_ROUTE_META.get(node["node_key"], {})
        enriched = []
        for paper in papers:
            enriched.append({
                **paper,
                "read_why": paper.get("rating_reason") or paper.get("paper_contribution") or paper.get("note") or meta.get("goal") or "路线推荐论文",
            })
        routes.append({
            **node,
            "papers": enriched,
            "progress": roadmap_progress_for_papers(papers),
            "audience": meta.get("audience", ""),
            "eta_days": meta.get("eta_days", ""),
            "goal": meta.get("goal", node.get("description") or ""),
        })
    if selected_node:
        selected = next((item for item in flat_branches if item["node_key"] == selected_node), flat_branches[0] if flat_branches else {})
    else:
        selected = flat_branches[0] if flat_branches else {}
    return {"stages": stages, "branches": branches, "flat_branches": flat_branches, "routes": routes, "selected_node": selected}


def build_roadmap_lite(profile_id: str) -> dict:
    if not profile_id:
        return {"paper_count": 0, "has_map": False, "stage_count": 0, "must_read_total": 0, "must_read_done": 0}
    paper_count = row("SELECT COUNT(*) AS c FROM papers WHERE profile_id = ?", (profile_id,)) or {}
    stage_count = row(
        "SELECT COUNT(*) AS c FROM roadmap_nodes WHERE profile = ? AND node_type = 'stage'",
        (profile_id,),
    ) or {}
    node_count = row("SELECT COUNT(*) AS c FROM roadmap_nodes WHERE profile = ?", (profile_id,)) or {}
    must_stats = row(
        """
        SELECT
            SUM(CASE WHEN rp.is_must_read = 1 THEN 1 ELSE 0 END) AS must_total,
            SUM(CASE WHEN rp.is_must_read = 1 AND p.reading_status = 'read' THEN 1 ELSE 0 END) AS must_done
        FROM roadmap_papers rp
        JOIN papers p ON p.stable_id = rp.paper_stable_id
        WHERE rp.profile = ?
        """,
        (profile_id,),
    ) or {}
    return {
        "paper_count": int(paper_count.get("c") or 0),
        "has_map": int(node_count.get("c") or 0) > 1,
        "stage_count": int(stage_count.get("c") or 0),
        "must_read_total": int(must_stats.get("must_total") or 0),
        "must_read_done": int(must_stats.get("must_done") or 0),
    }


def list_roadmap_reports(profile_id: str, limit: int = 5) -> list[str]:
    if not ROADMAP_REPORTS_DIR.exists() or not profile_id:
        return []
    slug = safe_slug(profile_id)
    names = sorted(
        (p.name for p in ROADMAP_REPORTS_DIR.glob(f"*_{slug}_*.md") if p.is_file()),
        reverse=True,
    )
    return names[:limit]


def build_roadmap_sidebar_summary(profile_id: str, roadmap: dict) -> dict:
    stats = row(
        """
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN abstract_is_complete = 1 THEN 1 ELSE 0 END) AS abstract_completed,
            SUM(CASE WHEN is_milestone = 1 THEN 1 ELSE 0 END) AS milestone_count,
            SUM(CASE WHEN """ + RATING_SQL + """ >= 4 THEN 1 ELSE 0 END) AS high_star
        FROM papers WHERE profile_id = ?
        """,
        (profile_id,),
    ) or {}
    flat = roadmap.get("flat_branches") or []
    trending = sorted(
        flat,
        key=lambda b: (1 if b.get("trend_label") == "增长明显" else 0, b.get("progress", {}).get("total") or 0),
        reverse=True,
    )[:3]
    suggestions = []
    if int(stats.get("total") or 0) < 20:
        suggestions.append("文献积累不足，建议先获取更多文献再生成路径。")
    elif trending:
        names = "、".join(b.get("title") or "" for b in trending[:2])
        suggestions.append(f"建议优先追踪：{names}。")
    else:
        suggestions.append("可进入想法打磨室，基于分支文献验证研究想法。")
    return {
        "total": stats.get("total") or 0,
        "abstract_completed": stats.get("abstract_completed") or 0,
        "high_star": stats.get("high_star") or 0,
        "milestone_count": stats.get("milestone_count") or 0,
        "trending_branches": trending,
        "suggestions": suggestions,
    }


def roadmap_page_context(request: Request, profile: str = "", view: str = "timeline", node: str = "", stage: str = "") -> dict:
    context = common_context(request)
    profile_id = profile or context.get("current_profile_id", "")
    doc = load_profiles_doc()
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    lite = build_roadmap_lite(profile_id)
    roadmap = build_roadmap_view_model(profile_id, node) if lite["paper_count"] > 0 else {
        "stages": [], "branches": [], "flat_branches": [], "routes": [], "selected_node": {},
    }
    candidates = roadmap_candidate_papers(profile_id, 20) if lite["paper_count"] > 0 else []
    year_row = row(
        "SELECT MIN(year) AS y_min, MAX(year) AS y_max FROM papers WHERE profile_id = ? AND year IS NOT NULL",
        (profile_id,),
    ) if profile_id else {}
    map_updated = row(
        "SELECT MAX(updated_at) AS ts FROM roadmap_nodes WHERE profile = ?",
        (profile_id,),
    ) if profile_id else {}
    normalized_view = view if view in ("timeline", "branches", "routes") else "timeline"
    if view == "tree":
        normalized_view = "branches"
    selected_stage = None
    if stage:
        selected_stage = next((s for s in roadmap.get("stages") or [] if s.get("node_key") == stage), None)
    sidebar = build_roadmap_sidebar_summary(profile_id, roadmap) if profile_id else {}
    return {
        **context,
        "selected_profile": profile_id,
        "selected_profile_name": profile_obj.get("display_name") or profile_obj.get("name") or profile_id,
        "view": normalized_view,
        "roadmap": roadmap,
        "roadmap_lite": lite,
        "sidebar_summary": sidebar,
        "profiles": rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id"),
        "papers": candidates,
        "paper_count": lite["paper_count"],
        "year_min": (year_row or {}).get("y_min") or "—",
        "year_max": (year_row or {}).get("y_max") or "—",
        "milestone_count": sidebar.get("milestone_count") or lite.get("must_read_total") or 0,
        "map_updated_at": (map_updated or {}).get("ts") or "—",
        "selected_stage": selected_stage,
        "reports": list_roadmap_reports(profile_id),
        "taxonomy_note": "本页面仅使用当前研究方向的文献，不会混入其他方向。阶段与分支结构基于 Ptychography 通用模板。",
    }


def recommended_papers(limit: int = 3) -> list[dict]:
    return rows(
        """
        SELECT * FROM papers
        ORDER BY
            CASE
                WHEN recommendation_level LIKE 'A+%' THEN 1
                WHEN recommendation_level LIKE 'A %' THEN 2
                WHEN recommendation_level LIKE 'B %' THEN 3
                ELSE 4
            END,
            final_score DESC,
            relevance_score DESC,
            updated_at DESC
        LIMIT ?
        """,
        (limit,),
    )


def all_papers_for_export() -> list[dict]:
    return rows("SELECT * FROM papers ORDER BY year DESC, final_score DESC, updated_at DESC")


def paper_line(paper: dict) -> str:
    return (
        f"{paper.get('title', '无标题')}（{paper.get('recommendation_level') or '未评级'}，"
        f"综合分 {paper.get('final_score') or paper.get('relevance_score') or 0}/30）"
    )


def build_daily_report_markdown(context: dict, papers: list[dict]) -> str:
    profile = context.get("current_profile", {})
    numbers = context.get("dashboard_numbers", {})
    date = today_string()
    lines = [
        f"# {date} 文献雷达日报",
        "",
        f"方向：{profile.get('display_name') or profile.get('name') or context.get('current_profile_id')}",
        "",
        "## 今日概览",
        f"- 候选论文：{numbers.get('total_found', 0)} 篇",
        f"- 完整摘要：{numbers.get('abstract_completed', 0)} 篇",
        f"- IF 匹配：{numbers.get('if_matched', 0)} 篇",
        f"- 推荐精读：{numbers.get('recommended', 0)} 篇",
        "",
        "## 今日最值得读 TOP 3",
        "",
    ]
    if not papers:
        lines.extend([
            "当前研究方向今天暂无可推荐论文。",
            "",
            "## 今日阅读建议",
            "先点击“获取文献”运行当前研究方向；如果仍然没有结果，请检查关键词、must_have_any 或年份范围是否过窄。",
            "",
        ])
    else:
        for idx, paper in enumerate(papers, 1):
            lines.extend([
                f"### {idx}. {paper_line(paper)}",
                f"- 期刊：{paper.get('journal') or '未知'}",
                f"- IF：{paper.get('impact_factor') or '未匹配'}",
                f"- 引用数：{paper.get('citation_count') or 0}",
                f"- 摘要来源：{paper.get('abstract_source') or 'missing'}",
                f"- 链接：{paper_display_url(paper)}",
                "",
                "一句话总结：",
                (paper.get("abstract_zh") or paper.get("abstract_original") or "暂无摘要")[:240],
                "",
            ])
        lines.extend([
            "## 今日阅读建议",
            "优先精读第 1 篇；其余 A/A+ 论文加入待读队列，并在周报中复盘。",
            "",
        ])
    return "\n".join(lines)


def build_seminar_markdown(context: dict, papers: list[dict]) -> str:
    profile = context.get("current_profile", {})
    lines = [
        f"# {week_string()} 组会汇报提纲",
        "",
        f"方向：{profile.get('display_name') or profile.get('name') or context.get('current_profile_id')}",
        "",
        "## 本周重点论文",
    ]
    for idx, paper in enumerate(papers, 1):
        rating = paper.get("display_rating") or paper.get("system_rating") or 3
        lines.append(f"{idx}. {paper.get('title')} ({paper.get('year')}) — {star_text(rating)}")
        lines.append(f"   - 链接：{paper_display_url(paper)}")
        lines.append(f"   - 一句话：{(paper.get('abstract_zh') or paper.get('abstract_original') or '')[:160]}")
    lines.extend(["", "## 讨论问题", "1. 方法对比与适用边界", "2. 与当前课题的结合点", ""])
    return "\n".join(lines)


def build_intro_draft_markdown(context: dict, papers: list[dict]) -> str:
    profile = context.get("current_profile", {})
    topic = profile.get("display_name") or profile.get("name") or "研究方向"
    lines = [
        f"# {topic} — 论文引言草稿（初稿）",
        "",
        "> AI 生成初稿，请核对引用与事实。",
        "",
        "## 1. 研究背景",
        "（根据下列高星文献摘要归纳领域问题与动机。）",
        "",
    ]
    kimi = build_kimi_review_body(f"{topic} 引言背景", "论文引言", "中文", papers)
    if kimi:
        lines.append(kimi)
    else:
        for paper in papers[:5]:
            lines.append(f"- {paper.get('title')}: {(paper.get('abstract_zh') or paper.get('abstract_original') or '')[:200]}")
    lines.extend(["", "## 2. 本文工作定位", "（待填写）", "", "## 参考文献候选"])
    for idx, paper in enumerate(papers, 1):
        lines.append(f"{idx}. {paper.get('title')}. {paper.get('journal')}, {paper.get('year')}. {paper_display_url(paper)}")
    return "\n".join(lines)


def build_weekly_report_markdown(context: dict, papers: list[dict], keywords: list[str]) -> str:
    profile = context.get("current_profile", {})
    numbers = context.get("dashboard_numbers", {})
    lines = [
        f"# {week_string()} 文献趋势周报",
        "",
        f"方向：{profile.get('display_name') or profile.get('name') or context.get('current_profile_id')}",
        "",
        "## 本周数据",
        f"- 本周新增：{numbers.get('weekly_new', 0)} 篇",
        f"- 推荐精读：{numbers.get('recommended', 0)} 篇",
        f"- 完整摘要：{numbers.get('abstract_completed', 0)} 篇",
        f"- IF 匹配：{numbers.get('if_matched', 0)} 篇",
        "",
        "## 热门关键词",
        ", ".join(keywords) if keywords else "暂无关键词统计",
        "",
        "## 本周 TOP 5",
    ]
    for idx, paper in enumerate(papers, 1):
        lines.append(f"{idx}. {paper_line(paper)}")
    lines.extend([
        "",
        "## 本周趋势总结",
        "本周文献建议重点关注相位恢复、4D-STEM 重建质量、低剂量成像和半导体表征等方向。",
        "",
        "## 下周阅读建议",
        "优先精读 TOP 1-2，将方法类论文整理进算法专题，将应用类论文进入综述候选。",
        "",
    ])
    return "\n".join(lines)


def write_markdown(path: Path, content: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return path


def export_papers_excel(papers: list[dict]) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORTS_DIR / f"research_radar_export_{today_string()}.xlsx"
    headers = [
        "标题", "年份", "期刊", "DOI", "原文链接", "PDF", "arXiv", "摘要来源", "摘要是否完整",
        "JCR IF", "JCR分区", "中科院分区", "中科院TOP", "预警状态", "核心标签",
        "期刊指标来源", "匹配方式", "引用数", "相关性评分",
        "综合分", "系统星级", "推荐等级", "阅读状态", "是否推送微信",
    ]
    if Workbook is None:
        csv_path = path.with_suffix(".csv")
        lines = [",".join(headers)]
        for paper in papers:
            display_rating = paper.get("display_rating") or paper.get("system_rating") or score_to_rating(paper.get("final_score"))
            warning = paper.get("cas_warning", "")
            warning_display = "预警" if str(warning).lower() in {"yes", "y", "1", "是", "warning", "true"} else "无"
            values = [
                paper.get("title", ""), paper.get("year", ""), paper.get("journal", ""),
                paper.get("doi", ""), paper_display_url(paper), paper.get("pdf_url", ""),
                paper.get("arxiv_url", ""), paper.get("abstract_source", ""),
                paper.get("abstract_is_complete", ""),
                paper.get("jcr_impact_factor") or paper.get("impact_factor", ""),
                paper.get("jcr_quartile", ""), paper.get("cas_quartile", ""),
                paper.get("cas_top", ""), warning_display, paper.get("core_tags", ""),
                paper.get("journal_rank_source", ""), paper.get("journal_match_method", ""),
                paper.get("citation_count", ""), paper.get("relevance_score", ""),
                paper.get("final_score", ""), display_rating, paper.get("recommendation_level", ""),
                paper.get("reading_status", ""), paper.get("pushed_to_wechat", ""),
            ]
            lines.append(",".join('"' + str(v).replace('"', '""') + '"' for v in values))
        csv_path.write_text("\n".join(lines), encoding="utf-8-sig")
        return csv_path
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Radar"
    ws.append(headers)
    for paper in papers:
        display_rating = paper.get("display_rating") or paper.get("system_rating") or score_to_rating(paper.get("final_score"))
        warning = paper.get("cas_warning", "")
        warning_display = "预警" if str(warning).lower() in {"yes", "y", "1", "是", "warning", "true"} else "无"
        ws.append([
            paper.get("title", ""), paper.get("year", ""), paper.get("journal", ""),
            paper.get("doi", ""), paper_display_url(paper), paper.get("pdf_url", ""),
            paper.get("arxiv_url", ""), paper.get("abstract_source", ""),
            "是" if paper.get("abstract_is_complete") else "否",
            paper.get("jcr_impact_factor") or paper.get("impact_factor", ""),
            paper.get("jcr_quartile", ""), paper.get("cas_quartile", ""),
            paper.get("cas_top", ""), warning_display, paper.get("core_tags", ""),
            paper.get("journal_rank_source", ""), paper.get("journal_match_method", ""),
            paper.get("citation_count", ""), paper.get("relevance_score", ""),
            paper.get("final_score", ""), display_rating, paper.get("recommendation_level", ""),
            paper.get("reading_status", ""), "是" if paper.get("pushed_to_wechat") else "否",
        ])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)
    wb.save(path)
    return path


LIBRARY_OUTPUT_ARTIFACTS = (
    ("全量库 Excel", "Ptychography_论文全量库.xlsx"),
    ("解析汇总", "Ptychography论文解析汇总.md"),
)


def _output_download_href(path: Path) -> str:
    try:
        rel = path.resolve().relative_to(OUTPUT_DIR.resolve())
    except ValueError:
        return ""
    return f"/output/download?rel={quote(rel.as_posix())}"


def _report_file_entry(label: str, path: Path) -> Optional[dict]:
    if not path.is_file():
        return None
    suffix = path.suffix.lower()
    if suffix not in {".md", ".xlsx", ".csv"}:
        return None
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0
    return {
        "type": label,
        "name": path.name,
        "path": str(path),
        "mtime": mtime,
        "mtime_display": datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M") if mtime else "",
        "ext": suffix,
        "download_href": _output_download_href(path),
    }


def recent_report_files(limit: int = 6) -> list[dict]:
    report_dirs = [
        ("日报", OUTPUT_DIR / "daily_reports"),
        ("周报", WEEKLY_REPORTS_DIR),
        ("年度", ANNUAL_REPORTS_DIR),
        ("综述", REVIEW_REPORTS_DIR),
        ("方向地图", ROADMAP_REPORTS_DIR),
        ("导出", EXPORTS_DIR),
    ]
    files: list[dict] = []
    seen_paths: set[str] = set()

    def add_entry(entry: Optional[dict]) -> None:
        if not entry:
            return
        key = os.path.normcase(entry["path"])
        if key in seen_paths:
            return
        seen_paths.add(key)
        files.append(entry)

    for label, filename in LIBRARY_OUTPUT_ARTIFACTS:
        add_entry(_report_file_entry(label, OUTPUT_DIR / filename))

    for label, folder in report_dirs:
        if not folder.exists():
            continue
        for path in folder.glob("*"):
            add_entry(_report_file_entry(label, path))

    try:
        init_db()
        with connect() as conn:
            for row in conn.execute(
                """
                SELECT report_path FROM runs
                WHERE report_path IS NOT NULL AND TRIM(report_path) != ''
                ORDER BY run_time DESC, id DESC
                LIMIT 8
                """
            ).fetchall():
                add_entry(_report_file_entry("运行报告", Path(row[0])))
    except (OSError, sqlite3.Error):
        pass

    files.sort(key=lambda item: item["mtime"], reverse=True)
    return files[:limit]


@app.get("/output/download")
def output_download(rel: str = ""):
    rel_norm = (rel or "").replace("\\", "/").strip().lstrip("/")
    if not rel_norm or ".." in rel_norm.split("/"):
        raise HTTPException(status_code=404, detail="文件不存在")
    path = (OUTPUT_DIR / rel_norm).resolve()
    if not str(path).startswith(str(OUTPUT_DIR.resolve())) or not path.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(path, filename=path.name)


def library_year_overview(limit: int = 4, profile: str = "") -> list[dict]:
    clauses = ["COALESCE(publication_year, year) IS NOT NULL"]
    params: list = []
    if profile and profile != "__all__":
        clauses.append("profile_id = ?")
        params.append(profile)
    params.append(limit)
    return rows(
        f"""
        SELECT COALESCE(publication_year, year) AS paper_year, COUNT(*) AS count
        FROM papers
        WHERE {' AND '.join(clauses)}
        GROUP BY COALESCE(publication_year, year)
        ORDER BY paper_year DESC
        LIMIT ?
        """,
        tuple(params),
    )


def annual_filter_query(year: int, filters: dict) -> tuple[str, tuple]:
    clauses = ["COALESCE(publication_year, year) = ?"]
    params: list = [year]
    if filters.get("profile"):
        clauses.append("profile_id = ?")
        params.append(filters["profile"])
    if filters.get("rating"):
        clauses.append(f"{RATING_SQL} = ?")
        params.append(int(filters["rating"]))
    if filters.get("abstract") == "complete":
        clauses.append("abstract_is_complete = 1")
    elif filters.get("abstract") == "incomplete":
        clauses.append("abstract_is_complete = 0")
    if filters.get("status"):
        clauses.append("reading_status = ?")
        params.append(filters["status"])
    if filters.get("favorite") == "yes":
        clauses.append("(favorite = 1 OR is_favorite = 1)")
    if filters.get("included") == "yes":
        clauses.append("included_in_review = 1")
    if filters.get("if_min"):
        clauses.append("CAST(NULLIF(impact_factor, '') AS REAL) >= ?")
        params.append(float(filters["if_min"]))
    if filters.get("citation_min"):
        clauses.append("COALESCE(citation_count, 0) >= ?")
        params.append(int(filters["citation_min"]))
    if filters.get("source"):
        clauses.append("source LIKE ?")
        params.append(f"%{filters['source']}%")
    if filters.get("q"):
        clauses.append("(title LIKE ? OR journal LIKE ? OR abstract_original LIKE ? OR abstract_zh LIKE ?)")
        keyword = f"%{filters['q']}%"
        params.extend([keyword, keyword, keyword, keyword])
    _apply_journal_rank_filters(clauses, params, filters)
    return "WHERE " + " AND ".join(clauses), tuple(params)


def annual_papers(year: int, filters: dict, limit: int = 500) -> list[dict]:
    where, params = annual_filter_query(year, filters)
    return rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        {where}
        ORDER BY display_rating DESC, COALESCE(final_score, relevance_score, 0) DESC, citation_count DESC, updated_at DESC
        LIMIT ?
        """,
        (*params, limit),
    )


def annual_stats(year: int, filters: dict) -> dict:
    where, params = annual_filter_query(year, filters)
    return row(
        f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN abstract_is_complete = 1 THEN 1 ELSE 0 END) AS abstract_completed,
            SUM(CASE WHEN impact_factor IS NOT NULL AND impact_factor != '' THEN 1 ELSE 0 END) AS if_matched,
            SUM(CASE WHEN {RATING_SQL} = 5 THEN 1 ELSE 0 END) AS five_star,
            SUM(CASE WHEN {RATING_SQL} = 4 THEN 1 ELSE 0 END) AS four_star,
            SUM(CASE WHEN {RATING_SQL} = 3 THEN 1 ELSE 0 END) AS three_star,
            SUM(CASE WHEN reading_status = 'read' THEN 1 ELSE 0 END) AS read_count,
            SUM(CASE WHEN favorite = 1 OR is_favorite = 1 THEN 1 ELSE 0 END) AS favorite_count,
            SUM(CASE WHEN included_in_review = 1 THEN 1 ELSE 0 END) AS review_count
        FROM papers
        {where}
        """,
        params,
    ) or {}


def available_years() -> list[int]:
    year_rows = rows(
        """
        SELECT DISTINCT COALESCE(publication_year, year) AS paper_year
        FROM papers
        WHERE COALESCE(publication_year, year) IS NOT NULL
        ORDER BY paper_year DESC
        """
    )
    return [int(item["paper_year"]) for item in year_rows if item.get("paper_year")] or [current_year()]


def build_annual_markdown(year: int, profile_label: str, stats: dict, papers: list[dict]) -> str:
    lines = [
        f"# {year} 年 {profile_label} 年度文献库",
        "",
        "## 年度概览",
        f"- 总论文数：{stats.get('total') or 0}",
        f"- 五星论文：{stats.get('five_star') or 0}",
        f"- 四星论文：{stats.get('four_star') or 0}",
        f"- 三星论文：{stats.get('three_star') or 0}",
        f"- 完整摘要：{stats.get('abstract_completed') or 0}",
        f"- IF 匹配：{stats.get('if_matched') or 0}",
        "",
        "## 五星论文",
    ]
    for rating in [5, 4, 3, 2, 1]:
        if rating != 5:
            lines.extend(["", f"## {star_text(rating)} 论文"])
        matches = [paper for paper in papers if int(paper.get("display_rating") or 1) == rating]
        if not matches:
            lines.append("暂无")
        for paper in matches:
            lines.extend([
                f"### {paper.get('title') or 'Untitled'}",
                f"- 期刊：{paper.get('journal') or '未知'}",
                f"- IF：{paper.get('impact_factor') or '未匹配'}",
                f"- 引用数：{paper.get('citation_count') or 0}",
                f"- 综合分：{paper.get('final_score') or paper.get('relevance_score') or 0}",
                f"- 系统建议：{paper.get('rating_reason') or rating_label(rating)}",
                f"- 链接：{paper.get('url') or ''}",
                "",
            ])
    lines.extend([
        "## 按方法分类",
        "- ePIE / iterative phase retrieval",
        "- WDD / SSB / direct methods",
        "- Multislice ptychography",
        "- Deep Learning assisted reconstruction",
        "",
        "## 年度趋势总结",
        "本节为年度文献库自动生成摘要，建议结合已读论文和个人笔记继续补充。",
        "",
        "## 可用于综述的代表性论文表",
        "见同目录年度 Excel 文件。",
    ])
    return "\n".join(lines)


def export_annual_excel(year: int, papers: list[dict], profile_slug: str) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORTS_DIR / f"{year}_{profile_slug}_annual_library.xlsx"
    headers = ["星级", "综合分", "推荐说明", "标题", "作者", "年份", "期刊", "IF", "JCR分区", "中科院分区", "引用数", "DOI", "URL", "摘要来源", "摘要是否完整", "中文摘要", "阅读状态", "是否收藏", "加入综述"]
    if Workbook is None:
        csv_path = path.with_suffix(".csv")
        lines = [",".join(headers)]
        for paper in papers:
            values = [
                star_text(paper.get("display_rating")),
                paper.get("final_score") or paper.get("relevance_score") or 0,
                paper.get("rating_reason") or rating_label(paper.get("display_rating")),
                paper.get("title", ""),
                parse_authors(paper.get("authors", "")),
                paper.get("year") or paper.get("publication_year") or year,
                paper.get("journal", ""),
                paper.get("impact_factor", ""),
                paper.get("jcr_quartile", ""),
                paper.get("cas_quartile", ""),
                paper.get("citation_count", ""),
                paper.get("doi", ""),
                paper.get("url", ""),
                paper.get("abstract_source", ""),
                "是" if paper.get("abstract_is_complete") else "否",
                paper.get("abstract_zh", ""),
                paper.get("reading_status", ""),
                "是" if paper.get("favorite") or paper.get("is_favorite") else "否",
                "是" if paper.get("included_in_review") else "否",
            ]
            lines.append(",".join('"' + str(value).replace('"', '""') + '"' for value in values))
        csv_path.write_text("\n".join(lines), encoding="utf-8-sig")
        return csv_path
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年论文"
    ws.append(headers)
    for paper in papers:
        ws.append([
            star_text(paper.get("display_rating")),
            paper.get("final_score") or paper.get("relevance_score") or 0,
            paper.get("rating_reason") or rating_label(paper.get("display_rating")),
            paper.get("title", ""),
            parse_authors(paper.get("authors", "")),
            paper.get("year") or paper.get("publication_year") or year,
            paper.get("journal", ""),
            paper.get("impact_factor", ""),
            paper.get("jcr_quartile", ""),
            paper.get("cas_quartile", ""),
            paper.get("citation_count", ""),
            paper.get("doi", ""),
            paper.get("url", ""),
            paper.get("abstract_source", ""),
            "是" if paper.get("abstract_is_complete") else "否",
            paper.get("abstract_zh", ""),
            paper.get("reading_status", ""),
            "是" if paper.get("favorite") or paper.get("is_favorite") else "否",
            "是" if paper.get("included_in_review") else "否",
        ])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 56)
    wb.save(path)
    return path


def ensure_review_jobs_table() -> None:
    execute(
        """
        CREATE TABLE IF NOT EXISTS review_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile TEXT,
            topic TEXT,
            review_type TEXT,
            language TEXT,
            paper_count INTEGER DEFAULT 0,
            status TEXT,
            output_path TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )


def selected_review_papers(
    profile: str,
    time_range: str,
    level_filter: str,
    abstract_filter: str,
    if_filter: str,
    reading_status: str,
    max_papers: int,
) -> list[dict]:
    from datetime import datetime

    clauses = []
    params: list = []
    if profile:
        clauses.append("profile_id = ?")
        params.append(profile)
    current_year = datetime.now().year
    if time_range == "1y":
        clauses.append("year >= ?")
        params.append(current_year - 1)
    elif time_range == "3y":
        clauses.append("year >= ?")
        params.append(current_year - 3)
    elif time_range == "5y":
        clauses.append("year >= ?")
        params.append(current_year - 5)
    if level_filter == "a":
        clauses.append("recommendation_level LIKE 'A%'")
    elif level_filter == "aplus":
        clauses.append("recommendation_level LIKE 'A+%'")
    elif level_filter == "b_plus":
        clauses.append("(recommendation_level LIKE 'A%' OR recommendation_level LIKE 'B%')")
    if abstract_filter == "complete":
        clauses.append("abstract_is_complete = 1")
    if if_filter == "matched":
        clauses.append("impact_factor IS NOT NULL AND impact_factor != ''")
    elif if_filter in {"ge1", "ge3", "ge5"}:
        threshold = {"ge1": 1, "ge3": 3, "ge5": 5}[if_filter]
        clauses.append("CAST(NULLIF(impact_factor, '') AS REAL) >= ?")
        params.append(threshold)
    if reading_status:
        clauses.append("reading_status = ?")
        params.append(reading_status)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    params.append(max(5, min(max_papers, 200)))
    return rows(
        f"""
        SELECT p.*, rn.paper_topic, rn.core_method, rn.paper_contribution,
               rn.inspiration, rn.possible_ideas, rn.reason
        FROM papers p
        LEFT JOIN reading_notes rn ON rn.paper_id = p.stable_id
        {where}
        ORDER BY
            CASE
                WHEN p.recommendation_level LIKE 'A+%' THEN 1
                WHEN p.recommendation_level LIKE 'A%' THEN 2
                WHEN p.recommendation_level LIKE 'B%' THEN 3
                ELSE 4
            END,
            p.abstract_is_complete DESC,
            p.final_score DESC,
            p.year DESC
        LIMIT ?
        """,
        tuple(params),
    )


def group_methods(papers: list[dict]) -> list[tuple[str, list[dict]]]:
    groups = {
        "ePIE / iterative phase retrieval": ["epie", "iterative", "phase retrieval", "probe recovery"],
        "WDD / SSB / direct methods": ["wdd", "ssb", "single sideband", "direct"],
        "Multislice ptychography": ["multislice", "thick", "multiple scattering"],
        "Deep learning assisted reconstruction": ["deep", "learning", "neural", "unrolling", "ai"],
        "Applications and characterization": ["strain", "semiconductor", "atomic", "low-dose", "dose"],
    }
    bucketed = {name: [] for name in groups}
    bucketed["Other relevant studies"] = []
    for paper in papers:
        text = " ".join(str(paper.get(key) or "").lower() for key in [
            "title", "abstract_original", "abstract_zh", "core_method", "paper_topic"
        ])
        placed = False
        for name, keywords in groups.items():
            if any(keyword in text for keyword in keywords):
                bucketed[name].append(paper)
                placed = True
                break
        if not placed:
            bucketed["Other relevant studies"].append(paper)
    return [(name, items) for name, items in bucketed.items() if items]


def build_review_outline(topic: str, review_type: str, language: str, papers: list[dict]) -> str:
    method_groups = group_methods(papers)
    lines = [
        f"# {topic}",
        "",
        "> 本综述为 AI 辅助生成初稿，请人工核对原文、引用和事实。",
        "",
        f"- 综述类型：{review_type}",
        f"- 输出语言：{language}",
        f"- 使用文献：{len(papers)} 篇",
        "",
        "## 1. 研究背景",
        "说明该方向的科学问题、实验/算法瓶颈，以及为什么需要系统梳理近期文献。",
        "",
        "## 2. 发展脉络",
        "按年份总结代表性工作如何从基础重建、鲁棒性提升，走向应用扩展和自动化分析。",
        "",
        "## 3. 方法分类",
    ]
    for idx, (name, items) in enumerate(method_groups, 1):
        lines.extend([
            f"### 3.{idx} {name}",
            f"代表论文：{'; '.join((paper.get('title') or 'Untitled')[:80] for paper in items[:3])}",
            "",
        ])
    lines.extend([
        "## 4. 代表性文献与关键贡献",
        "用表格整理论文、年份、期刊、IF、核心方法、贡献、局限和推荐等级。",
        "",
        "## 5. 当前挑战",
        "从数据质量、剂量、厚样品、多重散射、算法稳定性、可解释性和实验迁移等角度展开。",
        "",
        "## 6. 未来方向与可创新点",
        "结合当前课题，提炼可落地的算法改进、实验验证和应用迁移方向。",
        "",
        "## 7. 参考文献列表",
        "后续人工核对 DOI、页码和引用格式。",
    ])
    return "\n".join(lines)


def build_review_draft(topic: str, review_type: str, language: str, papers: list[dict]) -> str:
    outline = build_review_outline(topic, review_type, language, papers)
    lines = [
        outline,
        "",
        "## 综述初稿",
        "",
        "### 研究背景",
        "从当前归档文献看，该方向的研究重点集中在重建质量、相位恢复稳定性、数据采集效率以及面向真实样品的应用迁移。高质量论文通常同时具备完整摘要、明确的方法贡献和可复现实验路径，因此适合作为综述主线的支撑材料。",
        "",
        "### 研究脉络",
    ]
    year_groups: dict[str, list[dict]] = {}
    for paper in papers:
        year_groups.setdefault(str(paper.get("year") or "未知年份"), []).append(paper)
    for year in sorted(year_groups.keys()):
        titles = "；".join((paper.get("title") or "Untitled")[:80] for paper in year_groups[year][:4])
        lines.append(f"- {year}：{titles}")
    lines.extend(["", "### 方法分类与代表性论文"])
    for name, items in group_methods(papers):
        lines.append(f"#### {name}")
        for paper in items[:5]:
            contribution = paper.get("paper_contribution") or paper.get("reason") or paper.get("abstract_zh") or paper.get("abstract_original") or "需要进一步阅读原文确认贡献。"
            lines.append(f"- **{paper.get('title') or 'Untitled'}**：{str(contribution)[:220]}")
        lines.append("")
    lines.extend([
        "### 当前挑战",
        "- 完整摘要和高质量元数据仍不均衡，需要人工核对代表性论文的实验条件与引用信息。",
        "- 不同方法在低剂量、厚样品、复杂噪声和真实实验数据上的适用边界仍需比较。",
        "- AI 辅助重建类方法需要重点关注泛化性、可解释性和与传统迭代方法的结合方式。",
        "",
        "### 未来方向与可创新点",
    ])
    ideas = [paper.get("possible_ideas") or paper.get("inspiration") for paper in papers if paper.get("possible_ideas") or paper.get("inspiration")]
    if ideas:
        for idx, idea in enumerate(ideas[:8], 1):
            lines.append(f"{idx}. {str(idea)[:260]}")
    else:
        lines.extend([
            "1. 将直接重建方法与迭代优化方法结合，提高初始相位估计稳定性。",
            "2. 构建面向真实实验噪声的模拟数据集，用于验证深度展开网络的鲁棒性。",
            "3. 围绕半导体器件和低剂量场景，设计可复用的 ptychographic reconstruction 流程。",
        ])
    lines.extend([
        "",
        "### 参考文献候选",
    ])
    for idx, paper in enumerate(papers, 1):
        lines.append(f"{idx}. {paper.get('title') or 'Untitled'}. {paper.get('journal') or 'Unknown Journal'}, {paper.get('year') or 'n.d.'}. DOI: {paper.get('doi') or '待补全'}")
    return "\n".join(lines)


def export_literature_table(papers: list[dict], base_path: Path) -> Path:
    headers = ["论文", "年份", "期刊", "IF", "方法/主题", "主要贡献", "局限/备注", "推荐等级", "链接"]
    if Workbook is None:
        path = base_path.with_suffix(".csv")
        lines = [",".join(headers)]
        for paper in papers:
            values = [
                paper.get("title", ""),
                paper.get("year", ""),
                paper.get("journal", ""),
                paper.get("impact_factor", ""),
                paper.get("core_method") or paper.get("paper_topic") or "",
                paper.get("paper_contribution") or paper.get("abstract_zh") or paper.get("abstract_original") or "",
                paper.get("reason") or "",
                paper.get("recommendation_level") or "",
                paper.get("url") or "",
            ]
            lines.append(",".join('"' + str(value).replace('"', '""') + '"' for value in values))
        path.write_text("\n".join(lines), encoding="utf-8-sig")
        return path
    path = base_path.with_suffix(".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Literature Table"
    ws.append(headers)
    for paper in papers:
        ws.append([
            paper.get("title", ""),
            paper.get("year", ""),
            paper.get("journal", ""),
            paper.get("impact_factor", ""),
            paper.get("core_method") or paper.get("paper_topic") or "",
            paper.get("paper_contribution") or paper.get("abstract_zh") or paper.get("abstract_original") or "",
            paper.get("reason") or "",
            paper.get("recommendation_level") or "",
            paper.get("url") or "",
        ])
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 56)
    wb.save(path)
    return path


def build_wechat_text(context: dict, papers: list[dict], report_path: str = "") -> tuple[str, str]:
    profile = context.get("current_profile", {})
    numbers = context.get("dashboard_numbers", {})
    title = f"今日高质量文献雷达｜{today_string()}"
    lines = [
        f"📚 今日高质量文献雷达｜{today_string()}",
        "",
        f"方向：{profile.get('display_name') or profile.get('name') or context.get('current_profile_id')}",
        f"候选论文：{numbers.get('total_found', 0)} 篇",
        f"完整摘要：{numbers.get('abstract_completed', 0)} 篇",
        f"推荐精读：{numbers.get('recommended', 0)} 篇",
        "",
    ]
    for idx, paper in enumerate(papers, 1):
        lines.extend([
            f"🔥 {idx}. {paper.get('title', '无标题')}",
            f"推荐等级：{paper.get('recommendation_level') or '未评级'}",
            f"综合分：{paper.get('final_score') or paper.get('relevance_score') or 0}/30",
            f"期刊：{paper.get('journal') or '未知'}",
            f"IF：{paper.get('impact_factor') or '未匹配'}",
            f"引用数：{paper.get('citation_count') or 0}",
            "",
            "为什么推荐：高相关，适合进入今日阅读队列。",
            f"链接：{paper_display_url(paper)}",
            "",
        ])
    lines.extend(["🧠 今日建议：优先读第 1 篇，并整理进本周总结。"])
    if report_path:
        lines.extend(["", f"完整日报：{report_path}"])
    return title, "\n".join(lines)


def send_serverchan_message(title: str, desp: str) -> tuple[bool, str]:
    if requests is None:
        return False, "缺少 requests 依赖"
    sct_key = os.environ.get("SCT_KEY", "")
    if not sct_key:
        return False, "未配置 SCT_KEY"
    try:
        response = requests.post(f"https://sctapi.ftqq.com/{sct_key}.send", data={"title": title, "desp": desp}, timeout=15)
        return response.status_code == 200, response.text[:300]
    except Exception as exc:
        return False, str(exc)


def mask_key(value: str) -> str:
    if not value:
        return "未配置"
    if len(value) <= 8:
        return "已配置"
    return f"{value[:4]}...{value[-4:]}"


def save_env_values(values: dict) -> None:
    existing = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if not line.strip() or line.lstrip().startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            existing[key.strip()] = value.strip()
    for key, value in values.items():
        value = value.strip()
        if value:
            existing[key] = value
    ordered = [
        "SERPAPI_API_KEY",
        "SCT_KEY",
        "SEMANTIC_SCHOLAR_API_KEY",
        "KIMI_API_KEY",
        "PAPER_AI_PROVIDER",
        "KIMI_PARSE_MODEL",
        "OPENAI_API_KEY",
        "ANTHROPIC_API_KEY",
    ]
    lines = [f"{key}={existing.get(key, '')}" for key in ordered]
    ENV_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    for key in ordered:
        if existing.get(key):
            os.environ[key] = existing[key]


def latest_run_for_profile(profile_id: str = "") -> dict | None:
    if profile_id:
        return row(
            "SELECT * FROM runs WHERE profile = ? ORDER BY run_time DESC, id DESC LIMIT 1",
            (profile_id,),
        )
    return latest_run()


def activate_profile(profile_id: str) -> bool:
    doc = load_profiles_doc()
    profiles = doc.get("profiles", {})
    profile_id = slugify_profile_id(profile_id)
    if profile_id not in profiles:
        return False
    if doc.get("default_profile") == profile_id:
        return True
    doc["default_profile"] = profile_id
    save_profiles_doc(doc)
    return True


def redirect_if_run_not_active(run: dict | None) -> RedirectResponse | None:
    if not run:
        return None
    current = load_profiles_doc().get("default_profile", "") or ""
    if not current or run.get("profile") == current:
        return None
    latest = latest_run_for_profile(current)
    if latest:
        return RedirectResponse(f"/runs/{latest['id']}", status_code=303)
    return RedirectResponse("/fetch?no_runs=1", status_code=303)


def load_app_settings() -> dict:
    if APP_SETTINGS_PATH.exists():
        try:
            return json.loads(APP_SETTINGS_PATH.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            pass
    return {"advanced_features_enabled": False}


def save_app_settings(doc: dict) -> None:
    APP_SETTINGS_PATH.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")


def advanced_features_enabled() -> bool:
    return bool(load_app_settings().get("advanced_features_enabled"))


def common_context(request: Request) -> dict:
    latest = None
    profiles_doc = load_profiles_doc() if PROFILES_PATH.exists() else {"profiles": {}}
    profile_id = profiles_doc.get("default_profile", "") or ""
    current_profile = profiles_doc.get("profiles", {}).get(profile_id, {})
    latest = latest_run_for_profile(profile_id) if profile_id else None
    if not profile_id and latest:
        profile_id = latest.get("profile") or ""
        current_profile = profiles_doc.get("profiles", {}).get(profile_id, {})
    profile_where = "WHERE profile_id = ?" if profile_id else ""
    profile_params = (profile_id,) if profile_id else ()
    stats = row(
        f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN abstract_is_complete = 1 THEN 1 ELSE 0 END) AS abstract_completed,
            SUM(CASE WHEN impact_factor IS NOT NULL AND impact_factor != '' THEN 1 ELSE 0 END) AS if_matched,
            SUM(CASE WHEN recommendation_level LIKE 'A%' THEN 1 ELSE 0 END) AS recommended,
            SUM(CASE WHEN pushed_to_wechat = 1 THEN 1 ELSE 0 END) AS pushed,
            SUM(CASE WHEN reading_status = 'read' THEN 1 ELSE 0 END) AS read_count,
            SUM(CASE WHEN reading_status = 'unread' OR reading_status IS NULL THEN 1 ELSE 0 END) AS unread_count
        FROM papers
        {profile_where}
        """,
        profile_params,
    ) or {}
    weekly = row(
        f"""
        SELECT COUNT(*) AS total
        FROM papers
        WHERE date(COALESCE(created_at, updated_at)) >= date('now', '-7 days')
        {"AND profile_id = ?" if profile_id else ""}
        """,
        profile_params,
    ) or {}
    dashboard_numbers = {
        "total_found": stats.get("total") or 0,
        "abstract_completed": stats.get("abstract_completed") or 0,
        "if_matched": stats.get("if_matched") or 0,
        "recommended": stats.get("recommended") or 0,
        "pushed": stats.get("pushed") or (latest or {}).get("pushed_count") or 0,
        "unread": stats.get("unread_count") or 0,
        "weekly_new": weekly.get("total") or 0,
        "total": stats.get("total") or 0,
    }
    return {
        "request": request,
        "latest": latest,
        "stats": stats,
        "dashboard_numbers": dashboard_numbers,
        "current_profile": current_profile,
        "current_profile_id": profile_id,
        "advanced_features_enabled": advanced_features_enabled(),
    }


@app.on_event("startup")
def startup() -> None:
    init_db()
    ensure_review_jobs_table()
    ensure_prompt_template_records()
    try:
        from journal_rank_enhancer import ensure_journal_metrics_file
        ensure_journal_metrics_file(str(JOURNAL_METRICS_PATH))
    except Exception:
        pass
    if PROFILES_PATH.exists():
        try:
            sync_profiles(load_profiles_doc())
        except (OSError, sqlite3.Error):
            pass


@app.get("/")
def dashboard(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    if request.query_params.get("switched") == "1":
        context["profile_switched"] = True
    if request.query_params.get("run_finished_other") == "1":
        context["run_finished_other"] = True
    recent_runs = rows(
        "SELECT * FROM runs WHERE (? = '' OR profile = ?) ORDER BY run_time DESC, id DESC LIMIT 3",
        (profile_id, profile_id),
    )
    current_profile = context["current_profile"]
    today_stats = row(
        """
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN date(created_at) = date('now') THEN 1 ELSE 0 END) AS added_today,
            SUM(CASE WHEN abstract_is_complete = 1 THEN 1 ELSE 0 END) AS abstract_completed,
            SUM(CASE WHEN impact_factor IS NOT NULL AND impact_factor != '' THEN 1 ELSE 0 END) AS if_matched,
            SUM(CASE WHEN recommendation_level LIKE 'A%' THEN 1 ELSE 0 END) AS recommended
        FROM papers
        WHERE date(COALESCE(created_at, updated_at)) = date('now')
          AND (? = '' OR profile_id = ?)
        """,
        (profile_id, profile_id),
    ) or {}
    context.update({
        "run_started": request.query_params.get("run_started") == "1",
        "papers": top_papers_for_profile(profile_id, 3),
        "recent_runs": recent_runs,
        "profile_latest": latest_run_for_profile(profile_id),
        "today_stats": today_stats,
        "year_overview": library_year_overview(3, profile_id),
        "recent_reports": recent_report_files(5),
        "roadmap_lite": build_roadmap_lite(profile_id),
        "db_path": DB_PATH,
    })
    return templates.TemplateResponse(request, "dashboard.html", context)


@app.get("/library")
def library_hub(request: Request):
    return RedirectResponse("/papers", status_code=307)


def idea_lab_redirect_url(**params: str) -> str:
    from urllib.parse import urlencode
    clean = {k: v for k, v in params.items() if v not in ("", None)}
    return f"/idea-lab?{urlencode(clean)}" if clean else "/idea-lab"


@app.get("/ask")
def ask_page_entry(request: Request, q: str = "", profile: str = ""):
    if not advanced_features_enabled():
        return RedirectResponse("/papers", status_code=307)
    context = common_context(request)
    return templates.TemplateResponse(
        request,
        "ask.html",
        {
            **context,
            "profiles": rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id"),
            "selected_profile": profile or context.get("current_profile_id", ""),
            "question": q,
            "result": {},
            "rag_status": rag_status(),
        },
    )


@app.get("/citations")
def mvp_redirect_citations():
    return RedirectResponse("/reports", status_code=307)


@app.get("/wechat")
@app.get("/wechat/push-today")
def mvp_redirect_wechat():
    return RedirectResponse("/", status_code=307)


@app.get("/writing")
def writing_hub(request: Request):
    return RedirectResponse("/reports", status_code=307)


@app.get("/reports")
def reports_page(request: Request):
    context = common_context(request)
    return templates.TemplateResponse(
        request,
        "reports.html",
        {
            **context,
            "years": available_years(),
            "current_year": current_year(),
            "recent_reports": recent_report_files(10),
        },
    )


@app.get("/writing/daily-report")
def writing_daily_report():
    return RedirectResponse("/generate/daily-report", status_code=307)


@app.get("/writing/weekly")
def writing_weekly():
    return RedirectResponse("/weekly/generate", status_code=307)


@app.get("/writing/review")
def writing_review():
    return RedirectResponse("/generate/review", status_code=307)


@app.get("/writing/seminar")
def writing_seminar_page(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    papers = papers_ingested_recent(7, profile_id, limit=8)
    content = build_seminar_markdown(context, papers)
    return templates.TemplateResponse(
        request,
        "writing_seminar.html",
        {**context, "papers": papers, "markdown_preview": content},
    )


@app.post("/writing/seminar")
def writing_seminar_action(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    papers = papers_ingested_recent(7, profile_id, limit=8)
    profile_slug = safe_slug(profile_id or "research")
    SEMINAR_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = SEMINAR_REPORTS_DIR / f"{week_string()}_{profile_slug}_seminar.md"
    content = build_seminar_markdown(context, papers)
    write_markdown(path, content)
    return templates.TemplateResponse(
        request,
        "writing_seminar.html",
        {**context, "papers": papers, "markdown_preview": content, "generated_path": path},
    )


@app.get("/writing/intro-draft")
def writing_intro_page(request: Request):
    context = common_context(request)
    papers = query_papers({"profile": context.get("current_profile_id", ""), "rating": "4"}, limit=15)
    return templates.TemplateResponse(
        request,
        "writing_intro.html",
        {**context, "papers": papers[:10], "markdown_preview": ""},
    )


@app.post("/writing/intro-draft")
def writing_intro_action(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    papers = query_papers({"profile": profile_id, "rating": "4"}, limit=15)
    body = build_intro_draft_markdown(context, papers[:10])
    profile_slug = safe_slug(profile_id or "research")
    INTRO_DRAFTS_DIR.mkdir(parents=True, exist_ok=True)
    path = INTRO_DRAFTS_DIR / f"{today_string()}_{profile_slug}_intro.md"
    write_markdown(path, body)
    return templates.TemplateResponse(
        request,
        "writing_intro.html",
        {**context, "papers": papers[:10], "markdown_preview": body, "generated_path": path},
    )


def answer_with_rag(question: str, profile: str = "", scope: str = "library", source_id: str = "", paper_id: str = "") -> dict:
    if paper_id:
        rag_index_paper(paper_id)
    chunks = rag_search(question, profile=profile, source_id=source_id, paper_id=paper_id, limit=8)
    sources = rag_source_list(chunks)
    if not chunks:
        answer = "## 结论\n未找到足够依据回答这个问题。请先重建 RAG 索引，或缩小到已有论文/笔记范围。\n\n## 依据来源\n暂无。\n\n## 可操作建议\n尝试补充摘要、上传 PDF、保存阅读笔记后重新提问。"
    elif ai_parse_configured():
        answer = call_llm(build_answer_prompt(question, chunks)) or ""
        if not answer:
            answer = "## 结论\nAI 请求失败，但已找到相关来源。请检查 API Key、网络或额度。\n\n## 依据来源\n" + "\n".join(
                f"- {src.get('title')}（{src.get('source_type')}）：{src.get('url') or src.get('source_id')}" for src in sources
            )
    else:
        answer = "## 结论\n已找到相关来源，但未配置 AI Key，暂不生成综合回答。\n\n## 依据来源\n" + "\n".join(
            f"- {src.get('title')}（{src.get('source_type')}）：{src.get('url') or src.get('source_id')}" for src in sources
        )
    save_rag_query(question, scope, profile, source_id or paper_id, answer, sources)
    return {"question": question, "answer": answer, "chunks": chunks, "sources": sources}


@app.post("/ask")
def ask_action(request: Request, question: str = Form(""), profile: str = Form("")):
    context = common_context(request)
    selected_profile = profile or context.get("current_profile_id", "")
    result = answer_with_rag(question, selected_profile, "library") if question.strip() else {}
    return templates.TemplateResponse(
        request,
        "ask.html",
        {
            **context,
            "profiles": rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id"),
            "selected_profile": selected_profile,
            "question": question,
            "result": result,
            "rag_status": rag_status(),
        },
    )


@app.post("/rag/reindex")
def rag_reindex(profile: str = Form("")):
    stats = rag_index_all(profile)
    suffix = f"?indexed={stats.get('chunks', 0)}&docs={stats.get('papers', 0) + stats.get('files', 0)}"
    return RedirectResponse(f"/ask{suffix}", status_code=303)


@app.get("/agents")
def agents_page(request: Request):
    if not advanced_features_enabled():
        return RedirectResponse("/", status_code=307)
    context = common_context(request)
    return templates.TemplateResponse(
        request,
        "agents.html",
        {**context, "agents": list_agents(), "agent_runs": recent_agent_runs(8)},
    )


@app.post("/agents/plan")
def agents_plan(agent_name: str = Form(...), user_request: str = Form(""), profile: str = Form("")):
    run_id = create_agent_plan(agent_name, profile, user_request)
    return RedirectResponse(f"/agents/{run_id}", status_code=303)


def agent_tool_executor(tool_name: str, run: dict) -> dict:
    profile_id = run.get("profile") or load_profiles_doc().get("default_profile", "")
    context = {
        "current_profile": load_profiles_doc().get("profiles", {}).get(profile_id, {}),
        "current_profile_id": profile_id,
        "dashboard_numbers": {},
    }
    if tool_name == "rag_index":
        stats = rag_index_all(profile_id)
        return {"message": f"已更新 RAG 索引：{stats['chunks']} 个片段。", "content": json.dumps(stats, ensure_ascii=False)}
    if tool_name in {"rag_search", "build_context"}:
        chunks = rag_search(run.get("user_request") or profile_id or "ptychography", profile=profile_id, limit=8)
        return {"message": f"找到 {len(chunks)} 条相关依据。", "content": format_rag_context(chunks)}
    if tool_name == "generate_daily_report":
        papers = daily_report_candidates(profile_id, 5)
        path = OUTPUT_DIR / "daily_reports" / f"{today_string()}_{safe_slug(profile_id)}_agent_daily.md"
        content = build_daily_report_markdown(context, papers)
        write_markdown(path, content)
        return {"message": "已生成 Agent 日报。", "content": content, "path": str(path), "output_type": "markdown"}
    if tool_name == "generate_review":
        papers = selected_review_papers(profile_id, "3y", "b_plus", "complete", "", "", 30)
        content = build_review_draft(f"{profile_id} RAG 综述", "领域进展综述", "中文", papers)
        path = REVIEW_REPORTS_DIR / f"{today_string()}_{safe_slug(profile_id)}_agent_review.md"
        write_markdown(path, content)
        return {"message": f"已基于 {len(papers)} 篇论文生成综述草稿。", "content": content, "path": str(path), "output_type": "markdown"}
    if tool_name == "progress_metrics":
        doc = load_profiles_doc()
        profile_obj = {**doc.get("profiles", {}).get(profile_id, {}), "id": profile_id}
        dashboard = build_progress_dashboard(profile_id, profile_obj)
        return {"message": "已统计研究进展。", "content": json.dumps(dashboard.get("metrics", {}), ensure_ascii=False)}
    if tool_name == "roadmap_candidates":
        candidates = roadmap_candidate_papers(profile_id, limit=80)
        preview = "\n".join(f"- {p.get('title')} ({p.get('year') or 'n.d.'})" for p in candidates[:20])
        return {"message": f"已筛选 {len(candidates)} 篇候选里程碑/代表论文。", "content": preview}
    if tool_name in {"roadmap_build", "roadmap_save"}:
        stats = generate_roadmap_map(profile_id)
        return {"message": f"已生成方向地图：候选 {stats['candidates']} 篇，路线 {stats['routes']} 条。", "content": json.dumps(stats, ensure_ascii=False)}
    if tool_name == "roadmap_report":
        doc = load_profiles_doc()
        profile_obj = {**doc.get("profiles", {}).get(profile_id, {}), "id": profile_id}
        papers = roadmap_candidate_papers(profile_id, limit=50)
        content = build_roadmap_markdown(profile_obj, papers)
        ROADMAP_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        path = ROADMAP_REPORTS_DIR / f"{today_string()}_{safe_slug(profile_id)}_research_map.md"
        write_markdown(path, content)
        return {"message": "已导出研究方向地图 Markdown 报告。", "content": content, "path": str(path), "output_type": "markdown"}
    if tool_name == "save_snapshot":
        doc = load_profiles_doc()
        profile_obj = {**doc.get("profiles", {}).get(profile_id, {}), "id": profile_id}
        dashboard = build_progress_dashboard(profile_id, profile_obj)
        save_progress_snapshot(profile_id, dashboard["metrics"], dashboard["suggestions"])
        return {"message": "已保存研究进展快照。", "content": "\n".join(dashboard["suggestions"])}
    if tool_name == "format_citations":
        collections = rows("SELECT * FROM citation_collections ORDER BY id DESC LIMIT 1")
        if not collections:
            return {"message": "暂无引用篮，请先从论文详情加入引用。"}
        items = rows(
            """
            SELECT p.* FROM citation_items ci JOIN papers p ON p.stable_id = ci.paper_id
            WHERE ci.collection_id = ? ORDER BY ci.sort_order
            """,
            (collections[0]["id"],),
        )
        from citation_format import format_citations
        preview = format_citations(items, collections[0].get("format_default") or "gbt7714")
        return {"message": f"已格式化 {len(items)} 条引用。", "content": preview}
    if tool_name == "mine_idea":
        query = run.get("user_request") or profile_id or "research idea"
        local_hits, rag_chunks = search_local_evidence(query, profile_id, limit=8)
        external_hits = search_external_evidence(query)
        doc = load_profiles_doc()
        profile_obj = doc.get("profiles", {}).get(profile_id, {})
        result = polish_idea_structured(query, profile_obj, "", local_hits, external_hits, call_llm, ai_parse_configured(), rag_chunks=rag_chunks)
        return {"message": "已生成想法打磨草稿。", "content": result.get("raw_markdown") or "", "output_type": "markdown"}
    if tool_name == "link_papers":
        query = run.get("user_request") or profile_id
        chunks = rag_search(query, profile=profile_id, limit=8)
        preview = "\n".join(f"- {c.get('metadata', {}).get('title') or c.get('title') or '片段'}" for c in chunks)
        return {"message": f"已关联 {len(chunks)} 条文献依据。", "content": preview}
    if tool_name == "save_idea":
        query = run.get("user_request") or "Agent 灵感"
        local_hits, rag_chunks = search_local_evidence(query, profile_id, limit=5)
        external_hits = search_external_evidence(query)
        doc = load_profiles_doc()
        profile_obj = doc.get("profiles", {}).get(profile_id, {})
        result = polish_idea_structured(query, profile_obj, "", local_hits, external_hits, call_llm, ai_parse_configured(), rag_chunks=rag_chunks)
        from datetime import datetime
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ensure_ideas_tables()
        with connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO ideas (title, body_md, profile_id, status, linked_paper_ids_json,
                possible_direction, next_tasks, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (query[:200], result.get("raw_markdown") or "", profile_id, "想法中", "", "", ", ".join(result.get("fetch_keywords") or []), now, now),
            )
            conn.commit()
            iid = int(cursor.lastrowid)
        return {"message": f"已保存灵感 #{iid}。", "content": result.get("raw_markdown") or ""}
    return {"message": "步骤已记录。请在对应页面继续确认或查看输出。"}


@app.get("/agents/{run_id}")
def agent_run_page(request: Request, run_id: int):
    if not advanced_features_enabled():
        return RedirectResponse("/", status_code=307)
    run = get_agent_run(run_id)
    if not run:
        return RedirectResponse("/agents", status_code=303)
    return templates.TemplateResponse(
        request,
        "agent_run.html",
        {**common_context(request), "run": run},
    )


@app.post("/agents/{run_id}/execute")
def agent_run_execute(run_id: int):
    execute_agent_run(run_id, agent_tool_executor)
    return RedirectResponse(f"/agents/{run_id}", status_code=303)


@app.get("/generate")
def generate_page(request: Request):
    return RedirectResponse("/reports", status_code=307)


@app.get("/fetch")
def fetch_page(request: Request):
    from datetime import datetime as dt
    from urllib.parse import unquote

    profile_rows = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    current_year = dt.now().year
    prefill = unquote(request.query_params.get("prefill_keywords") or "")
    branch = request.query_params.get("branch") or ""
    from_source = request.query_params.get("from") or ""
    branch_title = branch_meta(branch).get("title") if branch else ""
    return templates.TemplateResponse(
        request,
        "fetch.html",
        {
            **common_context(request),
            "profiles": profile_rows,
            "current_year": current_year,
            "years": list(range(current_year, current_year - 10, -1)),
            "year_range_years": list(range(current_year, 1994, -1)),
            "prefill_keywords": prefill,
            "prefill_branch": branch,
            "prefill_branch_title": branch_title,
            "prefill_from": from_source,
        },
    )


def web_run_log_tail(max_chars: int = 2500) -> str:
    if not WEB_RUN_LOG.exists():
        return ""
    try:
        text = WEB_RUN_LOG.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return ""
    return text[-max_chars:]


RUN_FINISH_MARKERS = (
    "===== Web run finished =====",
    "论文追踪完成",
    "每日论文追踪完成",
    "无新增论文，仍将更新",
    "今日未检索到相关论文",
    "今日暂无新增论文",
    "无需解析",
    "SQLite 数据库已更新",
    "数据库已更新",
    "error: unrecognized arguments",
)


def _web_run_log_blocks() -> list[str]:
    text = web_run_log_tail(80000)
    parts = text.split("===== Web run started =====")
    if len(parts) < 2:
        return []
    return parts[1:]


def _block_is_finished(block: str) -> bool:
    return bool(block) and any(marker in block for marker in RUN_FINISH_MARKERS)


def web_run_log_has_success(since_epoch: float = 0) -> bool:
    """True if any session finished; ignores a trailing in-progress block."""
    if not WEB_RUN_LOG.exists():
        return False
    if since_epoch > 0:
        try:
            if WEB_RUN_LOG.stat().st_mtime < since_epoch - 2:
                return False
        except OSError:
            return False
    blocks = _web_run_log_blocks()
    if not blocks:
        return False
    for block in blocks[:-1]:
        if _block_is_finished(block):
            return True
    return _block_is_finished(blocks[-1])


def run_id_newer_than(after: int) -> Optional[int]:
    newer = row("SELECT id FROM runs WHERE id > ? ORDER BY id DESC LIMIT 1", (after,))
    return int(newer["id"]) if newer else None


def wait_session_done(after: int, since: float) -> tuple[bool, Optional[int]]:
    newer_id = run_id_newer_than(after)
    if newer_id:
        return True, newer_id
    if web_run_log_has_success(since):
        # Log finished but runs.id did not advance — do not reuse an older run row.
        return True, None
    return False, None


def resolve_waiting_run_id(after: int) -> Optional[int]:
    return run_id_newer_than(after)


@app.get("/api/run-status")
def api_run_status(after: int = 0, since: float = 0, tries: int = 0):
    latest = latest_run()
    latest_id = int((latest or {}).get("id") or 0)
    done, run_id = wait_session_done(after, since)
    timed_out = tries >= 90
    resolved_id = run_id or resolve_waiting_run_id(after)
    return JSONResponse({
        "done": done or timed_out,
        "run_id": resolved_id if (done or timed_out) else None,
        "latest_id": latest_id,
        "finished_marker": done,
        "timed_out": timed_out,
        "log_tail": web_run_log_tail(2000),
    })


@app.get("/runs/waiting")
def runs_waiting(request: Request, after: int = 0, since: float = 0, tries: int = 0):
    done, run_id = wait_session_done(after, since)
    if done:
        if run_id:
            mismatch = redirect_if_run_not_active(get_run(run_id))
            if mismatch:
                return RedirectResponse("/?run_finished_other=1", status_code=303)
            return RedirectResponse(f"/runs/{run_id}?from_waiting=1", status_code=303)
        return RedirectResponse("/fetch?run_finished_no_db=1", status_code=303)
    if tries >= 90:
        run_id = resolve_waiting_run_id(after)
        if run_id:
            mismatch = redirect_if_run_not_active(get_run(run_id))
            if mismatch:
                return RedirectResponse("/?run_finished_other=1", status_code=303)
            return RedirectResponse(f"/runs/{run_id}?wait_timeout=1", status_code=303)
        return RedirectResponse("/fetch?wait_timeout=1", status_code=303)
    return templates.TemplateResponse(
        request,
        "run_waiting.html",
        {
            **common_context(request),
            "after": after,
            "since": since or time.time(),
            "tries": tries,
            "log_tail": web_run_log_tail(1500),
        },
    )


@app.get("/runs/latest")
def runs_latest(request: Request):
    profile_id = load_profiles_doc().get("default_profile", "") or ""
    latest = latest_run_for_profile(profile_id) if profile_id else latest_run()
    if not latest:
        return RedirectResponse("/fetch?no_runs=1", status_code=303)
    return RedirectResponse(f"/runs/{latest['id']}", status_code=303)


@app.get("/runs/{run_id}")
def run_result_page(request: Request, run_id: int):
    run = get_run(run_id)
    if not run:
        return templates.TemplateResponse(
            request,
            "not_found.html",
            {**common_context(request), "paper_id": f"run #{run_id}"},
            status_code=404,
        )
    mismatch = redirect_if_run_not_active(run)
    if mismatch:
        return mismatch
    filter_stats = {}
    raw_filter = run.get("filter_stats_json") if run else None
    if raw_filter:
        try:
            filter_stats = json.loads(raw_filter) if isinstance(raw_filter, str) else dict(raw_filter)
        except (json.JSONDecodeError, TypeError):
            filter_stats = {}
    profiles_doc = load_profiles_doc()
    run_profile_id = run.get("profile") or ""
    run_profile_obj = profiles_doc.get("profiles", {}).get(run_profile_id, {})
    run_profile_name = (
        run_profile_obj.get("display_name")
        or run_profile_obj.get("name")
        or run_profile_id
        or "—"
    )
    latest_for_profile = latest_run_for_profile(run_profile_id) if run_profile_id else latest_run()
    newer_run = None
    if latest_for_profile and int(latest_for_profile.get("id") or 0) > run_id:
        newer_run = latest_for_profile
    stale_empty_run = (
        int(run.get("total_found") or 0) == 0
        and int(run.get("ingested_count") or run.get("new_papers") or 0) == 0
        and run.get("mode") == "daily"
    )
    return templates.TemplateResponse(
        request,
        "run_result.html",
        {
            **common_context(request),
            "run": run,
            "run_candidates": get_run_candidates(run_id),
            "filter_stats": filter_stats,
            "run_profile_name": run_profile_name,
            "newer_run": newer_run,
            "stale_empty_run": stale_empty_run,
        },
    )


@app.post("/run")
def run_tracker(
    profile: str = Form(""),
    mode: str = Form("daily"),
    max_results: int = Form(30),
    time_scope: str = Form("latest"),
    latest_preset: str = Form(""),
    time_range_days: int = Form(7),
    year: str = Form(""),
    start_year: str = Form(""),
    end_year: str = Form(""),
    write_db: str = Form("on"),
    write_excel: str = Form("on"),
    write_markdown: str = Form("on"),
    enrich_metadata: str = Form("on"),
    journal_rank: str = Form("on"),
    auto_rating: str = Form("on"),
    notify_wechat: str = Form(""),
    ingest_policy: str = Form("relevance"),
    dry_run: str = Form(""),
    extra_keywords: str = Form(""),
):
    if WEB_RUN_LOCK.exists():
        try:
            age = time.time() - WEB_RUN_LOCK.stat().st_mtime
        except OSError:
            age = 9999
        if age < 600:
            return RedirectResponse("/runs/waiting?busy=1", status_code=303)
    WEB_RUN_LOCK.write_text(str(time.time()), encoding="utf-8")

    before = latest_run()
    after_id = int((before or {}).get("id") or 0)

    from datetime import datetime as dt

    effective_mode = mode
    if mode not in ("high_quality", "google_scholar"):
        effective_mode = "annual_summary" if time_scope in ("range", "year") else "daily"

    args = [
        sys.executable,
        "-u",
        str(TRACKER_SCRIPT),
        "--mode",
        effective_mode,
        "--max_results",
        str(max_results),
    ]
    if profile:
        args.extend(["--profile", profile])
    if effective_mode == "daily":
        preset = (latest_preset or "").strip()
        if preset == "pure":
            days = 0
        elif preset:
            days = int(preset)
        else:
            days = time_range_days
        args.extend(["--time_range_days", str(days)])
    elif effective_mode == "annual_summary":
        args.extend(["--max_papers_per_year", str(max_results)])
        year_s = str(year or "").strip()
        start_s = str(start_year or "").strip()
        end_s = str(end_year or "").strip()
        current = str(dt.now().year)
        if year_s and time_scope == "year":
            start_s, end_s = year_s, year_s
        if not start_s:
            start_s = "2024"
        if not end_s:
            end_s = current
        try:
            start_i = int(start_s)
            end_i = int(end_s)
        except ValueError:
            start_i, end_i = int(current), int(current)
        if start_i > end_i:
            start_i, end_i = end_i, start_i
        start_s, end_s = str(start_i), str(end_i)
        args.extend(["--start_year", start_s, "--end_year", end_s])
        run_year_label = start_s if start_i == end_i else f"{start_s}-{end_s}"

    run_env = os.environ.copy()
    # Skip per-paper Semantic Scholar calls in Web runs (faster, fewer SSL/rate-limit failures).
    run_env.setdefault("TRACKER_SKIP_S2_METADATA", "1")
    if write_db != "on":
        run_env["TRACKER_SKIP_DB"] = "1"
    if write_excel != "on":
        run_env["TRACKER_SKIP_EXCEL"] = "1"
    if write_markdown != "on":
        run_env["TRACKER_SKIP_MARKDOWN"] = "1"
    if enrich_metadata != "on":
        run_env["TRACKER_SKIP_METADATA"] = "1"
    if journal_rank != "on":
        run_env["TRACKER_SKIP_JOURNAL_RANK"] = "1"
    if auto_rating != "on":
        run_env["TRACKER_SKIP_AUTO_RATING"] = "1"
    if notify_wechat == "on":
        run_env["TRACKER_NOTIFY"] = "serverchan"
    run_env["TRACKER_INGEST_POLICY"] = ingest_policy or "relevance"
    if ingest_policy == "all":
        run_env["TRACKER_NO_RELEVANCE_FILTER"] = "1"
        args.append("--no_relevance_filter")
    run_env["TRACKER_MAX_RESULTS"] = str(max_results)
    if (extra_keywords or "").strip():
        run_env["TRACKER_EXTRA_KEYWORDS"] = extra_keywords.strip()
    if effective_mode == "annual_summary":
        run_env["TRACKER_MAX_PAPERS_PER_YEAR"] = str(max_results)
        run_env["TRACKER_SKIP_AI_PARSE"] = "1"
    if effective_mode == "daily":
        run_env["TRACKER_DATA_SOURCES"] = "arXiv, Semantic Scholar"
    else:
        run_env["TRACKER_DATA_SOURCES"] = "Google Scholar, OpenAlex, Crossref"
        if effective_mode == "annual_summary":
            run_env["TRACKER_RUN_YEAR"] = run_year_label
        else:
            run_env["TRACKER_RUN_YEAR"] = str(year or "").strip() or str(dt.now().year)
    profiles_doc = load_profiles_doc() if PROFILES_PATH.exists() else {}
    prof_obj = profiles_doc.get("profiles", {}).get(profile or profiles_doc.get("default_profile", ""), {})
    if prof_obj.get("google_scholar_query"):
        run_env["TRACKER_GOOGLE_QUERY"] = prof_obj["google_scholar_query"]
    if dry_run == "on":
        args.append("--dry_run")
        run_env["TRACKER_SKIP_EXCEL"] = "1"
        run_env["TRACKER_SKIP_MARKDOWN"] = "1"

    WEB_RUN_LOG.parent.mkdir(parents=True, exist_ok=True)
    log_file = WEB_RUN_LOG.open("a", encoding="utf-8")
    log_file.write("\n\n===== Web run started =====\n")
    log_file.write(" ".join(args) + "\n")
    log_file.flush()
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    subprocess.Popen(
        args,
        cwd=ROOT_DIR,
        stdout=log_file,
        stderr=subprocess.STDOUT,
        env=run_env,
        creationflags=creationflags,
    )
    started = time.time()
    return RedirectResponse(f"/runs/waiting?after={after_id}&since={started}&tries=0", status_code=303)


@app.get("/papers")
def papers(
    request: Request,
    profile: str = "",
    year: str = "",
    level: str = "",
    rating: str = "",
    abstract: str = "",
    has_if: str = "",
    status: str = "",
    q: str = "",
    date: str = "",
    if_min: str = "",
    if_max: str = "",
    citation_min: str = "",
    tier: str = "",
    jcr_quartile: str = "",
    cas_quartile: str = "",
    cas_top: str = "",
    cas_warning: str = "",
    core_tag: str = "",
    journal_matched: str = "",
):
    context = common_context(request)
    raw_profile = profile if profile != "" else request.query_params.get("profile")
    selected_profile = context.get("current_profile_id", "") if raw_profile in (None, "") else raw_profile
    filters = {
        "profile": selected_profile,
        "year": year,
        "level": level,
        "rating": rating,
        "abstract": abstract,
        "has_if": has_if,
        "status": status,
        "q": q,
        "date": date,
        "if_min": if_min,
        "if_max": if_max,
        "citation_min": citation_min,
        "tier": tier,
        "jcr_quartile": jcr_quartile,
        "cas_quartile": cas_quartile,
        "cas_top": cas_top,
        "cas_warning": cas_warning,
        "core_tag": core_tag,
        "journal_matched": journal_matched,
    }
    paper_rows = query_papers(filters, limit=500)
    profile_rows = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    if selected_profile and selected_profile != "__all__":
        years = rows(
            "SELECT DISTINCT year FROM papers WHERE year IS NOT NULL AND profile_id = ? ORDER BY year DESC",
            (selected_profile,),
        )
    else:
        years = rows("SELECT DISTINCT year FROM papers WHERE year IS NOT NULL ORDER BY year DESC")
    profile_total = row(
        "SELECT COUNT(*) AS total FROM papers WHERE profile_id = ?",
        (selected_profile,),
    ) if selected_profile and selected_profile != "__all__" else row("SELECT COUNT(*) AS total FROM papers")
    export_query = export_filters_query_string(filters)
    empty_hint = (
        "当前方向暂无论文，请先获取文献。"
        if (profile_total or {}).get("total") == 0 and selected_profile and selected_profile != "__all__"
        else None
    )
    return templates.TemplateResponse(
        request,
        "papers.html",
        {
            **context,
            "papers": paper_rows,
            "profiles": profile_rows,
            "years": years,
            "filters": filters,
            "profile_total": (profile_total or {}).get("total") or 0,
            "export_query": export_query,
            "empty_profile_hint": empty_hint,
        },
    )


def load_profile_for_paper(paper: dict) -> dict:
    doc = load_profiles_doc()
    profile_id = paper.get("profile_id") or doc.get("default_profile", "")
    profile = doc.get("profiles", {}).get(profile_id, {})
    profile = {**profile, "id": profile_id}
    return profile


def ensure_ai_reading_note(paper: dict, note: dict, force: bool = False) -> tuple[dict, str]:
    if not force and not note_needs_generation(note):
        return note, ""
    if not ai_parse_configured():
        return note, "no_key"
    profile = load_profile_for_paper(paper)
    parsed = parse_paper_with_ai(paper, profile)
    if not parsed:
        return note, "failed"
    payload = parsed_to_reading_note_payload(paper, parsed)
    save_reading_note_for_paper(paper["stable_id"], payload)
    if parsed.get("摘要中文翻译") and is_missing_zh(paper.get("abstract_zh")):
        update_paper_abstract_zh(paper["stable_id"], parsed["摘要中文翻译"])
    refreshed = row("SELECT * FROM reading_notes WHERE paper_id = ?", (paper["stable_id"],)) or {}
    return refreshed, "generated"


def build_pdf_parse_prompt(paper: dict, profile: dict, pdf_text: str) -> str:
    focus_lines = "\n".join(f"- {item}" for item in profile.get("parse_focus", []))
    fallback = f"""你是科研人员的论文精读助手。请基于上传 PDF 抽取文本完成结构化中文精读。
要求：只根据提供文本，不要编造；文本缺失时明确写“PDF 文本未覆盖，需人工确认”；输出要可直接转成 Markdown 笔记。

【研究方向】
{profile.get('display_name') or profile.get('name') or profile.get('id') or '当前方向'}
{focus_lines}

【论文信息】
标题：{paper.get('title') or ''}
期刊：{paper.get('journal') or ''}
年份：{paper.get('year') or ''}
DOI：{paper.get('doi') or ''}
原文链接：{paper_display_url(paper)}

【PDF 正文节选】
{pdf_text[:24000]}

请严格按以下字段输出，每段以【字段名】：开头：
【摘要中文翻译】：
【研究背景】：
【研究目的】：
【核心方法】：
【论文创新点】：
【实验结果】：
【总结】：
【未来展望】：
【可创新点】：
【对我的启发】：
【是否值得精读】：
"""
    return render_template(
        "full_paper_reading_prompt",
        {
            "profile_name": profile.get("display_name") or profile.get("name") or profile.get("id") or "当前方向",
            "profile_focus": focus_lines,
            "title": paper.get("title") or "",
            "journal": paper.get("journal") or "",
            "year": paper.get("year") or "",
            "doi": paper.get("doi") or "",
            "link": paper_display_url(paper),
            "pdf_text": pdf_text[:24000],
        },
        fallback,
    )


def ensure_pdf_reading_note(paper: dict, pdf_text: str) -> tuple[dict, str]:
    if not ai_parse_configured():
        if not load_my_notes(paper):
            save_my_notes(paper, default_my_notes_template(paper, {}))
        return {}, "template"
    profile = load_profile_for_paper(paper)
    raw = call_llm(build_pdf_parse_prompt(paper, profile, pdf_text))
    parsed = parse_structured_fields(raw or "")
    if not parsed:
        if not load_my_notes(paper):
            save_my_notes(paper, default_my_notes_template(paper, {}))
        return {}, "failed"
    payload = parsed_to_reading_note_payload(paper, parsed)
    save_reading_note_for_paper(paper["stable_id"], payload)
    if parsed.get("摘要中文翻译") and is_missing_zh(paper.get("abstract_zh")):
        update_paper_abstract_zh(paper["stable_id"], parsed["摘要中文翻译"])
    refreshed = row("SELECT * FROM reading_notes WHERE paper_id = ?", (paper["stable_id"],)) or {}
    if not load_my_notes(paper):
        save_my_notes(paper, default_my_notes_template(paper, refreshed))
    return refreshed, "generated"


@app.get("/papers/{paper_id}")
def paper_detail(request: Request, paper_id: str):
    paper = row(f"SELECT *, {RATING_SQL} AS display_rating FROM papers WHERE stable_id = ?", (paper_id,))
    if not paper:
        return templates.TemplateResponse(request, "not_found.html", {**common_context(request), "paper_id": paper_id}, status_code=404)
    paper["computed_system_rating"] = paper.get("system_rating") or score_to_rating(paper.get("final_score") or paper.get("relevance_score"))
    previous_zh = paper.get("abstract_zh")
    paper = ensure_paper_chinese_abstract(paper)
    if is_missing_zh(previous_zh) and not is_missing_zh(paper.get("abstract_zh")):
        update_paper_abstract_zh(paper_id, paper.get("abstract_zh", ""))
    note = row("SELECT * FROM reading_notes WHERE paper_id = ?", (paper_id,)) or {}
    auto_generate = request.query_params.get("generate", "1") != "0"
    ai_status = ""
    if auto_generate:
        note, ai_status = ensure_ai_reading_note(paper, note)
        if ai_status == "generated":
            paper = row(f"SELECT *, {RATING_SQL} AS display_rating FROM papers WHERE stable_id = ?", (paper_id,)) or paper
            paper["computed_system_rating"] = paper.get("system_rating") or score_to_rating(paper.get("final_score") or paper.get("relevance_score"))
    my_notes = load_my_notes(paper)
    if request.query_params.get("notes_saved") == "1":
        my_notes = load_my_notes(paper)
    return templates.TemplateResponse(
        request,
        "paper_detail.html",
        {
            **common_context(request),
            "paper": paper,
            "note": note,
            "ai_status": ai_status,
            "my_notes": my_notes,
            "my_notes_saved": request.query_params.get("notes_saved") == "1",
        },
    )


@app.post("/papers/{paper_id}/generate-notes")
def generate_paper_notes(paper_id: str):
    paper = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,))
    if not paper:
        return RedirectResponse("/papers", status_code=303)
    note = row("SELECT * FROM reading_notes WHERE paper_id = ?", (paper_id,)) or {}
    ensure_ai_reading_note(paper, note, force=True)
    rag_index_paper(paper_id)
    return RedirectResponse(f"/papers/{paper_id}?generated=1", status_code=303)


@app.get("/papers/{paper_id}/ask")
def paper_ask_page(request: Request, paper_id: str, q: str = ""):
    paper = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,))
    if not paper:
        return RedirectResponse("/papers", status_code=303)
    result = answer_with_rag(q, paper.get("profile_id") or "", "paper", paper_id=paper_id) if q.strip() else {}
    return templates.TemplateResponse(
        request,
        "ask.html",
        {
            **common_context(request),
            "profiles": rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id"),
            "selected_profile": paper.get("profile_id") or "",
            "question": q,
            "result": result,
            "rag_status": rag_status(),
            "paper": paper,
            "scope_label": "问这篇论文",
        },
    )


@app.post("/papers/{paper_id}/ask")
def paper_ask_action(request: Request, paper_id: str, question: str = Form("")):
    paper = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,))
    if not paper:
        return RedirectResponse("/papers", status_code=303)
    result = answer_with_rag(question, paper.get("profile_id") or "", "paper", paper_id=paper_id) if question.strip() else {}
    return templates.TemplateResponse(
        request,
        "ask.html",
        {
            **common_context(request),
            "profiles": rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id"),
            "selected_profile": paper.get("profile_id") or "",
            "question": question,
            "result": result,
            "rag_status": rag_status(),
            "paper": paper,
            "scope_label": "问这篇论文",
        },
    )


@app.post("/papers/{paper_id}/read")
def mark_read(paper_id: str):
    execute("UPDATE papers SET reading_status = 'read', updated_at = datetime('now') WHERE stable_id = ?", (paper_id,))
    return RedirectResponse(f"/papers/{paper_id}", status_code=303)


@app.post("/papers/{paper_id}/my-notes")
def save_my_notes_action(paper_id: str, content: str = Form(""), use_template: str = Form("")):
    paper = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,))
    if not paper:
        return RedirectResponse("/papers", status_code=303)
    note = row("SELECT * FROM reading_notes WHERE paper_id = ?", (paper_id,)) or {}
    if use_template == "1" and not content.strip():
        content = default_my_notes_template(paper, note)
    elif use_template == "ai" and not content.strip():
        content = default_my_notes_template(paper, note)
        ai_block = "\n".join([
            "## AI 阅读摘要（自动插入）",
            f"**主题**：{note.get('paper_topic') or ''}",
            f"**方法**：{note.get('core_method') or ''}",
            f"**贡献**：{note.get('paper_contribution') or ''}",
            f"**启发**：{note.get('inspiration') or ''}",
            f"**可创新点**：{note.get('possible_ideas') or ''}",
        ])
        content = content.rstrip() + "\n\n" + ai_block + "\n"
    save_my_notes(paper, content)
    rag_index_paper(paper_id)
    return RedirectResponse(f"/papers/{paper_id}?notes_saved=1#my-notes", status_code=303)


@app.post("/papers/{paper_id}/favorite")
def toggle_favorite(paper_id: str):
    execute(
        """
        UPDATE papers
        SET favorite = CASE WHEN favorite = 1 THEN 0 ELSE 1 END,
            is_favorite = CASE WHEN favorite = 1 THEN 0 ELSE 1 END,
            updated_at = datetime('now')
        WHERE stable_id = ?
        """,
        (paper_id,),
    )
    return RedirectResponse(f"/papers/{paper_id}", status_code=303)


@app.get("/papers/{paper_id}/export")
def export_note(paper_id: str):
    paper = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,))
    note = row("SELECT * FROM reading_notes WHERE paper_id = ?", (paper_id,)) or {}
    if not paper:
        return RedirectResponse("/papers", status_code=303)
    profile = paper.get("profile_id") or "default"
    note_dir = PAPER_NOTES_DIR / safe_slug(profile)
    note_dir.mkdir(parents=True, exist_ok=True)
    path = note_dir / f"{paper.get('year') or 'unknown'}_{safe_slug(paper['title'])}.md"
    content = [
        f"# {paper['title']}",
        "",
        f"- 期刊：{paper.get('journal') or ''}",
        f"- 年份：{paper.get('year') or ''}",
        f"- DOI：{paper.get('doi') or ''}",
        f"- 链接：{paper_display_url(paper)}",
        f"- PDF：{paper.get('pdf_url') or ''}",
        f"- DOI：{paper.get('doi_url') or paper.get('doi') or ''}",
        f"- 推荐等级：{paper.get('recommendation_level') or ''}",
        f"- 综合分：{paper.get('final_score') or 0}",
        "",
        "## 原始摘要",
        "",
        paper.get("abstract_original") or "",
        "",
        "## AI 摘要阅读",
        "",
        f"### 论文主题\n{note.get('paper_topic') or ''}",
        f"### 研究背景\n{note.get('research_background') or ''}",
        f"### 研究目的\n{note.get('research_purpose') or ''}",
        f"### 核心方法\n{note.get('core_method') or ''}",
        f"### 主要结果\n{note.get('main_results') or ''}",
        f"### 论文贡献\n{note.get('paper_contribution') or ''}",
        f"### 对我的启发\n{note.get('inspiration') or ''}",
        f"### 是否值得精读\n{note.get('worth_reading') or ''}",
        f"### 可创新点\n{note.get('possible_ideas') or ''}",
    ]
    path.write_text("\n\n".join(content), encoding="utf-8")
    return FileResponse(path, filename=path.name)


@app.post("/profiles/set-default")
def set_default_profile(profile_id: str = Form(...)):
    if not activate_profile(profile_id):
        return RedirectResponse("/profiles", status_code=303)
    return RedirectResponse("/?switched=1", status_code=303)


@app.get("/profiles")
def profiles_page(request: Request, selected: str = ""):
    doc = load_profiles_doc()
    profile_map = doc.get("profiles", {})
    default_profile = doc.get("default_profile", "")
    selected = selected or default_profile or next(iter(profile_map), "")
    if selected and selected in profile_map and selected != default_profile:
        activate_profile(selected)
        return RedirectResponse("/?switched=1", status_code=303)
    profile = profile_map.get(selected, {})
    return templates.TemplateResponse(
        request,
        "profiles.html",
        {
            **common_context(request),
            "profiles": profile_map,
            "selected": selected,
            "profile": profile,
            "is_new": False,
            "default_profile": default_profile,
        },
    )


@app.get("/profiles/new")
def profiles_new_page(request: Request):
    doc = load_profiles_doc()
    profile_map = doc.get("profiles", {})
    return templates.TemplateResponse(
        request,
        "profiles.html",
        {
            **common_context(request),
            "profiles": profile_map,
            "selected": "",
            "profile": {
                "include_keywords": [],
                "must_have_any": [],
                "exclude_keywords": [],
                "research_focus": [],
                "arxiv_categories": ["physics.app-ph", "cond-mat.mtrl-sci"],
                "default_sources": {"arxiv": True, "semantic_scholar": True, "google_scholar": False},
                "score_rules": {"min_score": 3},
            },
            "is_new": True,
            "default_profile": doc.get("default_profile", ""),
            "can_delete": False,
        },
    )


def _profile_form_values(
    profile_id: str = Form(...),
    name: str = Form(""),
    display_name: str = Form(""),
    description: str = Form(""),
    include_keywords: str = Form(""),
    must_have_any: str = Form(""),
    exclude_keywords: str = Form(""),
    research_focus: str = Form(""),
    google_scholar_query: str = Form(""),
    arxiv_categories: str = Form(""),
    min_score: int = Form(3),
    enable_wechat: str = Form(""),
    use_arxiv: str = Form(""),
    use_semantic: str = Form(""),
    use_google: str = Form(""),
    set_as_default: str = Form(""),
    ingest_min_score: int = Form(1),
    ingest_below_must_have: str = Form(""),
):
    return {
        "profile_id": slugify_profile_id(profile_id),
        "name": name.strip(),
        "display_name": display_name.strip(),
        "description": description.strip(),
        "include_keywords": include_keywords,
        "must_have_any": must_have_any,
        "exclude_keywords": exclude_keywords,
        "research_focus": research_focus,
        "google_scholar_query": google_scholar_query.strip(),
        "arxiv_categories": arxiv_categories,
        "min_score": min_score,
        "enable_wechat": enable_wechat == "on",
        "use_arxiv": use_arxiv == "on",
        "use_semantic": use_semantic == "on",
        "use_google": use_google == "on",
        "set_as_default": set_as_default == "on",
        "ingest_min_score": ingest_min_score,
        "ingest_below_must_have": ingest_below_must_have == "on",
    }


@app.post("/profiles/create")
def create_profile(
    profile_id: str = Form(...),
    name: str = Form(""),
    display_name: str = Form(""),
    description: str = Form(""),
    include_keywords: str = Form(""),
    must_have_any: str = Form(""),
    exclude_keywords: str = Form(""),
    research_focus: str = Form(""),
    google_scholar_query: str = Form(""),
    arxiv_categories: str = Form(""),
    min_score: int = Form(3),
    enable_wechat: str = Form(""),
    use_arxiv: str = Form("on"),
    use_semantic: str = Form("on"),
    use_google: str = Form(""),
    set_as_default: str = Form(""),
    ingest_min_score: int = Form(1),
    ingest_below_must_have: str = Form(""),
):
    values = _profile_form_values(
        profile_id,
        name,
        display_name,
        description,
        include_keywords,
        must_have_any,
        exclude_keywords,
        research_focus,
        google_scholar_query,
        arxiv_categories,
        min_score,
        enable_wechat,
        use_arxiv,
        use_semantic,
        use_google,
        set_as_default,
        ingest_min_score,
        ingest_below_must_have,
    )
    profile_id = values.pop("profile_id")
    set_as_default = values.pop("set_as_default")
    doc = load_profiles_doc()
    profiles = doc.setdefault("profiles", {})
    if profile_id in profiles:
        return RedirectResponse(f"/profiles/new?error=exists", status_code=303)
    profiles[profile_id] = build_profile_payload(**values, profile_id=profile_id, old={})
    doc["default_profile"] = profile_id
    save_profiles_doc(doc)
    return RedirectResponse(f"/profiles?selected={profile_id}&saved=1", status_code=303)


@app.post("/profiles")
def save_profile(
    profile_id: str = Form(...),
    name: str = Form(""),
    display_name: str = Form(""),
    description: str = Form(""),
    include_keywords: str = Form(""),
    must_have_any: str = Form(""),
    exclude_keywords: str = Form(""),
    research_focus: str = Form(""),
    google_scholar_query: str = Form(""),
    arxiv_categories: str = Form(""),
    min_score: int = Form(3),
    enable_wechat: str = Form(""),
    use_arxiv: str = Form(""),
    use_semantic: str = Form(""),
    use_google: str = Form(""),
    set_as_default: str = Form(""),
    ingest_min_score: int = Form(1),
    ingest_below_must_have: str = Form(""),
):
    values = _profile_form_values(
        profile_id,
        name,
        display_name,
        description,
        include_keywords,
        must_have_any,
        exclude_keywords,
        research_focus,
        google_scholar_query,
        arxiv_categories,
        min_score,
        enable_wechat,
        use_arxiv,
        use_semantic,
        use_google,
        set_as_default,
        ingest_min_score,
        ingest_below_must_have,
    )
    profile_id = values.pop("profile_id")
    set_as_default = values.pop("set_as_default")
    doc = load_profiles_doc()
    profiles = doc.setdefault("profiles", {})
    if profile_id not in profiles:
        return RedirectResponse("/profiles/new?error=missing", status_code=303)
    profiles[profile_id] = build_profile_payload(**values, profile_id=profile_id, old=profiles.get(profile_id, {}))
    doc["default_profile"] = profile_id
    save_profiles_doc(doc)
    return RedirectResponse(f"/profiles?selected={profile_id}&saved=1", status_code=303)


@app.post("/profiles/delete")
def delete_profile(profile_id: str = Form(...)):
    profile_id = slugify_profile_id(profile_id)
    doc = load_profiles_doc()
    profiles = doc.get("profiles", {})
    if len(profiles) <= 1:
        return RedirectResponse("/profiles?error=last", status_code=303)
    if profile_id == doc.get("default_profile"):
        return RedirectResponse(f"/profiles?selected={profile_id}&error=default", status_code=303)
    if profile_id in profiles:
        del profiles[profile_id]
        if doc.get("default_profile") not in profiles:
            doc["default_profile"] = next(iter(profiles))
        save_profiles_doc(doc)
    return RedirectResponse(f"/profiles?selected={doc.get('default_profile', '')}", status_code=303)


@app.post("/settings/advanced-features")
def save_advanced_features(advanced_features_enabled: str = Form("")):
    save_app_settings({"advanced_features_enabled": advanced_features_enabled == "on"})
    return RedirectResponse("/settings?saved=advanced", status_code=303)


@app.get("/settings")
def settings_page(request: Request):
    app_settings = load_app_settings()
    return templates.TemplateResponse(
        request,
        "settings.html",
        {
            **common_context(request),
            "app_settings": app_settings,
            "serpapi_configured": bool(os.environ.get("SERPAPI_API_KEY")),
            "sct_configured": bool(os.environ.get("SCT_KEY")),
            "semantic_configured": bool(os.environ.get("SEMANTIC_SCHOLAR_API_KEY")),
            "kimi_configured": bool(os.environ.get("KIMI_API_KEY") or os.environ.get("MOONSHOT_API_KEY")),
            "openai_configured": bool(os.environ.get("OPENAI_API_KEY")),
            "anthropic_configured": bool(os.environ.get("ANTHROPIC_API_KEY")),
            "serpapi_masked": mask_key(os.environ.get("SERPAPI_API_KEY", "")),
            "sct_masked": mask_key(os.environ.get("SCT_KEY", "")),
            "semantic_masked": mask_key(os.environ.get("SEMANTIC_SCHOLAR_API_KEY", "")),
            "kimi_masked": mask_key(os.environ.get("KIMI_API_KEY", "") or os.environ.get("MOONSHOT_API_KEY", "")),
            "openai_masked": mask_key(os.environ.get("OPENAI_API_KEY", "")),
            "anthropic_masked": mask_key(os.environ.get("ANTHROPIC_API_KEY", "")),
            "paper_ai_provider": os.environ.get("PAPER_AI_PROVIDER", "kimi" if os.environ.get("KIMI_API_KEY") else ""),
            "kimi_parse_model": os.environ.get("KIMI_PARSE_MODEL", "moonshot-v1-32k"),
            "journal_metrics_exists": JOURNAL_METRICS_PATH.exists(),
            "journal_metrics_path": JOURNAL_METRICS_PATH,
            "journal_metrics_stats": _journal_metrics_dashboard_stats(),
            "journal_example_path": SKILL_DIR / "journal_metrics.example.csv",
            "excel_path": OUTPUT_DIR / "Ptychography_论文全量库.xlsx",
            "daily_report_path": OUTPUT_DIR / "daily_reports",
            "pdfs_path": PDFS_DIR,
            "db_path": DB_PATH,
            "doctor_checks": run_doctor_checks(),
            "prompt_templates": prompt_status(),
            "prompt_runs": recent_prompt_runs(8),
            "rag_status": rag_status(),
            "agent_runs": recent_agent_runs(8),
        },
    )


@app.post("/settings/api-keys")
def save_api_keys(
    serpapi_key: str = Form(""),
    sct_key: str = Form(""),
    semantic_scholar_key: str = Form(""),
    kimi_key: str = Form(""),
    paper_ai_provider: str = Form(""),
    kimi_model: str = Form(""),
    openai_key: str = Form(""),
    anthropic_key: str = Form(""),
):
    save_env_values({
        "SERPAPI_API_KEY": serpapi_key,
        "SCT_KEY": sct_key,
        "SEMANTIC_SCHOLAR_API_KEY": semantic_scholar_key,
        "KIMI_API_KEY": kimi_key,
        "PAPER_AI_PROVIDER": paper_ai_provider,
        "KIMI_PARSE_MODEL": kimi_model,
        "OPENAI_API_KEY": openai_key,
        "ANTHROPIC_API_KEY": anthropic_key,
    })
    return RedirectResponse("/settings?saved=1", status_code=303)


@app.get("/daily")
def daily_page(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    run_rows = rows(
        "SELECT * FROM runs WHERE (? = '' OR profile = ?) ORDER BY run_time DESC, id DESC LIMIT 30",
        (profile_id, profile_id),
    )
    return templates.TemplateResponse(request, "daily.html", {**context, "runs": run_rows})


@app.get("/daily/run")
def daily_run_page(request: Request):
    profiles = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    return templates.TemplateResponse(request, "daily_run.html", {**common_context(request), "profiles": profiles})


@app.get("/annual")
def annual_index():
    return RedirectResponse(f"/annual/{available_years()[0]}", status_code=303)


@app.get("/annual/{year}")
def annual_page(
    request: Request,
    year: int,
    profile: str = "",
    rating: str = "",
    abstract: str = "",
    if_min: str = "",
    citation_min: str = "",
    status: str = "",
    favorite: str = "",
    included: str = "",
    source: str = "",
    q: str = "",
    generated: str = "",
    jcr_quartile: str = "",
    cas_quartile: str = "",
    cas_top: str = "",
    cas_warning: str = "",
    core_tag: str = "",
    journal_matched: str = "",
    if_max: str = "",
):
    ctx = common_context(request)
    default_profile = ctx.get("current_profile_id", "")
    filters = {
        "profile": profile if profile else default_profile,
        "rating": rating,
        "abstract": abstract,
        "if_min": if_min,
        "if_max": if_max,
        "citation_min": citation_min,
        "status": status,
        "favorite": favorite,
        "included": included,
        "source": source,
        "q": q,
        "jcr_quartile": jcr_quartile,
        "cas_quartile": cas_quartile,
        "cas_top": cas_top,
        "cas_warning": cas_warning,
        "core_tag": core_tag,
        "journal_matched": journal_matched,
    }
    paper_rows = annual_papers(year, filters)
    profiles = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    return templates.TemplateResponse(
        request,
        "annual.html",
        {
            **common_context(request),
            "year": year,
            "years": available_years(),
            "profiles": profiles,
            "papers": paper_rows,
            "annual_stats": annual_stats(year, filters),
            "filters": filters,
            "generated": generated,
        },
    )


@app.post("/annual/{year}/run")
def annual_run(year: int, profile: str = Form(""), max_results: int = Form(200)):
    if WEB_RUN_LOCK.exists():
        try:
            age = time.time() - WEB_RUN_LOCK.stat().st_mtime
        except OSError:
            age = 9999
        if age < 600:
            return RedirectResponse("/runs/waiting?busy=1", status_code=303)
    WEB_RUN_LOCK.write_text(str(time.time()), encoding="utf-8")
    before = latest_run()
    after_id = int((before or {}).get("id") or 0)
    args = [
        sys.executable,
        str(TRACKER_SCRIPT),
        "--mode",
        "annual_summary",
        "--start_year",
        str(year),
        "--end_year",
        str(year),
        "--max_papers_per_year",
        str(max_results),
    ]
    if profile:
        args.extend(["--profile", profile])
    run_env = os.environ.copy()
    run_env.setdefault("TRACKER_SKIP_S2_METADATA", "1")
    WEB_RUN_LOG.parent.mkdir(parents=True, exist_ok=True)
    log_file = WEB_RUN_LOG.open("a", encoding="utf-8")
    log_file.write(f"\n\n===== Annual run {year} started =====\n")
    log_file.write(" ".join(args) + "\n")
    log_file.flush()
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    subprocess.Popen(args, cwd=ROOT_DIR, stdout=log_file, stderr=subprocess.STDOUT, env=run_env, creationflags=creationflags)
    started = time.time()
    return RedirectResponse(f"/runs/waiting?after={after_id}&since={started}&tries=0", status_code=303)


@app.post("/annual/{year}/markdown")
def annual_markdown(year: int, profile: str = Form(""), rating: str = Form(""), abstract: str = Form(""), if_min: str = Form(""), citation_min: str = Form(""), status: str = Form(""), favorite: str = Form(""), included: str = Form(""), source: str = Form(""), q: str = Form("")):
    filters = {"profile": profile, "rating": rating, "abstract": abstract, "if_min": if_min, "citation_min": citation_min, "status": status, "favorite": favorite, "included": included, "source": source, "q": q}
    paper_rows = annual_papers(year, filters, limit=2000)
    stats = annual_stats(year, filters)
    profile_label = profile or "all_profiles"
    profile_slug = safe_slug(profile_label)
    content = build_annual_markdown(year, profile_label, stats, paper_rows)
    path = ANNUAL_REPORTS_DIR / f"{year}_{profile_slug}_annual.md"
    write_markdown(path, content)
    return RedirectResponse(f"/annual/{year}?generated=markdown", status_code=303)


@app.post("/annual/{year}/excel")
def annual_excel(year: int, profile: str = Form(""), rating: str = Form(""), abstract: str = Form(""), if_min: str = Form(""), citation_min: str = Form(""), status: str = Form(""), favorite: str = Form(""), included: str = Form(""), source: str = Form(""), q: str = Form("")):
    filters = {"profile": profile, "rating": rating, "abstract": abstract, "if_min": if_min, "citation_min": citation_min, "status": status, "favorite": favorite, "included": included, "source": source, "q": q}
    paper_rows = annual_papers(year, filters, limit=2000)
    export_annual_excel(year, paper_rows, safe_slug(profile or "all_profiles"))
    return RedirectResponse(f"/annual/{year}?generated=excel", status_code=303)


@app.post("/papers/{paper_id}/rating")
def update_paper_rating(paper_id: str, rating: int = Form(...), return_to: str = Form("")):
    rating = max(1, min(int(rating), 5))
    execute("UPDATE papers SET user_rating = ?, updated_at = datetime('now') WHERE stable_id = ?", (rating, paper_id))
    return RedirectResponse(return_to or f"/papers/{paper_id}", status_code=303)


@app.post("/papers/{paper_id}/status")
def update_paper_status(paper_id: str, reading_status: str = Form("unread"), return_to: str = Form("")):
    allowed = {"unread", "todo", "reading", "read", "ignored"}
    if reading_status not in allowed:
        reading_status = "unread"
    execute("UPDATE papers SET reading_status = ?, updated_at = datetime('now') WHERE stable_id = ?", (reading_status, paper_id))
    return RedirectResponse(return_to or f"/papers/{paper_id}", status_code=303)


@app.post("/papers/{paper_id}/review-toggle")
def toggle_review_inclusion(paper_id: str, return_to: str = Form("")):
    execute(
        """
        UPDATE papers
        SET included_in_review = CASE WHEN included_in_review = 1 THEN 0 ELSE 1 END,
            updated_at = datetime('now')
        WHERE stable_id = ?
        """,
        (paper_id,),
    )
    return RedirectResponse(return_to or f"/papers/{paper_id}", status_code=303)


@app.get("/weekly")
def weekly_page(request: Request):
    if not advanced_features_enabled():
        return RedirectResponse("/reports", status_code=307)
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    papers = papers_ingested_recent(7, profile_id, limit=10)
    context["dashboard_numbers"]["weekly_new"] = len(papers_ingested_recent(7, profile_id, limit=500))
    keywords = (context["current_profile"].get("research_focus") or [])[:10]
    return templates.TemplateResponse(request, "weekly.html", {**context, "papers": papers, "keywords": keywords})


@app.get("/weekly/generate")
def weekly_generate_page(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    papers = papers_ingested_recent(7, profile_id, limit=10)
    context["dashboard_numbers"]["weekly_new"] = len(papers_ingested_recent(7, profile_id, limit=500))
    keywords = (context["current_profile"].get("research_focus") or [])[:10]
    return templates.TemplateResponse(request, "weekly_generate.html", {**context, "papers": papers, "keywords": keywords})


@app.post("/weekly/generate")
def weekly_generate_action(request: Request):
    context = common_context(request)
    profile_id = context.get("current_profile_id", "")
    papers = papers_ingested_recent(7, profile_id, limit=10)
    context["dashboard_numbers"]["weekly_new"] = len(papers_ingested_recent(7, profile_id, limit=500))
    keywords = (context["current_profile"].get("research_focus") or [])[:10]
    profile_slug = safe_slug(profile_id or "research")
    path = WEEKLY_REPORTS_DIR / f"{week_string()}_{profile_slug}_weekly.md"
    content = build_weekly_report_markdown(context, papers, keywords)
    write_markdown(path, content)
    return templates.TemplateResponse(
        request,
        "weekly_generate.html",
        {**context, "papers": papers, "keywords": keywords, "generated_path": path, "markdown_preview": content},
    )


@app.get("/reading")
def reading_page(request: Request):
    return RedirectResponse("/papers", status_code=307)


@app.get("/reading-ideas")
def reading_ideas_page(request: Request):
    return RedirectResponse("/papers", status_code=307)


@app.get("/generate/review")
def generate_review_page(request: Request):
    if not advanced_features_enabled():
        return RedirectResponse("/reports", status_code=307)
    context = common_context(request)
    profiles = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    return templates.TemplateResponse(
        request,
        "generate_review.html",
        {
            **context,
            "profiles": profiles,
            "form": {
                "topic": "Electron Ptychography in 4D-STEM: Methods, Applications, and Challenges",
                "profile": context.get("current_profile_id", ""),
                "review_type": "领域进展综述",
                "language": "中文",
                "time_range": "5y",
                "level_filter": "b_plus",
                "abstract_filter": "complete",
                "if_filter": "matched",
                "reading_status": "",
                "max_papers": 50,
            },
        },
    )


@app.post("/generate/review")
def generate_review_action(
    request: Request,
    topic: str = Form(""),
    profile: str = Form(""),
    review_type: str = Form("领域进展综述"),
    language: str = Form("中文"),
    time_range: str = Form("5y"),
    level_filter: str = Form("b_plus"),
    abstract_filter: str = Form("complete"),
    if_filter: str = Form("matched"),
    reading_status: str = Form(""),
    max_papers: int = Form(50),
):
    ensure_review_jobs_table()
    context = common_context(request)
    profiles = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    topic = topic.strip() or "Untitled Research Review"
    papers = selected_review_papers(profile, time_range, level_filter, abstract_filter, if_filter, reading_status, max_papers)
    form = {
        "topic": topic,
        "profile": profile,
        "review_type": review_type,
        "language": language,
        "time_range": time_range,
        "level_filter": level_filter,
        "abstract_filter": abstract_filter,
        "if_filter": if_filter,
        "reading_status": reading_status,
        "max_papers": max_papers,
    }
    if len(papers) < 5:
        execute(
            """
            INSERT INTO review_jobs (profile, topic, review_type, language, paper_count, status, output_path, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
            """,
            (profile, topic, review_type, language, len(papers), "文献不足", ""),
        )
        return templates.TemplateResponse(
            request,
            "generate_review.html",
            {
                **context,
                "profiles": profiles,
                "form": form,
                "paper_count": len(papers),
                "warning": "当前满足条件的论文不足 5 篇，建议先运行 high_quality 模式，或放宽筛选条件。",
            },
        )

    REVIEW_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    profile_slug = safe_slug(profile or context.get("current_profile_id") or "research")
    topic_slug = safe_slug(topic, limit=46)
    base = REVIEW_REPORTS_DIR / f"{today_string()}_{profile_slug}_{topic_slug}"
    outline = build_review_outline(topic, review_type, language, papers)
    draft = build_review_draft(topic, review_type, language, papers)
    kimi_body = build_kimi_review_body(topic, review_type, language, papers)
    if kimi_body:
        draft = draft + "\n\n## Kimi 综述正文（初稿）\n\n" + kimi_body
    outline_path = write_markdown(base.with_name(base.name + "_review_outline.md"), outline)
    draft_path = write_markdown(base.with_name(base.name + "_review_draft.md"), draft)
    table_path = export_literature_table(papers, base.with_name(base.name + "_literature_table"))
    execute(
        """
        INSERT INTO review_jobs (profile, topic, review_type, language, paper_count, status, output_path, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
        """,
        (profile, topic, review_type, language, len(papers), "已生成", str(draft_path)),
    )
    return templates.TemplateResponse(
        request,
        "generate_review.html",
        {
            **context,
            "profiles": profiles,
            "form": form,
            "papers": papers[:10],
            "paper_count": len(papers),
            "outline_path": outline_path,
            "draft_path": draft_path,
            "table_path": table_path,
            "outline_preview": outline,
        },
    )


@app.post("/generate/review/from-rag")
def generate_review_from_rag(request: Request, question: str = Form("当前研究方向综述"), profile: str = Form("")):
    context = common_context(request)
    profile_id = profile or context.get("current_profile_id", "")
    chunks = rag_search(question, profile=profile_id, limit=12)
    profile_obj = context.get("current_profile", {})
    fallback = f"""请基于以下 RAG 证据生成综述资料包初稿。

主题：{question}
研究方向：{profile_obj.get('display_name') or profile_obj.get('name') or profile_id}
证据：
{format_rag_context(chunks)}
"""
    prompt = render_template(
        "review_writing_prompt",
        {
            "language": "中文",
            "review_type": "基于 RAG 的领域进展综述",
            "topic": question,
            "paper_context": format_rag_context(chunks),
        },
        fallback,
    )
    body = call_llm(prompt) if ai_parse_configured() and chunks else ""
    if not body:
        body = "未能生成 AI 综述。请先在「问文献库」重建索引，或配置 AI Key。\n\n## 已检索依据\n" + format_rag_context(chunks)
    REVIEW_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = REVIEW_REPORTS_DIR / f"{today_string()}_{safe_slug(profile_id)}_rag_review.md"
    write_markdown(path, body)
    return templates.TemplateResponse(
        request,
        "generate_review.html",
        {
            **context,
            "profiles": rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id"),
            "draft_path": path,
            "outline_path": path,
            "outline_preview": body,
            "draft_preview": body,
            "table_path": "",
            "paper_count": len(chunks),
            "form": {
                "topic": question,
                "profile": profile_id,
                "review_type": "基于 RAG 的领域进展综述",
                "language": "中文",
                "time_range": "all",
                "level_filter": "b_plus",
                "abstract_filter": "all",
                "if_filter": "all",
                "reading_status": "",
                "max_papers": 30,
            },
            "papers": [],
        },
    )


@app.get("/generate/daily-report")
def generate_daily_report_page(request: Request):
    context = common_context(request)
    papers = daily_report_candidates(context.get("current_profile_id", ""), limit=3)
    return templates.TemplateResponse(request, "generate_daily_report.html", {**context, "papers": papers})


@app.post("/generate/daily-report")
def generate_daily_report_action(request: Request):
    context = common_context(request)
    papers = daily_report_candidates(context.get("current_profile_id", ""), limit=3)
    profile_id = safe_slug(context.get("current_profile_id", "research"))
    path = OUTPUT_DIR / "daily_reports" / f"{today_string()}_{profile_id}_daily.md"
    content = build_daily_report_markdown(context, papers)
    write_markdown(path, content)
    return templates.TemplateResponse(
        request,
        "generate_daily_report.html",
        {**context, "papers": papers, "generated_path": path, "markdown_preview": content},
    )


@app.get("/roadmap")
def roadmap_page(request: Request, profile: str = "", view: str = "timeline", node: str = "", stage: str = ""):
    if view == "tree":
        from urllib.parse import urlencode
        params = {"profile": profile or "", "view": "branches", "node": node, "stage": stage}
        return RedirectResponse(f"/roadmap?{urlencode({k: v for k, v in params.items() if v})}", status_code=301)
    return templates.TemplateResponse(
        request,
        "roadmap.html",
        roadmap_page_context(request, profile, view, node, stage),
    )


@app.get("/roadmap/stage/{node_key}")
def roadmap_stage_json(request: Request, node_key: str, profile: str = ""):
    context = common_context(request)
    profile_id = profile or context.get("current_profile_id", "")
    roadmap = build_roadmap_view_model(profile_id)
    stage = next((s for s in roadmap.get("stages") or [] if s.get("node_key") == node_key), None)
    if not stage:
        return JSONResponse({"error": "not found"}, status_code=404)
    return JSONResponse(stage)


@app.post("/roadmap/generate-map")
def roadmap_generate_map(profile: str = Form("")):
    profile_id = profile or load_profiles_doc().get("default_profile", "")
    stats = generate_roadmap_map(profile_id)
    return RedirectResponse(
        f"/roadmap?profile={profile_id}&view=timeline&map_generated=1&candidates={stats.get('candidates', 0)}",
        status_code=303,
    )


@app.post("/roadmap/paper-status")
def roadmap_paper_status(paper_id: str = Form(...), reading_status: str = Form("todo"), return_to: str = Form("/roadmap")):
    allowed = {"unread", "todo", "reading", "read", "ignored"}
    if reading_status not in allowed:
        reading_status = "todo"
    execute("UPDATE papers SET reading_status = ?, updated_at = datetime('now') WHERE stable_id = ?", (reading_status, paper_id))
    return RedirectResponse(return_to or "/roadmap", status_code=303)


@app.post("/roadmap/generate")
def roadmap_generate_action(request: Request, profile: str = Form(""), min_rating: int = Form(4)):
    context = common_context(request)
    profile_id = profile or context.get("current_profile_id", "")
    doc = load_profiles_doc()
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    profile_obj = {**profile_obj, "id": profile_id}
    papers = rows(
        f"""
        SELECT *, {RATING_SQL} AS display_rating
        FROM papers
        WHERE profile_id = ? AND ({RATING_SQL} >= ? OR is_milestone = 1)
        ORDER BY year DESC, display_rating DESC, final_score DESC
        LIMIT 50
        """,
        (profile_id, min_rating),
    )
    kimi_body = ""
    if ai_parse_configured() and papers:
        kimi_body = build_kimi_review_body(
            f"{profile_obj.get('display_name') or profile_id} 研究脉络",
            "方向进展脉络",
            "中文",
            papers,
        )
    content = build_roadmap_markdown(profile_obj, papers, kimi_body, profile_id)
    ROADMAP_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = ROADMAP_REPORTS_DIR / f"{today_string()}_{safe_slug(profile_id)}_roadmap.md"
    write_markdown(path, content)
    return RedirectResponse(f"/roadmap?profile={profile_id}&generated={path.name}", status_code=303)


def build_idea_lab_page_context(
    request: Request,
    profile: str = "",
    branch: str = "",
    from_paper: str = "",
    idea_id: str = "",
    result: Optional[dict] = None,
    evidence: Optional[dict] = None,
) -> dict:
    context = common_context(request)
    if not advanced_features_enabled():
        return context
    profile_id = profile or context.get("current_profile_id", "")
    doc = load_profiles_doc()
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    branch_info = branch_meta(branch) if branch else {"title": "全方向", "keywords": "", "description": ""}
    local_count = row("SELECT COUNT(*) AS c FROM papers WHERE profile_id = ?", (profile_id,)) or {}
    branch_count = 0
    if branch and profile_id:
        ids = roadmap_branch_paper_ids(profile_id, branch)
        branch_count = len(ids)
        if not branch_count:
            kw = (branch_info.get("keywords") or "").split(",")[0].strip()
            if kw:
                bc = row(
                    "SELECT COUNT(*) AS c FROM papers WHERE profile_id = ? AND (title LIKE ? OR abstract_original LIKE ?)",
                    (profile_id, f"%{kw}%", f"%{kw}%"),
                )
                branch_count = int((bc or {}).get("c") or 0)
    raw_idea = ""
    idea = {}
    if idea_id:
        idea = row("SELECT * FROM ideas WHERE id = ?", (int(idea_id),)) or {}
        raw_idea = idea.get("title") or ""
        if idea.get("body_md"):
            raw_idea = (idea.get("title") or "") + "\n\n" + (idea.get("body_md") or "")
    elif from_paper:
        paper = row(f"SELECT p.*, rn.inspiration, rn.possible_ideas FROM papers p LEFT JOIN reading_notes rn ON rn.paper_id = p.stable_id WHERE p.stable_id = ?", (from_paper,))
        if paper:
            raw_idea = paper.get("title") or ""
            hints = " ".join(filter(None, [paper.get("possible_ideas"), paper.get("inspiration")]))
            if hints:
                raw_idea += f"\n\n相关灵感：{hints[:500]}"
    return {
        **context,
        "selected_profile": profile_id,
        "selected_profile_name": profile_obj.get("display_name") or profile_obj.get("name") or profile_id,
        "branch_key": branch,
        "branch_info": branch_info,
        "local_paper_count": local_count.get("c") or 0,
        "branch_paper_count": branch_count,
        "raw_idea": raw_idea,
        "idea": idea,
        "from_paper": from_paper,
        "result": result or {},
        "evidence": evidence or {},
    }


@app.get("/idea-lab")
def idea_lab_page(
    request: Request,
    profile: str = "",
    branch: str = "",
    from_paper: str = "",
    idea_id: str = "",
):
    if not advanced_features_enabled():
        return RedirectResponse("/settings?need_advanced=1", status_code=307)
    ctx = build_idea_lab_page_context(request, profile, branch, from_paper, idea_id)
    return templates.TemplateResponse(request, "idea_lab.html", ctx)


@app.post("/idea-lab/polish")
def idea_lab_polish(
    request: Request,
    profile: str = Form(""),
    branch: str = Form(""),
    raw_idea: str = Form(""),
    scope_branch: str = Form(""),
    scope_profile: str = Form("on"),
    scope_external: str = Form(""),
):
    if not advanced_features_enabled():
        return RedirectResponse("/settings?need_advanced=1", status_code=307)
    profile_id = profile or load_profiles_doc().get("default_profile", "")
    doc = load_profiles_doc()
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    query = raw_idea.strip()
    local_hits, rag_chunks = search_local_evidence(
        query, profile_id, branch if scope_branch == "on" else "", limit=10
    )
    if scope_profile != "on" and scope_branch != "on":
        local_hits, rag_chunks = [], []
    external_hits = []
    if scope_external == "on":
        external_hits = search_external_evidence(query)
    result = polish_idea_structured(
        raw_idea,
        profile_obj,
        branch,
        local_hits,
        external_hits,
        call_llm,
        ai_parse_configured(),
        rag_chunks=rag_chunks,
    )
    evidence = group_evidence(local_hits, external_hits)
    ctx = build_idea_lab_page_context(
        request, profile_id, branch, result=result, evidence=evidence
    )
    ctx["raw_idea"] = raw_idea
    return templates.TemplateResponse(request, "idea_lab.html", ctx)


@app.post("/idea-lab/save")
def idea_lab_save(
    request: Request,
    title: str = Form(""),
    body_md: str = Form(""),
    profile_id: str = Form(""),
    branch: str = Form(""),
    linked_paper_ids: str = Form(""),
    possible_direction: str = Form(""),
    next_tasks: str = Form(""),
    idea_id: str = Form(""),
):
    if not advanced_features_enabled():
        return RedirectResponse("/settings?need_advanced=1", status_code=307)
    branch_info = branch_meta(branch)
    title = (title or "").strip() or "未命名想法"
    if branch and not possible_direction:
        possible_direction = branch_info.get("title") or branch
    return ideas_save(
        idea_id=idea_id,
        title=title,
        body_md=body_md,
        profile_id=profile_id,
        status="想法中",
        linked_paper_ids=linked_paper_ids,
        possible_direction=possible_direction,
        next_tasks=next_tasks,
    )


@app.post("/idea-lab/fetch-task")
def idea_lab_fetch_task(
    profile: str = Form(""),
    branch: str = Form(""),
    keywords: str = Form(""),
    raw_idea: str = Form(""),
):
    if not advanced_features_enabled():
        return RedirectResponse("/settings?need_advanced=1", status_code=307)
    from urllib.parse import urlencode
    profile_id = profile or load_profiles_doc().get("default_profile", "")
    kw_list = [k.strip() for k in re.split(r"[,，;\n]+", keywords) if k.strip()]
    if not kw_list and raw_idea:
        kw_list = extract_fetch_keywords("", raw_idea)
    if branch and not kw_list:
        kw_list = [t.strip() for t in branch_meta(branch).get("keywords", "").split(",") if t.strip()][:4]
    prefill = ",".join(kw_list)
    params = urlencode({
        "prefill_keywords": prefill,
        "branch": branch or "",
        "from": "idea-lab",
        "profile_hint": profile_id,
    })
    return RedirectResponse(f"/fetch?{params}", status_code=303)


@app.get("/ideas")
def ideas_list_redirect(request: Request, profile: str = "", status: str = ""):
    return RedirectResponse(idea_lab_redirect_url(profile=profile), status_code=307)


@app.get("/ideas/new")
def ideas_new_redirect(request: Request, from_paper: str = ""):
    return RedirectResponse(idea_lab_redirect_url(from_paper=from_paper), status_code=307)


@app.get("/ideas/{idea_id}")
def ideas_edit_redirect(request: Request, idea_id: int):
    return RedirectResponse(idea_lab_redirect_url(idea_id=str(idea_id)), status_code=307)


@app.get("/wechat")
def wechat_page(request: Request):
    return RedirectResponse("/", status_code=307)


@app.get("/wechat/push-today")
def wechat_push_today_page(request: Request):
    return RedirectResponse("/", status_code=307)


@app.post("/wechat/push-today")
def wechat_push_today_action(request: Request):
    context = common_context(request)
    papers = top_papers_for_profile(context.get("current_profile_id", ""), 3)
    title, preview = build_wechat_text(context, papers)
    ok, message = send_serverchan_message(title, preview)
    if ok:
        ids = [(paper["stable_id"],) for paper in papers if paper.get("stable_id")]
        execute_many("UPDATE papers SET pushed_to_wechat = 1, updated_at = datetime('now') WHERE stable_id = ?", ids)
    return templates.TemplateResponse(
        request,
        "wechat_push_today.html",
        {**context, "papers": papers, "push_title": title, "push_preview": preview, "push_ok": ok, "push_message": message},
    )


def ensure_ideas_tables() -> None:
    init_db()


def _keyword_in_paper_text(keyword: str, paper: dict) -> bool:
    text = " ".join(
        [
            str(paper.get("title") or ""),
            str(paper.get("abstract_original") or ""),
            str(paper.get("abstract_zh") or ""),
        ]
    ).lower()
    kw = keyword.lower().replace("-", " ").replace("_", " ")
    hay = text.replace("-", " ").replace("_", " ")
    return kw in hay or kw.replace(" ", "") in hay.replace(" ", "")


def progress_metrics(profile_id: str) -> dict:
    """Aggregate KPIs for the research progress dashboard."""
    if not profile_id:
        return {"total": 0, "review_readiness": 0}
    base = row(
        f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN reading_status = 'read' THEN 1 ELSE 0 END) AS read_count,
            SUM(CASE WHEN reading_status = 'reading' THEN 1 ELSE 0 END) AS reading_count,
            SUM(CASE WHEN reading_status = 'todo' THEN 1 ELSE 0 END) AS todo_only,
            SUM(CASE WHEN reading_status IS NULL OR reading_status IN ('', 'unread') THEN 1 ELSE 0 END) AS unread_count,
            SUM(CASE WHEN reading_status IN ('unread', '', NULL) OR reading_status = 'todo' THEN 1 ELSE 0 END) AS todo_count,
            SUM(CASE WHEN reading_status = 'read' AND (
                (my_notes_path IS NOT NULL AND my_notes_path != '')
                OR stable_id IN (
                    SELECT paper_id FROM reading_notes
                    WHERE COALESCE(paper_topic, '') != '' OR COALESCE(core_method, '') != ''
                )
            ) THEN 1 ELSE 0 END) AS deep_read_count,
            SUM(CASE WHEN my_notes_path IS NOT NULL AND my_notes_path != '' THEN 1 ELSE 0 END) AS notes_count,
            SUM(CASE WHEN favorite = 1 OR is_favorite = 1 THEN 1 ELSE 0 END) AS favorite_count,
            SUM(CASE WHEN is_milestone = 1 THEN 1 ELSE 0 END) AS milestone_total,
            SUM(CASE WHEN is_milestone = 1 AND reading_status = 'read' THEN 1 ELSE 0 END) AS milestone_read,
            SUM(CASE WHEN {RATING_SQL} >= 4 THEN 1 ELSE 0 END) AS high_star_count,
            SUM(CASE WHEN abstract_is_complete = 1 THEN 1 ELSE 0 END) AS abstract_complete,
            SUM(CASE WHEN {RATING_SQL} >= 4 AND abstract_is_complete = 1 THEN 1 ELSE 0 END) AS review_ready,
            SUM(CASE WHEN COALESCE(doi, '') != '' OR COALESCE(doi_url, '') != '' THEN 1 ELSE 0 END) AS citation_ok
        FROM papers
        WHERE profile_id = ?
        """,
        (profile_id,),
    ) or {}
    ideas_row = row(
        """
        SELECT
            COUNT(*) AS ideas_count,
            SUM(CASE WHEN status IN ('准备实验', '准备写作') THEN 1 ELSE 0 END) AS idea_writable,
            SUM(CASE WHEN linked_paper_ids_json IS NOT NULL AND linked_paper_ids_json != '' THEN 1 ELSE 0 END) AS ideas_linked
        FROM ideas WHERE profile_id = ?
        """,
        (profile_id,),
    ) or {}
    base.update(ideas_row)
    total = int(base.get("total") or 0)
    ready = int(base.get("review_ready") or 0)
    base["review_readiness"] = int(100 * ready / total) if total else 0
    base["milestone_pct"] = int(100 * (base.get("milestone_read") or 0) / (base.get("milestone_total") or 1)) if base.get("milestone_total") else 0
    base["readiness"] = compute_writing_readiness(base)
    return base


def compute_writing_readiness(metrics: dict) -> dict:
    total = max(int(metrics.get("total") or 0), 1)
    milestone_pct = int(metrics.get("milestone_pct") or 0)

    def blend(weights: dict) -> int:
        score = 0.0
        score += weights.get("papers", 0) * min(100, total * 100 / 120)
        score += weights.get("high_star", 0) * (int(metrics.get("high_star_count") or 0) * 100 / total)
        score += weights.get("abstract", 0) * (int(metrics.get("abstract_complete") or 0) * 100 / total)
        score += weights.get("read", 0) * (int(metrics.get("read_count") or 0) * 100 / total)
        score += weights.get("notes", 0) * min(100, (int(metrics.get("notes_count") or 0)) * 12)
        score += weights.get("ideas", 0) * min(100, (int(metrics.get("ideas_count") or 0)) * 15)
        score += weights.get("milestone", 0) * milestone_pct
        score += weights.get("citation", 0) * (int(metrics.get("citation_ok") or 0) * 100 / total)
        return min(100, int(score))

    review = blend(
        {
            "papers": 0.12,
            "high_star": 0.22,
            "abstract": 0.18,
            "read": 0.15,
            "notes": 0.12,
            "ideas": 0.08,
            "milestone": 0.08,
            "citation": 0.05,
        }
    )
    paper_writing = blend(
        {
            "papers": 0.10,
            "high_star": 0.18,
            "abstract": 0.12,
            "read": 0.22,
            "notes": 0.18,
            "ideas": 0.12,
            "milestone": 0.08,
        }
    )
    seminar = blend(
        {
            "papers": 0.15,
            "high_star": 0.25,
            "abstract": 0.20,
            "read": 0.20,
            "notes": 0.10,
            "ideas": 0.05,
            "milestone": 0.05,
        }
    )
    return {
        "review": review,
        "paper": paper_writing,
        "seminar": seminar,
    }


def progress_yearly_counts(profile_id: str) -> list[dict]:
    rows_out = rows(
        """
        SELECT y, COUNT(*) AS c FROM (
            SELECT COALESCE(publication_year, year, CAST(substr(COALESCE(created_at, updated_at), 1, 4) AS INTEGER)) AS y
            FROM papers WHERE profile_id = ?
        )
        WHERE y IS NOT NULL AND y > 1900
        GROUP BY y
        ORDER BY y DESC
        LIMIT 12
        """,
        (profile_id,),
    )
    return [{"year": r["y"], "count": r["c"]} for r in rows_out]


def progress_keyword_coverage(profile_id: str, profile_obj: dict) -> list[dict]:
    keywords: list[str] = []
    for key in ("include_keywords", "research_focus"):
        for item in profile_obj.get(key) or []:
            if item and item not in keywords:
                keywords.append(item)
    if not keywords:
        return []
    papers = rows(
        "SELECT title, abstract_original, abstract_zh FROM papers WHERE profile_id = ?",
        (profile_id,),
    )
    total = len(papers) or 1
    threshold = max(3, total // 10)
    result = []
    for kw in keywords:
        count = sum(1 for p in papers if _keyword_in_paper_text(kw, p))
        result.append(
            {
                "keyword": kw,
                "count": count,
                "low": count < threshold,
                "pct": int(100 * count / total),
            }
        )
    result.sort(key=lambda x: x["count"], reverse=True)
    return result


def progress_milestone_papers(profile_id: str) -> list[dict]:
    return rows(
        f"""
        SELECT stable_id, title, year, reading_status, milestone_stage, milestone_reason,
               {RATING_SQL} AS display_rating, recommendation_level
        FROM papers
        WHERE profile_id = ? AND is_milestone = 1
        ORDER BY display_rating DESC, final_score DESC
        LIMIT 30
        """,
        (profile_id,),
    )


def progress_suggestions(profile_id: str, profile_obj: dict, metrics: dict, keywords: list[dict]) -> list[str]:
    tips: list[str] = []
    total = int(metrics.get("total") or 0)
    if total < 20:
        tips.append("当前文献积累不足，建议先点击「补全文献」或运行 annual_summary 年度模式。")
    this_year = current_year()
    yearly = {y["year"]: y["count"] for y in progress_yearly_counts(profile_id)}
    if yearly.get(this_year, 0) < max(5, total // 15):
        tips.append(f"建议补充 {this_year} 年最新文献（当前仅 {yearly.get(this_year, 0)} 篇）。")
    low_kw = [k["keyword"] for k in keywords if k.get("low")][:4]
    if low_kw:
        tips.append(f"关键词覆盖偏弱：{', '.join(low_kw)}，建议在获取更多文献时放宽检索或补充 Google Scholar。")
    if int(metrics.get("milestone_total") or 0) > int(metrics.get("milestone_read") or 0):
        left = int(metrics.get("milestone_total") or 0) - int(metrics.get("milestone_read") or 0)
        tips.append(f"尚有 {left} 篇里程碑论文未读，建议按「精读路线」顺序推进。")
    readiness = metrics.get("readiness") or {}
    if readiness.get("review", 0) >= 60:
        tips.append("你已具备生成综述大纲的基础，可在写作中心启动「一键综述」。")
    if readiness.get("paper", 0) < 50:
        tips.append("若准备写小论文，建议继续补充近 2–3 年高星论文并精读至少 3 篇里程碑文献。")
    if not tips:
        tips.append("进展良好，可继续积累笔记与灵感，并定期导出研究进展报告留档。")
    return tips


def build_progress_dashboard(profile_id: str, profile_obj: dict) -> dict:
    metrics = progress_metrics(profile_id) if profile_id else {"total": 0, "readiness": {}, "review_readiness": 0}
    keywords = progress_keyword_coverage(profile_id, profile_obj) if profile_id else []
    return {
        "metrics": metrics,
        "yearly": progress_yearly_counts(profile_id) if profile_id else [],
        "keywords": keywords,
        "milestones": progress_milestone_papers(profile_id) if profile_id else [],
        "suggestions": progress_suggestions(profile_id, profile_obj, metrics, keywords) if profile_id else [],
    }


def save_progress_snapshot(profile_id: str, metrics: dict, suggestions: list[str]) -> None:
    from datetime import datetime as dt
    init_db()
    now = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    today = today_string()
    suggestion = suggestions[0] if suggestions else ""
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO progress_snapshots (
                profile_id, snapshot_date, paper_count, read_count, todo_count,
                note_count, idea_count, milestone_total, milestone_read,
                review_readiness, next_suggestion, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                profile_id,
                today,
                int(metrics.get("total") or 0),
                int(metrics.get("read_count") or 0),
                int(metrics.get("todo_count") or 0),
                int(metrics.get("notes_count") or 0),
                int(metrics.get("ideas_count") or 0),
                int(metrics.get("milestone_total") or 0),
                int(metrics.get("milestone_read") or 0),
                int(metrics.get("review_readiness") or 0),
                suggestion,
                now,
            ),
        )
        conn.commit()


def render_progress_page(request: Request, profile_id: str):
    context = common_context(request)
    profile_id = profile_id or context.get("current_profile_id", "")
    doc = load_profiles_doc()
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    profile_obj = {**profile_obj, "id": profile_id}
    dashboard = build_progress_dashboard(profile_id, profile_obj)
    profiles = rows("SELECT profile_id, display_name, name FROM profiles ORDER BY profile_id")
    snapshots = rows(
        "SELECT * FROM progress_snapshots WHERE profile_id = ? ORDER BY snapshot_date DESC LIMIT 5",
        (profile_id,),
    ) if profile_id else []
    context.update(
        {
            "dashboard": dashboard,
            "metrics": dashboard["metrics"],
            "yearly": dashboard["yearly"],
            "keywords": dashboard["keywords"],
            "milestones": dashboard["milestones"],
            "suggestions": dashboard["suggestions"],
            "profiles": profiles,
            "selected_profile": profile_id,
            "profile_obj": profile_obj,
            "snapshots": snapshots,
            "insufficient": int(dashboard["metrics"].get("total") or 0) < 20,
        }
    )
    return templates.TemplateResponse(request, "progress.html", context)


@app.get("/progress")
def progress_page(request: Request, profile: str = ""):
    if not advanced_features_enabled():
        return RedirectResponse("/roadmap", status_code=307)
    ctx = common_context(request)
    return render_progress_page(request, profile or ctx.get("current_profile_id", ""))


@app.get("/progress/{profile_id}")
def progress_profile_page(request: Request, profile_id: str):
    if not advanced_features_enabled():
        return RedirectResponse("/roadmap", status_code=307)
    return render_progress_page(request, profile_id)


@app.post("/progress/generate")
def progress_generate(request: Request, profile: str = Form(""), snapshot: str = Form("")):
    doc = load_profiles_doc()
    profile_id = profile or doc.get("default_profile", "")
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    profile_obj = {**profile_obj, "id": profile_id}
    dashboard = build_progress_dashboard(profile_id, profile_obj)
    metrics = dashboard["metrics"]
    body = build_progress_report_markdown(profile_obj, dashboard)
    PROGRESS_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = PROGRESS_REPORTS_DIR / f"{safe_slug(profile_id)}_{today_string()}_progress.md"
    write_markdown(path, body)
    if snapshot == "on":
        save_progress_snapshot(profile_id, metrics, dashboard["suggestions"])
    return RedirectResponse(f"/progress/{profile_id}?generated=1", status_code=303)


@app.post("/progress/review-analysis")
def progress_review_analysis(profile: str = Form("")):
    doc = load_profiles_doc()
    profile_id = profile or doc.get("default_profile", "")
    profile_obj = doc.get("profiles", {}).get(profile_id, {})
    profile_obj = {**profile_obj, "id": profile_id}
    dashboard = build_progress_dashboard(profile_id, profile_obj)
    body = build_progress_review_analysis_markdown(profile_obj, dashboard)
    PROGRESS_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = PROGRESS_REPORTS_DIR / f"{safe_slug(profile_id)}_{today_string()}_review_readiness.md"
    write_markdown(path, body)
    return RedirectResponse(f"/progress/{profile_id}?generated=review", status_code=303)


def build_progress_review_analysis_markdown(profile: dict, dashboard: dict) -> str:
    name = profile.get("display_name") or profile.get("name") or profile.get("id")
    metrics = dashboard["metrics"]
    readiness = metrics.get("readiness") or {}
    lines = [
        f"# {name} 综述准备度分析",
        "",
        f"- 综述准备度：{readiness.get('review', metrics.get('review_readiness', 0))}%",
        f"- 小论文准备度：{readiness.get('paper', 0)}%",
        f"- 组会汇报准备度：{readiness.get('seminar', 0)}%",
        "",
        "## 判断依据",
        f"- 收录文献：{metrics.get('total', 0)}",
        f"- 四星及以上：{metrics.get('high_star_count', 0)}",
        f"- 摘要完整：{metrics.get('abstract_complete', 0)}",
        f"- 已读：{metrics.get('read_count', 0)}",
        f"- 笔记：{metrics.get('notes_count', 0)}",
        f"- 灵感：{metrics.get('ideas_count', 0)}",
        f"- 里程碑已读：{metrics.get('milestone_read', 0)}/{metrics.get('milestone_total', 0)}",
        "",
        "## 下一步建议",
    ]
    lines.extend(f"- {tip}" for tip in dashboard.get("suggestions", []))
    return "\n".join(lines)


def build_progress_report_markdown(profile: dict, dashboard: dict) -> str:
    name = profile.get("display_name") or profile.get("name") or profile.get("id")
    metrics = dashboard["metrics"]
    readiness = metrics.get("readiness") or {}
    lines = [
        f"# {name} 研究进展报告",
        f"日期：{today_string()}",
        "",
        "## 核心指标",
        f"- 已收录文献：{metrics.get('total', 0)} 篇",
        f"- 已精读：{metrics.get('deep_read_count', 0)} 篇",
        f"- 待读：{metrics.get('todo_count', 0)} 篇",
        f"- 我的笔记：{metrics.get('notes_count', 0)} 篇",
        f"- 灵感记录：{metrics.get('ideas_count', 0)} 条",
        f"- 综述准备度：{metrics.get('review_readiness', 0)}%",
        "",
        "## 写作准备度",
        f"- 综述：{readiness.get('review', 0)}%",
        f"- 小论文：{readiness.get('paper', 0)}%",
        f"- 组会汇报：{readiness.get('seminar', 0)}%",
        "",
        "## 按年文献积累",
    ]
    for item in dashboard.get("yearly", []):
        lines.append(f"- {item['year']} 年：{item['count']} 篇")
    lines.extend(["", "## 关键词覆盖"])
    for item in dashboard.get("keywords", []):
        flag = "（偏少）" if item.get("low") else ""
        lines.append(f"- {item['keyword']}：{item['count']} 篇 {flag}")
    lines.extend(["", "## 里程碑论文"])
    for paper in dashboard.get("milestones", [])[:20]:
        lines.append(f"- [{paper.get('reading_status')}] {paper.get('title')} ({paper.get('year')})")
    lines.extend(["", "## 建议"])
    lines.extend(f"- {tip}" for tip in dashboard.get("suggestions", []))
    return "\n".join(lines)


@app.post("/ideas/{idea_id}/generate-plan")
def idea_generate_plan(request: Request, idea_id: int):
    context = common_context(request)
    idea = row("SELECT * FROM ideas WHERE id = ?", (idea_id,))
    if not idea:
        return RedirectResponse("/idea-lab", status_code=303)
    query = " ".join([idea.get("title") or "", idea.get("possible_direction") or "", idea.get("body_md") or ""]).strip()
    chunks = rag_search(query, profile=idea.get("profile_id") or context.get("current_profile_id", ""), limit=8) if query else []
    profile = load_profiles_doc().get("profiles", {}).get(idea.get("profile_id") or context.get("current_profile_id", ""), {})
    fallback = f"""请基于以下文献依据和用户笔记，挖掘可执行科研灵感。

【研究方向】{profile.get('display_name') or profile.get('name') or idea.get('profile_id') or ''}
【用户问题或目标】{idea.get('title') or ''}

【文献与笔记依据】
{format_rag_context(chunks)}

请输出：可能创新点、对应文献依据、可行性、需要的数据、实验/代码方案、风险、小论文题目。
"""
    prompt = render_template(
        "idea_mining_prompt",
        {
            "profile_name": profile.get("display_name") or profile.get("name") or idea.get("profile_id") or "",
            "user_question": idea.get("title") or "",
            "rag_context": format_rag_context(chunks),
        },
        fallback,
    )
    generated = call_llm(prompt) if ai_parse_configured() else ""
    generated = generated or "未配置 AI Key 或生成失败。已列出可用支持文献，可手动整理研究方案。"
    merged = "\n\n".join(part for part in [idea.get("body_md") or "", "## RAG 支持文献生成的研究方案", generated] if part.strip())
    execute("UPDATE ideas SET body_md = ?, updated_at = datetime('now') WHERE id = ?", (merged, idea_id))
    return RedirectResponse(f"/idea-lab?idea_id={idea_id}&generated=1", status_code=303)


@app.post("/ideas/save")
def ideas_save(
    idea_id: str = Form(""),
    title: str = Form(""),
    body_md: str = Form(""),
    profile_id: str = Form(""),
    status: str = Form("想法中"),
    linked_paper_ids: str = Form(""),
    possible_direction: str = Form(""),
    next_tasks: str = Form(""),
):
    from datetime import datetime
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ensure_ideas_tables()
    title = (title or "").strip() or "未命名灵感"
    profile_id = (profile_id or "").strip() or (load_profiles_doc().get("default_profile") or "")
    try:
        if idea_id:
            execute(
                """
                UPDATE ideas SET title=?, body_md=?, profile_id=?, status=?, linked_paper_ids_json=?,
                possible_direction=?, next_tasks=?, updated_at=? WHERE id=?
                """,
                (title, body_md, profile_id, status, linked_paper_ids, possible_direction, next_tasks, now, int(idea_id)),
            )
            iid = int(idea_id)
        else:
            with connect() as conn:
                cursor = conn.execute(
                    """
                    INSERT INTO ideas (title, body_md, profile_id, status, linked_paper_ids_json,
                    possible_direction, next_tasks, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (title, body_md, profile_id, status, linked_paper_ids, possible_direction, next_tasks, now, now),
                )
                conn.commit()
                iid = int(cursor.lastrowid)
        IDEA_NOTES_DIR.mkdir(parents=True, exist_ok=True)
        slug = safe_slug(title or f"idea_{iid}")
        path = IDEA_NOTES_DIR / f"{safe_slug(profile_id or 'default')}/{today_string()}_{slug}.md"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(body_md or "", encoding="utf-8")
        from rag_indexer import index_markdown_file
        index_markdown_file(path, "idea", profile_id)
    except Exception as exc:
        return RedirectResponse("/idea-lab?error=save_failed", status_code=303)
    return RedirectResponse(f"/idea-lab?saved=1&id={iid}", status_code=303)


@app.get("/citations")
def citations_page(request: Request, collection_id: int = 0):
    return RedirectResponse("/reports", status_code=307)


@app.post("/citations/add")
def citations_add(paper_id: str = Form(...), collection_id: int = Form(0), name: str = Form("默认引用篮")):
    ensure_ideas_tables()
    if not collection_id:
        with connect() as conn:
            cursor = conn.execute(
                "INSERT INTO citation_collections (name, profile_id, format_default, created_at) VALUES (?, ?, ?, datetime('now'))",
                (name, "", "gbt7714"),
            )
            collection_id = int(cursor.lastrowid)
    execute(
        """
        INSERT INTO citation_items (collection_id, paper_id, sort_order)
        VALUES (?, ?, (SELECT COALESCE(MAX(sort_order),0)+1 FROM citation_items WHERE collection_id=?))
        ON CONFLICT(collection_id, paper_id) DO NOTHING
        """,
        (collection_id, paper_id, collection_id),
    )
    return RedirectResponse(f"/citations?collection_id={collection_id}", status_code=303)


@app.post("/citations/export")
def citations_export(request: Request, collection_id: int = Form(...), style: str = Form("gbt7714")):
    items = rows(
        """
        SELECT p.* FROM citation_items ci JOIN papers p ON p.stable_id = ci.paper_id
        WHERE ci.collection_id = ? ORDER BY ci.sort_order
        """,
        (collection_id,),
    )
    sys.path.insert(0, str(SCRIPTS_DIR))
    from citation_format import format_citations
    preview = format_citations(items, style)
    collections = rows("SELECT * FROM citation_collections ORDER BY id DESC")
    return templates.TemplateResponse(
        request,
        "citations.html",
        {**common_context(request), "collections": collections, "collection_id": collection_id, "items": items, "preview": preview, "style": style},
    )


@app.post("/papers/{paper_id}/upload-pdf")
async def upload_paper_pdf(paper_id: str, file: UploadFile = File(...)):
    paper = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,))
    if not paper:
        return RedirectResponse("/papers", status_code=303)
    profile = safe_slug(paper.get("profile_id") or "default")
    PDFS_DIR.mkdir(parents=True, exist_ok=True)
    dest = PDFS_DIR / profile / f"{paper_id}.pdf"
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(await file.read())
    sys.path.insert(0, str(SCRIPTS_DIR))
    from pdf_extract import extract_pdf_text, guess_metadata_from_text
    result = extract_pdf_text(dest)
    if result.get("error") and not result.get("text"):
        return RedirectResponse(f"/papers/{paper_id}?pdf_error={result['error'][:80]}", status_code=303)
    meta = guess_metadata_from_text(result.get("text", ""))
    execute(
        "UPDATE papers SET pdf_url = ?, updated_at = datetime('now') WHERE stable_id = ?",
        (str(dest.as_uri()), paper_id),
    )
    if meta.get("doi") and not paper.get("doi"):
        execute("UPDATE papers SET doi = ? WHERE stable_id = ?", (meta["doi"], paper_id))
    refreshed = row("SELECT * FROM papers WHERE stable_id = ?", (paper_id,)) or paper
    _, pdf_status = ensure_pdf_reading_note(refreshed, result.get("text", ""))
    rag_index_pdf_text(refreshed, result.get("text", ""), str(dest))
    rag_index_paper(paper_id)
    suffix = "pdf_ai=1" if pdf_status == "generated" else "template=1" if pdf_status == "template" else "pdf_ai=failed"
    return RedirectResponse(f"/papers/{paper_id}?pdf_uploaded=1&{suffix}#my-notes", status_code=303)


@app.post("/papers/bulk-status")
def papers_bulk_status(paper_ids: str = Form(""), reading_status: str = Form("read")):
    for pid in split_lines(paper_ids.replace(",", "\n")):
        if pid.strip():
            execute(
                "UPDATE papers SET reading_status = ?, updated_at = datetime('now') WHERE stable_id = ?",
                (reading_status, pid.strip()),
            )
    return RedirectResponse("/papers", status_code=303)


@app.get("/archive")
def archive_page(request: Request):
    return RedirectResponse("/reports", status_code=307)


@app.get("/archive/export-excel/preview")
def archive_export_excel_preview(request: Request):
    context = common_context(request)
    filters = parse_paper_filters_from_request(request, context.get("current_profile_id", ""))
    count = count_papers_for_filters(filters)
    return {"count": count, "scope_lines": describe_export_scope(filters)}


@app.get("/archive/export-excel")
def archive_export_excel_page(request: Request):
    context = common_context(request)
    filters = parse_paper_filters_from_request(request, context.get("current_profile_id", ""))
    export_count = count_papers_for_filters(filters)
    return templates.TemplateResponse(
        request,
        "archive_export_excel.html",
        {
            **context,
            "filters": filters,
            "export_count": export_count,
            "export_scope_lines": describe_export_scope(filters),
        },
    )


@app.post("/archive/export-excel")
def archive_export_excel_action(
    request: Request,
    profile: str = Form(""),
    year: str = Form(""),
    rating: str = Form(""),
    status: str = Form(""),
    q: str = Form(""),
    abstract: str = Form(""),
    if_min: str = Form(""),
    if_max: str = Form(""),
    jcr_quartile: str = Form(""),
    cas_quartile: str = Form(""),
):
    context = common_context(request)
    selected_profile = profile if profile not in ("", None) else context.get("current_profile_id", "")
    filters = {
        "profile": selected_profile or context.get("current_profile_id", ""),
        "year": year,
        "rating": rating,
        "status": status,
        "q": q,
        "abstract": abstract,
        "if_min": if_min,
        "if_max": if_max,
        "jcr_quartile": jcr_quartile,
        "cas_quartile": cas_quartile,
    }
    export_rows = query_papers(filters, limit=5000)
    path = export_papers_excel(export_rows)
    return templates.TemplateResponse(
        request,
        "archive_export_excel.html",
        {
            **context,
            "generated_path": path,
            "export_count": len(export_rows),
            "filters": filters,
            "export_scope_lines": describe_export_scope(filters),
        },
    )


def _journal_metrics_dashboard_stats() -> dict:
    stats = metrics_file_stats(str(JOURNAL_METRICS_PATH)) if metrics_file_stats else {"exists": False, "count": 0}
    total = row("SELECT COUNT(*) AS c FROM papers", ()) or {}
    matched = row("SELECT COUNT(*) AS c FROM papers WHERE journal_matched = 1", ()) or {}
    stats["papers_total"] = total.get("c", 0)
    stats["papers_matched"] = matched.get("c", 0)
    stats["papers_unmatched"] = stats["papers_total"] - stats["papers_matched"]
    return stats


@app.post("/settings/journal-metrics")
def upload_journal_metrics(file: UploadFile = File(...)):
    with JOURNAL_METRICS_PATH.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    return RedirectResponse("/settings?journal_uploaded=1", status_code=303)


@app.post("/settings/journal-metrics/rematch")
def rematch_journal_metrics():
    if rematch_journal_ranks is None:
        return RedirectResponse("/settings?journal_error=module", status_code=303)
    rematch_journal_ranks(str(JOURNAL_METRICS_PATH))
    return RedirectResponse("/settings?journal_rematched=1", status_code=303)


@app.get("/settings/journal-metrics/unmatched.csv")
def download_unmatched_journals():
    paper_rows = rows("SELECT title, journal, issn, eissn, journal_matched FROM papers", ())
    if export_unmatched_journals is None:
        return RedirectResponse("/settings", status_code=303)
    unmatched = export_unmatched_journals([dict(p) for p in paper_rows])
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=["journal_name", "issn", "eissn", "paper_count", "example_title"],
    )
    writer.writeheader()
    writer.writerows(unmatched)
    out_path = EXPORTS_DIR / "unmatched_journals.csv"
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path.write_text(buffer.getvalue(), encoding="utf-8-sig")
    return FileResponse(out_path, filename="unmatched_journals.csv")


@app.get("/settings/journal-metrics/example")
def download_journal_metrics_example():
    example = SKILL_DIR / "journal_metrics.example.csv"
    if not example.exists():
        return RedirectResponse("/settings", status_code=303)
    return FileResponse(example, filename="journal_metrics.example.csv")


def _normalize_doi(value: str) -> str:
    text = str(value or "").strip().lower()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if text.startswith(prefix):
            text = text[len(prefix):]
    return text.strip("/")


def lookup_paper_in_library(title: str = "", doi: str = "", journal: str = "") -> Optional[dict]:
    doi_key = _normalize_doi(doi)
    if doi_key:
        found = row(
            "SELECT *, {rating} AS display_rating FROM papers WHERE LOWER(COALESCE(doi, '')) = ? OR stable_id = ? LIMIT 1".format(
                rating=RATING_SQL
            ),
            (doi_key, f"doi:{doi_key}"),
        )
        if found:
            return found
    title_text = str(title or "").strip()
    if title_text:
        found = row(
            f"SELECT *, {RATING_SQL} AS display_rating FROM papers WHERE title = ? LIMIT 1",
            (title_text,),
        )
        if found:
            return found
        sid = paper_stable_id({"title": title_text, "doi": doi_key, "year": ""})
        found = row(
            f"SELECT *, {RATING_SQL} AS display_rating FROM papers WHERE stable_id = ? LIMIT 1",
            (sid,),
        )
        if found:
            return found
        found = row(
            f"SELECT *, {RATING_SQL} AS display_rating FROM papers WHERE title LIKE ? LIMIT 1",
            (f"%{title_text[:120]}%",),
        )
        if found:
            return found
    return None


@app.get("/api/journal-rank")
def api_journal_rank(journal: str = "", issn: str = "", eissn: str = "", doi: str = ""):
    if match_journal_rank is None:
        return JSONResponse({"matched": False, "error": "module_unavailable"})
    rank = match_journal_rank(journal, issn, eissn, doi=doi)
    return JSONResponse(
        {
            "journal": journal or rank.get("journal_name", ""),
            "matched": bool(rank.get("matched")),
            "jcr_impact_factor": rank.get("jcr_impact_factor"),
            "jcr_year": rank.get("jcr_year"),
            "jcr_quartile": rank.get("jcr_quartile"),
            "cas_quartile": rank.get("cas_quartile"),
            "cas_top": rank.get("cas_top"),
            "cas_warning": rank.get("cas_warning"),
            "core_tags": rank.get("core_tags"),
            "journal_quality_score": rank.get("journal_quality_score"),
            "journal_match_method": rank.get("journal_match_method"),
            "journal_rank_source": rank.get("journal_rank_source"),
        }
    )


@app.get("/api/paper-match")
def api_paper_match(title: str = "", doi: str = "", journal: str = "", issn: str = "", eissn: str = ""):
    if match_journal_rank is None:
        return JSONResponse({"matched": False, "in_library": False, "error": "module_unavailable"})
    rank = match_journal_rank(journal, issn, eissn, doi=doi)
    library_paper = lookup_paper_in_library(title=title, doi=doi, journal=journal)
    system_rating = None
    if library_paper:
        system_rating = (
            library_paper.get("display_rating")
            or library_paper.get("system_rating")
            or score_to_rating(library_paper.get("final_score"))
        )
    metrics_summary = ""
    if paper_metrics_line is not None and library_paper:
        metrics_summary = paper_metrics_line(library_paper)
    elif paper_metrics_line is not None and rank.get("matched"):
        metrics_summary = paper_metrics_line({**rank, "abstract_is_complete": None, "citation_count": None})
    return JSONResponse(
        {
            "title": title,
            "doi": _normalize_doi(doi),
            "journal": journal or rank.get("journal_name", ""),
            "journal_match": {
                "matched": bool(rank.get("matched")),
                "jcr_impact_factor": rank.get("jcr_impact_factor"),
                "jcr_year": rank.get("jcr_year"),
                "jcr_quartile": rank.get("jcr_quartile"),
                "cas_quartile": rank.get("cas_quartile"),
                "cas_top": rank.get("cas_top"),
                "cas_warning": rank.get("cas_warning"),
                "core_tags": rank.get("core_tags"),
                "journal_quality_score": rank.get("journal_quality_score"),
                "journal_match_method": rank.get("journal_match_method"),
                "journal_rank_source": rank.get("journal_rank_source"),
            },
            "in_library": bool(library_paper),
            "stable_id": library_paper.get("stable_id") if library_paper else None,
            "system_rating": system_rating,
            "metrics_summary": metrics_summary,
        }
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("web.web_app:app", host="127.0.0.1", port=8000, reload=True)
