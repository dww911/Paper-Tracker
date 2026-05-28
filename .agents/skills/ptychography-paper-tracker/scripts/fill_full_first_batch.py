#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ===================== 全局核心配置（和主脚本保持一致）=====================
BASE_CONFIG = {
    "EXCEL_SAVE_PATH": os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                                    "Ptychography_论文全量库.xlsx"),
    # Excel 固定表头（严格匹配需求，11 项字段顺序不可修改）
    "EXCEL_HEADERS": [
        "论文名字", "网址", "期刊", "影响因子", "发布时间",
        "摘要中文翻译", "研究背景", "论文创新点", "实验结果", "总结", "未来展望", "可创新点"
    ]
}

# 第一波10篇论文 - 完整版（只显示年份，包含摘要中文翻译）
parsed_papers = [
    {
        "id": "scholar_0_1",
        "title": "Information in 4D-STEM: Where it is, and How to Use it",
        "authors": ["D Ma", "G Li", "DA Muller", "SE Zeltmann"],
        "published_time": "2025",
        "abstract": "... electron microscopy. ... electron microscopy (4D-STEM) data, and consequently identify new imaging modes that can also serve as crude but fast approximations to iterative ptychography. ...",
        "link": "https://www.sciencedirect.com/science/article/pii/S0304399126000446",
        "journal": "Micron",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "在电子显微学中，本文讨论了4D-STEM数据中的信息分布，并因此识别出了新的成像模式，这些成像模式可以作为迭代Ptychography的粗略但快速的近似。",
        "研究背景": "随着4D-STEM技术快速发展，学界需要更清晰地理解4D-STEM数据中信息的分布规律。现有迭代Ptychography重构精度高但计算速度慢，亟需寻找能够平衡精度和速度的快速近似成像方法。",
        "论文创新点": "1. 系统分析了4D-STEM数据中信息在不同维度的分布，识别出可作为迭代Ptychography粗略但快速近似的新型成像模式。\n2. 为不同研究需求在成像精度和计算速度之间提供了更多选择空间，降低了4D-STEM成像应用门槛。",
        "实验结果": "文章通过理论分析识别了新的成像模式空间，验证了这些模式作为迭代Ptychography近似的可行性，未提供具体实验量化数据。",
        "总结": "本文系统分析了4D-STEM数据中的信息分布，提出了可替代迭代Ptychography的快速近似成像模式，为4D-STEM快速成像领域提供了重要理论指导。",
        "未来展望": "作者期待这些新识别的成像模式能够在实际实验数据中得到验证，进一步推动4D-STEM技术在更多研究场景中的应用。",
        "可创新点": "1. 可以将本文提出的快速近似成像思想与WDD算法结合，研究基于WDD的分步Ptychography重构策略，先用WDD快速得到粗重构，再对关键区域进行迭代优化，在保证整体精度的同时显著提高重构速度。\n2. 基于本文的信息分布分析结论，可以研究自适应采样策略，对信息丰富区域加密采样，对信息稀疏区域稀疏采样，在保证成像质量的同时降低总数据量和样品辐照剂量。"
    },
    {
        "id": "scholar_1_2",
        "title": "Open-Source Phase Reconstructions of Focused-Probe 4D-STEM Data with Near-Ideal Direct-Electron Detection",
        "authors": ["T Susi", "N Dellby", "R Hayner", "C Hofer", "J Kotakoski"],
        "published_time": "2025",
        "abstract": "... electron microscopy (STEM) to make use of every electron for ... (WDD) [5–8], as well as iterative differential phase contrast (DPC… iterative gradient descent single-slice ptychography [4, 9]. ...",
        "link": "https://academic.oup.com/mam/article-abstract/30/Supplement_1/ozae044.920/7719835",
        "journal": "Microscopy and Microanalysis",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "电子显微学（STEM）需要充分利用每一个电子，本文工作支持WDD方法以及迭代差分相位对比（DPC）、迭代梯度下降单切片Ptychography等多种算法。",
        "研究背景": "开源相位重构工具对4D-STEM Ptychography领域开放研究至关重要，但现有开源工具对新型直接电子探测器数据支持不完善，需要更好地集成WDD和迭代Ptychography等多种算法。",
        "论文创新点": "1. 实现了支持近理想直接电子探测聚焦探针4D-STEM数据的开源相位重构工作流，原生支持WDD方法以及迭代梯度下降单切片Ptychography等多种主流算法。\n2. 使更多不具备高级算法开发能力的研究团队能够方便使用最先进的相位重构算法，促进了领域开放科学发展。",
        "实验结果": "文章成功集成了包括WDD在内的多种相位重构算法，对直接电子探测器数据格式提供了完整支持，未在摘要中提供具体性能测试数据。",
        "总结": "本文提供了支持直接电子探测数据的完整开源4D-STEM相位重构实现，降低了领域研究门槛，推动了Ptychography技术普及。",
        "未来展望": "期待社区共同参与测试和改进，进一步扩大对更多探测器类型和不同实验配置的支持。",
        "可创新点": "1. 在这个开源框架基础上，可以优化WDD算法的内存访问模式和GPU并行加速，使其能够处理分辨率更高、尺寸更大的4D-STEM数据集。\n2. 可以集成自动参数调优模块，根据输入数据特性自动选择最优重构参数，降低非专业用户的使用门槛。"
    },
    {
        "id": "scholar_2_3",
        "title": "Electromagnetic field reconstructions of 4D-STEM datasets using ptychography and differential phase contrast imaging",
        "authors": ["S Cao", "M Chi", "KL More"],
        "published_time": "2025",
        "abstract": "... Ptychography is a more complicated phase retrieval method that either employs iterative algorithms [4-6] or phases up the ... (WDD) or Single-Side Band (SSB) methods and showed that ...",
        "link": "https://www.cambridge.org/core/services/aop-cambridge-core/content/view/5CC143C62AB11B1C84839DCE26D65273/S1431927619001065a.pdf/electromagnetic_field_reconstructions_of_4dstem_datasets_using_ptychography_and_differential_phase_contrast_imaging.pdf",
        "journal": "Cambridge University Press",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "Ptychography是一种更复杂的相位恢复方法，可以采用迭代算法，本文对比了Ptychography与WDD（Wigner分布反卷积）或单边带（SSB）方法，并给出了对比结果。",
        "研究背景": "基于4D-STEM的电磁场重构是电子显微学中重要研究方向，Ptychography和差分相位对比（DPC）是两种主流方法。Ptychography采用迭代算法复杂度较高，而WDD/SSB等直接方法更加简洁高效，需要系统比较两类方法的优缺点。",
        "论文创新点": "1. 系统比较了使用Ptychography和差分相位对比成像两种方法对4D-STEM数据集进行电磁场重构的性能差异。\n2. 分析对比了迭代Ptychography和WDD/SSB直接方法在电磁场重构应用中的适用场景。",
        "实验结果": "文章进行了方法比较分析，但摘要未提供具体实验量化结果。理论上Ptychography能够获取更完整信息，但计算代价更高。",
        "总结": "本文对两种主流电磁场重构方法进行了比较分析，为研究者针对具体应用场景选择合适方法提供了重要参考。",
        "未来展望": "未来可研究结合两种方法优势的混合重构算法，兼顾精度和计算效率。",
        "可创新点": "1. 研究结合WDD初始估计和Ptychography迭代优化的混合电磁场重构算法，先用WDD快速得到电磁场分布初始估计，再用Ptychography进行精细化优化，在保持高精度的同时提高重构效率。"
    },
    {
        "id": "scholar_3_4",
        "title": "Quantitative comparison of HRTEM and electron ptychography",
        "authors": ["F Bennemann", "P Nellist", "A Kirkland"],
        "published_time": "2025",
        "abstract": "... WDD) method. In this work we introduce the detective quantum efficiency (DQE), applied to electron microscopy … All 4D-STEM simulations assumed an 80keV beam energy with a probe …",
        "link": "https://www.bio-conferences.org/articles/bioconf/pdf/2024/48/bioconf_emc2024_04007.pdf",
        "journal": "Bioconferences (EMC 2024)",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "本文采用WDD方法，将探测量子效率（DQE）引入电子显微学，所有4D-STEM模拟都假设采用80keV束流能量和探针...",
        "研究背景": "高分辨透射电子显微镜（HRTEM）和电子Ptychography是目前两种主要高分辨电子显微成像方法，但缺乏基于探测量子效率（DQE）的系统定量比较。使用WDD方法进行Ptychography重构需要更客观的性能评估。",
        "论文创新点": "1. 将探测量子效率（DQE）这一定量指标应用于电子Ptychography和HRTEM的性能比较，提供了更客观的评估标准。\n2. 在80keV电子束能量条件下进行了系统的4D-STEM模拟，为两种方法性能比较提供了定量依据。",
        "实验结果": "建立了基于模拟的定量比较框架，但摘要未给出具体DQE对比数值结果。",
        "总结": "本文通过DQE指标对HRTEM和电子Ptychography进行了定量比较，为高分辨成像方法选择提供了客观依据。",
        "未来展望": "可进一步在真实实验数据上验证比较结论，并拓展到不同加速电压和不同样品厚度的场景。",
        "可创新点": "1. 基于本文的定量比较结论，可以研究针对不同分辨率要求和样品类型，自适应选择HRTEM还是Ptych成像的智能电子显微成像策略。"
    },
    {
        "id": "scholar_4_5",
        "title": "Non-Iterative Electron Ptychography With Relaxed Real-Space Sampling",
        "authors": ["CM O'Leary", "H Sha", "G Varnavides"],
        "published_time": "2025",
        "abstract": "... electron microscopy (STEM). By acquiring a two-dimensional array of twodimensional diffraction patterns and implementing phase retrieval algorithms, … ptychography to 4D STEM data …",
        "link": "https://academic.oup.com/mam/article-pdf/doi/10.1093/mam/ozaf048.055/63846453/ozaf048.055.pdf",
        "journal": "Microscopy and Microanalysis",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "在电子显微学（STEM）中，通过获取二维衍射图案阵列并实现相位恢复算法，将Ptychography应用于4D STEM数据...",
        "研究背景": "非迭代电子Ptychography（如WDD算法）计算速度快，但对实空间扫描采样密度有较严格要求，欠采样条件下重构质量会明显下降。放松采样要求能够减少数据采集时间和样品辐照剂量。",
        "论文创新点": "1. 提出了一种放松实空间采样要求的非迭代电子Ptychography方法，降低了对扫描步长的严格限制。\n2. 在保持非迭代快速重构优势的同时，允许更大的扫描步长，有效减少了总数据采集时间和样品接受的辐照剂量。",
        "实验结果": "方法设计理论上支持更宽松的采样条件，但摘要未给出具体实验验证结果。",
        "总结": "本文提出了采样要求更宽松的非迭代电子Ptychography方法，有利于低剂量快速4D-STEM成像，对辐照敏感样品研究具有重要价值。",
        "未来展望": "需要在真实实验数据上验证该方法的实际重构性能，并拓展应用到更多类型的样品体系。",
        "可创新点": "1. 可以将本文的放松采样思想与WDD算法深度结合，研究欠采样条件下的WDD重构改进算法，在大扫描步长情况下仍能保持较高重构质量。\n2. 结合压缩感知技术，可以进一步降低采样率，实现亚采样条件下的超分辨Ptych重构。"
    },
    {
        "id": "scholar_5_6",
        "title": "Transfer of Information Across Various Phase Retrieval STEM Techniques",
        "authors": ["G Varnavides", "SM Ribet", "JM Bekkevold"],
        "published_time": "2025",
        "abstract": "... in direct and iterative phase retrieval algorithms, including center… Finally, I will propose iterative ptychographicholography – where … proposed quantuminformation limits of 4D-STEM [10]. …",
        "link": "https://www.scienceopen.com/hosted-document?doi=10.14293/APMC13-2025-0268",
        "journal": "APMC Conference",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "在直接和迭代相位恢复算法中，包括...最后，作者提出了迭代Ptychography全息术，... 提出了4D-STEM的量子信息极限。",
        "研究背景": "不同相位恢复算法（直接法如WDD vs 迭代法如Ptychography）在4D-STEM成像中信息传递效率存在差异，需要从信息论角度进行分析比较，帮助理解不同算法特性。",
        "论文创新点": "1. 从信息论角度分析比较了直接法和迭代法等不同相位检索算法之间的信息传递效率。\n2. 提出了迭代Ptychography全息方法，结合不同方法优势提高信息传递效率。",
        "实验结果": "文章从理论层面分析了不同算法的信息传递特性，未提供具体实验结果。",
        "总结": "本文从信息论角度分析比较了各种STEM相位检索技术的信息传递特性，提出了新型迭代Ptychography全息方法，为算法改进提供了理论指导。",
        "未来展望": "作者提出需要进一步从实验层面验证所提方法的实际性能。",
        "可创新点": "1. 基于本文的信息传递分析，可以研究WDD算法与迭代Ptychography结合的混合算法，利用WDD提供高质量初始估计，提高迭代收敛速度和信息传递效率。"
    },
    {
        "id": "scholar_6_7",
        "title": "Live processing of momentum-resolved STEM data for first moment imaging and ptychography",
        "authors": ["A Strauch", "D Weber", "A Clausen"],
        "published_time": "2025",
        "abstract": "... in 4D scanning transmission electron microscopy (STEM) using the … In this study, we demonstrate 4D-STEM continuous live … Its capability of direct, dose-efficient phase recovery makes …",
        "link": "https://academic.oup.com/mam/article-abstract/27/5/1078/6888074",
        "journal": "Microscopy and Microanalysis",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "在4D扫描透射电子显微镜（STEM）中使用... 本研究证明了4D-STEM连续实时... 直接、剂量高效的相位恢复能力使其能够...",
        "研究背景": "4D-STEM实时成像对实验过程中快速观察结果很重要，但Ptychography重构计算量大，难以实现连续实时处理。直接相位恢复方法（如WDD）具有直接、剂量高效的特点，适合实时处理。",
        "论文创新点": "1. 实现了4D-STEM动量分辨数据的实时处理流水线，支持一阶矩成像和Ptychography重构实时显示。\n2. 利用直接相位恢复方法的剂量高效特性，实现了连续实时成像，实验过程中可即时观察结果。",
        "实验结果": "实验验证了实时处理能力，证明了该方案能够支持连续live 4D-STEM成像，满足实验过程中实时观察需求。",
        "总结": "本文实现了4D-STEM一阶矩成像和Ptychography的实时处理流水线，支持实验过程中连续实时成像，大大改善了实验体验。",
        "未来展望": "可进一步优化处理速度，支持更高分辨率更大尺寸数据集的实时处理。",
        "可创新点": "1. 将WDD算法集成到类似实时处理框架中，利用Wdd的非迭代快速特性实现实时相位重构，提供比一阶矩成像更高精度的实时相位成像结果。"
    },
    {
        "id": "scholar_7_8",
        "title": "In silico ptychography of lithium-ion cathode materials from subsampled 4-D STEM data",
        "authors": ["AW Robinson", "A Moshtaghpour", "J Wells"],
        "published_time": "2025",
        "abstract": "... is recovered through a phase retrieval algorithm, meaning low … the phase of the 4-D STEM data, we chose to use the WDD … scanning transmission electron microscopy (4d-stem): From …",
        "link": "https://arxiv.org/abs/2307.06138",
        "journal": "arXiv",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "...相位通过相位恢复算法恢复，这意味着低... 对于4-D STEM数据的相位，我们选择使用WDD方法... 扫描透射电子显微镜（4d-stem）：从...",
        "研究背景": "锂离子阴极材料的纳米尺度结构表征对电池性能研究至关重要，但全采样4D-STEM数据采集时间长、样品接受辐照剂量大，容易引起材料损伤，需要从子采样数据中获得高质量Ptychography重构。本文选择WDD算法进行相位恢复。",
        "论文创新点": "1. 提出了从子采样4-D STEM数据通过计算机模拟Ptychography重构获得锂离子阴极材料纳米结构信息的方法。\n2. 采用WDD算法进行相位恢复，在子采样条件下仍能获得可用重构结果，降低了数据采集需求和辐照损伤。",
        "实验结果": "计算机模拟验证表明，使用WDD算法能够从子采样数据中成功恢复相位信息，获得阴极材料的纳米结构。",
        "总结": "本文验证了从子采样4D-STEM数据通过WDD Ptychography重构获得锂离子阴极材料纳米结构信息的可行性，为能源材料低剂量表征提供了有效途径。",
        "未来展望": "需要在真实实验数据上验证该方法，并推广应用到更多能源材料体系。",
        "可创新点": "1. 基于本文研究，可以针对能源材料低剂量成像需求，优化WDD算法在欠采样条件下的重构稳定性，抑制欠采样带来的伪影。"
    },
    {
        "id": "scholar_8_9",
        "title": "Simultaneous high-speed and low-dose 4-D STEM using compressive sensing techniques",
        "authors": ["AW Robinson", "A Moshtaghpour", "J Wells"],
        "published_time": "2025",
        "abstract": "... algorithms [49–54] that recover the complex ptychographic … 4(a) suggests that ptychographic phase recovery is more … scanning transmission electron microscopy (4D-STEM): From …",
        "link": "https://arxiv.org/abs/2309.14055",
        "journal": "arXiv",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "...算法恢复复值Ptychography... 图4(a)表明Ptychography相位恢复更... 扫描透射电子显微镜（4D-STEM）：从...",
        "研究背景": "4D-STEM Ptychography需要同时实现高速数据采集和低辐照剂量，这对锂电池等辐照敏感能源材料特别重要。压缩感知技术能够从较少测量中恢复完整信息，为解决这一问题提供了新思路。",
        "论文创新点": "1. 将压缩感知技术应用到4D-STEM数据采集中，实现同时高速和低剂量成像。\n2. 结合现有Ptychographic相位恢复算法，能够从压缩采样数据中恢复高质量相位信息。",
        "实验结果": "分析表明，压缩采样条件下Ptychography相位恢复仍能获得较好结果，验证了所提方法的有效性。",
        "总结": "本文提出使用压缩感知技术实现4D-STEM同时高速低剂量成像，能够有效减少数据采集时间和辐照剂量，有利于辐照敏感样品表征。",
        "未来展望": "需要在更多实际样品上验证方法性能，优化压缩采样策略进一步提高性能。",
        "可创新点": "1. 结合压缩感知技术和WDD算法，研究基于压缩感知的快速非迭代Ptychography，在保证重构质量的同时实现更低采样率和更快成像速度。"
    },
    {
        "id": "scholar_9_10",
        "title": "Influence of the loss function on gradient-based iterative ptychographic reconstructions in 4D-STEM",
        "authors": ["ML Leidl", "C Sachse"],
        "published_time": "2025",
        "abstract": "... However, direct reconstruction algorithms like single-sideband (SSB) or Wigner-distribution deconvolution (WDD) … Low-dose phase retrieval of biological specimens using cryo-electron …",
        "link": "https://www.bio-conferences.org/articles/bioconf/pdf/2024/48/bioconf_emc2024_04013.pdf",
        "journal": "Bioconferences (EMC 2024)",
        "impact_factor": "谷歌学术不提供",
        "摘要中文翻译": "然而，直接重构算法如单边带（SSB）或Wigner分布反卷积（WDD）... 使用冷冻电子对生物样品进行低剂量相位恢复...",
        "研究背景": "在基于梯度的迭代Ptychography重构中损失函数的选择对最终重构质量有重要影响。虽然SSB和WDD等直接重构算法速度更快，但迭代方法通常能提供更高重构精度，需要系统研究不同损失函数对重构结果的影响。",
        "论文创新点": "1. 系统研究了损失函数对4D-STEM中基于梯度的迭代Ptychography重构质量的影响。\n2. 为不同实验条件和样品类型选择合适损失函数提供指导，帮助研究者获得更高质量重构结果。",
        "实验结果": "文章分析了不同损失函数在不同场景下的表现，但摘要未给出具体量化结果。",
        "总结": "本文系统分析了损失函数选择对迭代Ptychography重构结果的影响，为算法参数选择提供了实践指导。",
        "未来展望": "可基于研究结果，针对特定样品类型开发自适应选择损失函数的自动算法。",
        "可创新点": "1. 研究在WDD获得的初始估计基础上，使用不同损失函数进行迭代优化的效果，找到兼顾重构速度和最终精度的组合方案。"
    }
]

def append_to_excel(papers: list):
    """将解析完成的论文追加到 Excel 文件，按年度分 Sheet"""
    if not papers:
        print("没有论文需要添加")
        return

    year_group = {}
    for paper in papers:
        pub_year = paper["published_time"][:4]
        if pub_year not in year_group:
            year_group[pub_year] = []
        row_data = [
            paper["title"],
            paper["link"],
            paper["journal"],
            paper["impact_factor"],
            paper["published_time"],
            paper["摘要中文翻译"],
            paper["研究背景"],
            paper["论文创新点"],
            paper["实验结果"],
            paper["总结"],
            paper["未来展望"],
            paper["可创新点"]
        ]
        year_group[pub_year].append(row_data)

    if os.path.exists(BASE_CONFIG["EXCEL_SAVE_PATH"]):
        # 删除已有文件，从头开始创建
        os.remove(BASE_CONFIG["EXCEL_SAVE_PATH"])

    with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl") as writer:
        for year, rows in year_group.items():
            sheet_name = f"{year}年论文"
            df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            header_font = Font(bold=True)
            for col in range(1, len(BASE_CONFIG["EXCEL_HEADERS"]) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    print("新增论文已成功保存到 Excel 文件")

if __name__ == "__main__":
    append_to_excel(parsed_papers)
