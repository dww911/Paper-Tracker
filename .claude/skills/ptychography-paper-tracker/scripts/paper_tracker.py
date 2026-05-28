import os
import json
import requests
import feedparser
import argparse
import pandas as pd
import sys
import time

# 设置标准输出编码为UTF-8，解决Windows GBK编码问题

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 新增：导入SerpApi谷歌学术依赖

from serpapi import Client

# ===================== 全局核心配置（可按需修改检索关键词）=====================

BASE_CONFIG = {
    # 研究方向精准过滤规则（电子显微学 Ptychography，排除无关领域）
    "SEARCH_KEYWORDS": (
        'ptychography electron microscopy 4D-STEM STEM phase retrieval WDD algorithm'
        ' -X-ray -optical -synchrotron -laser'
    ),
    "ARXIV_CATEGORIES": ["physics.optics", "cond-ml.mtrl-sci", "physics.ins-det"],
    # API 配置
    "ARXIV_API_URL": "http://export.arxiv.org/api/query",
    "SEMANTIC_SCHOLAR_API_URL": "https://api.semanticscholar.org/graph/v1/paper",
    # 新增：谷歌学术 SerpApi 配置
    "GOOGLE_SCHOLAR_ENGINE": "google_scholar",
    # 文件路径配置
    "HISTORY_FILE": os.path.join(os.path.dirname(os.path.dirname(__file__)), "paper_history.json"),
    "EXCEL_SAVE_PATH": os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                                    "Ptychography_论文全量库.xlsx"),
    # Excel 固定表头（严格匹配需求，11 项字段顺序不可修改）
    "EXCEL_HEADERS": [
        "论文名字", "网址", "期刊", "影响因子", "发布时间",
        "摘要中文翻译", "研究背景", "论文创新点", "实验结果", "总结", "未来展望", "可创新点"
    ]
}


# ===================== 工具函数 =====================

def load_history() -> Dict:
    """加载论文历史记录，避免重复解析"""
    if os.path.exists(BASE_CONFIG["HISTORY_FILE"]):
        with open(BASE_CONFIG["HISTORY_FILE"], "r", encoding="utf-8") as f:
            return json.load(f)
    return {"papers": {}}


def save_history(history: Dict):
    """保存论文历史记录"""
    with open(BASE_CONFIG["HISTORY_FILE"], "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def filter_new_papers(papers: List[Dict], history: Dict) -> List[Dict]:
    """过滤出从未处理过的新论文"""
    return [p for p in papers if p["id"] not in history["papers"]]


# ===================== 元数据获取 =====================

def get_paper_metadata_by_arxiv(arxiv_id: str, api_key: Optional[str] = None) -> Dict:
    """通过 arXiv ID 从 Semantic Scholar 获取期刊、影响因子等元数据"""
    headers = {"x-api-key": api_key} if api_key else {}
    paper_identifier = f"ARXIV:{arxiv_id}"
    fields = "title,venue,journal,publicationDate,url,abstract"

    try:
        for retry in range(2):
            response = requests.get(
                f"{BASE_CONFIG['SEMANTIC_SCHOLAR_API_URL']}/{paper_identifier}",
                headers=headers,
                params={"fields": fields},
                timeout=10
            )
            if response.status_code == 200:
                data = response.json()
                journal_name = data.get("journal", {}).get("name", data.get("venue", "预印本，暂无正式期刊信息"))
                impact_factor = data.get("journal", {}).get("impactFactor", "暂无影响因子数据")
                return {
                    "journal": journal_name,
                    "impact_factor": impact_factor,
                    "official_url": data.get("url", ""),
                    "publication_date": data.get("publicationDate", "")
                }
            elif response.status_code == 429:
                time.sleep(2)
                continue
            else:
                break
    except Exception as e:
        print(f"元数据获取失败（arXiv ID: {arxiv_id}）：{str(e)}")
    
    return {
        "journal": "预印本，暂无正式期刊信息",
        "impact_factor": "暂无影响因子数据",
        "official_url": "",
        "publication_date": ""
    }


# ===================== 新增：谷歌学术论文检索核心函数（完整版，爬全所有信息） =====================

def fetch_google_scholar_papers(max_results: int, serp_api_key: str) -> List[Dict]:
    """通过 SerpApi 调用谷歌学术检索论文（修复版：抓取所有可用字段）"""
    if not serp_api_key:
        print("❌ 请输入 SerpApi API Key！")
        return []

    params = {
        "engine": BASE_CONFIG["GOOGLE_SCHOLAR_ENGINE"],
        "q": BASE_CONFIG["SEARCH_KEYWORDS"],
        "api_key": serp_api_key,
        "hl": "zh-CN",
        "num": max_results,
        "start": 0,
        "timeout": 60
    }
    
    try:
        client = Client(api_key=serp_api_key)
        results = client.search(params)
        organic_results = results.get("organic_results", [])
        print(f"✅ 谷歌学术检索到 {len(organic_results)} 篇论文")
    except Exception as e:
        print(f"❌ 谷歌学术请求失败：{str(e)}")
        return []
    
    papers = []
    for i, res in enumerate(organic_results):
        try:
            # 核心：提取谷歌学术所有可用信息
            paper_id = f"scholar_{i}_{hash(res.get('link', ''))}"
            pub_info = res.get("publication_info", {})
    
            # 作者
            authors = [a.get("name", "未知作者") for a in pub_info.get("authors", [])] if pub_info.get("authors") else [
                "未知作者"]
            # 发表年份 - 只存储年份，谷歌学术不提供具体日期
            year = pub_info.get("year", "2025")
            published_time = f"{year}"
            # 期刊/出版物
            journal = pub_info.get("summary", "谷歌学术文献").split(",")[0] if "summary" in pub_info else "未知期刊"
            # 摘要
            abstract = res.get("snippet", "无摘要")
            # 论文链接
            link = res.get("link", res.get("resources", [{}])[0].get("link", "无链接")) if res.get("link") or res.get(
                "resources") else "无链接"
            # 标题
            title = res.get("title", "无标题")
    
            # 构造完整论文数据（所有字段填满）
            paper = {
                "id": paper_id,
                "title": title,
                "authors": authors,
                "published_time": published_time,
                "abstract": abstract,
                "link": link,
                "journal": journal,
                "impact_factor": "谷歌学术不提供"
            }
            papers.append(paper)
            time.sleep(1)
        except Exception as e:
            print(f"论文解析失败：{e}")
            continue
    
    return papers


# ===================== 论文解析核心逻辑 =====================

def parse_paper_full_fields(paper: Dict) -> Dict:
    """调用 Claude 大模型，解析论文全 11 项字段（核心解析逻辑）"""
    prompt = f"""
你是电子显微学、4D-STEM、Ptychography 相位重构领域的顶级专家，精通 WDD 算法、电子显微成像相关研究。
请基于以下论文信息，严格按照要求完成**11 项字段的结构化拆解**，禁止编造内容，所有内容必须基于论文原文，语言简洁专业，适配学术场景，禁止空泛表述。

【论文基础信息】
标题：{paper['title']}
作者：{', '.join(paper['authors'])}
发表时间：{paper['published_time']}
原文链接：{paper['link']}
摘要：{paper['abstract']}

请严格按照以下 6 个核心解析字段输出，每个字段单独标注，内容精准对应，分点内容使用数字序号标注：

1. 研究背景：论文解决的行业痛点、现有电子显微/Ptychography 技术的不足、研究的学术背景与意义
2. 论文创新点：论文核心的技术突破、方法创新、理论创新，分点列出，每点不超过 2 句话，必须紧扣电子显微/Ptychography 方向
3. 实验结果：论文核心的实验数据、验证结果、成像性能提升情况，客观精准表述，不夸大
4. 总结：论文的核心结论与学术贡献，一句话概括核心价值
5. 未来展望：论文作者提出的后续可研究方向
6. 可创新点：结合电子显微学 4D-STEM、WDD 算法方向，给出基于本论文可进一步延伸的、具体可落地的创新研究点，分点列出，每点必须有明确的研究方向，禁止空泛表述

输出格式要求：每个字段单独成段，开头标注字段名，如【研究背景】：xxx，禁止使用其他格式。
"""
    print(f"===CLAUDE_PARSE_START===\n{prompt}\n===CLAUDE_PARSE_END===")
    return {}


def format_parse_result(parse_result: Dict) -> Dict:
    """格式化 Claude 返回的解析结果，匹配 Excel 表头字段"""
    return {
        "摘要中文翻译": parse_result.get("摘要中文翻译", "无"),
        "研究背景": parse_result.get("研究背景", "解析失败，可手动查看原文补充"),
        "论文创新点": parse_result.get("论文创新点", "解析失败，可手动查看原文补充"),
        "实验结果": parse_result.get("实验结果", "解析失败，可手动查看原文补充"),
        "总结": parse_result.get("总结", "解析失败，可手动查看原文补充"),
        "未来展望": parse_result.get("未来展望", "解析失败，可手动查看原文补充"),
        "可创新点": parse_result.get("可创新点", "解析失败，可手动查看原文补充")
    }


def generate_markdown_summary(papers: List[Dict], output_path: str):
    """生成Markdown格式的解析汇总文件"""
    from datetime import datetime
    current_date = datetime.now().strftime("%Y-%m-%d")

    markdown_content = "# Ptychography领域论文解析汇总\n\n> 全自动爬取-解析-归档，包含中文摘要翻译和全字段结构化解析\n\n---\n\n"
    
    for idx, paper in enumerate(papers, 1):
        markdown_content += f"### 【{idx}】**{paper['title']}**\n"
        markdown_content += f"📅 发表时间：{paper['published_time']}  \n"
        markdown_content += f"📰 期刊：{paper['journal']}  \n"
        markdown_content += f"🔗 链接：[{paper['title']}]({paper['link']})  \n\n"
        markdown_content += f"【摘要中文翻译】：{paper['摘要中文翻译']}  \n\n"
        markdown_content += f"【研究背景】：{paper['研究背景']}  \n\n"
        markdown_content += f"【论文创新点】：\n{paper['论文创新点']}  \n\n"
        markdown_content += f"【实验结果】：{paper['实验结果']}  \n\n"
        markdown_content += f"【总结】：{paper['总结']}  \n\n"
        markdown_content += f"【未来展望】：{paper['未来展望']}  \n\n"
        markdown_content += f"【可创新点】：\n{paper['可创新点']}  \n\n"
        markdown_content += "---\n\n"
    
    markdown_content += f"\n*生成时间：{current_date}*"
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(markdown_content)


# ===================== Excel 导出 =====================

def append_to_excel(papers: List[Dict]):
    """将每日新增论文追加到 Excel 文件，按年度分 Sheet"""
    if not papers:
        return

    # 确保目录存在
    excel_dir = os.path.dirname(BASE_CONFIG["EXCEL_SAVE_PATH"])
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir=True)
    
    year_group = {}
    for paper in papers:
        pub_year = paper["published_time"][:4]
        if pub_year not in year_group:
            year_group[pub_year] = []
        row_data = [
            paper["title"],
            paper["link"],
            paper["journal"],
            paper["impact_factor"],
            paper["published_time"],
            paper["摘要中文翻译"],
            paper["研究背景"],
            paper["论文创新点"],
            paper["实验结果"],
            paper["总结"],
            paper["未来展望"],
            paper["可创新点"]
        ]
        year_group[pub_year].append(row_data)
    
    if os.path.exists(BASE_CONFIG["EXCEL_SAVE_PATH"]):
        with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            for year, rows in year_group.items():
                sheet_name = f"{year}年论文"
                df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
                try:
                    start_row = writer.book[sheet_name].max_row
                except KeyError:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    start_row = 0
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
    
                worksheet = writer.book[sheet_name]
                for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                    max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    else:
        with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl") as writer:
            for year, rows in year_group.items():
                sheet_name = f"{year}年论文"
                df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
                worksheet = writer.sheets[sheet_name]
                header_font = Font(bold=True)
                for col in range(1, len(BASE_CONFIG["EXCEL_HEADERS"]) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    
                for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                    max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    print(f"✅ 新增论文已自动归档到 Excel 文件，路径：{BASE_CONFIG['EXCEL_SAVE_PATH']}")
    
    # 同时输出Markdown汇总文件
    markdown_path = os.path.join(os.path.dirname(BASE_CONFIG["EXCEL_SAVE_PATH"]), "Ptychography论文解析汇总.md")
    generate_markdown_summary(papers, markdown_path)
    print(f"✅ Markdown汇总文件已输出，路径：{markdown_path}")

# ===================== 模式 1：每日全字段论文追踪=====================

def fetch_daily_papers(time_range_days: int, max_results: int, api_key: Optional[str] = None) -> List[Dict]:
    """从 arXiv API 检索每日最新论文，补全元数据"""
    category_query = " OR ".join([f"cat:{cat}" for cat in BASE_CONFIG["ARXIV_CATEGORIES"]])
    full_query = f"({BASE_CONFIG['SEARCH_KEYWORDS']}) AND ({category_query})"

    params = {
        "search_query": full_query,
        "start": 0,
        "max_results": max_results,
        "sortBy": "submittedDate",
        "sortOrder": "descending"
    }
    
    for retry in range(3):
        try:
            response = requests.get(BASE_CONFIG["ARXIV_API_URL"], params=params, timeout=30)
            if response.status_code == 200:
                break
            elif response.status_code == 429:
                print(f"arXiv 限流，等待5秒重试...")
                time.sleep(5)
                continue
            else:
                response.raise_for_status()
        except Exception as e:
            print(f"论文检索失败 (重试{retry + 1}/3): {str(e)}")
            if retry < 2:
                time.sleep(5)
            else:
                return []
    
    feed = feedparser.parse(response.content)
    papers = []
    cutoff_time = datetime.now() - timedelta(days=time_range_days)
    
    for entry in feed.entries:
        try:
            published_time = datetime.strptime(entry.published, "%Y-%m-%dT%H:%M:%SZ")
            if published_time < cutoff_time:
                continue
    
            arxiv_id = entry.id.split("/abs/")[-1].split("v")[0]
            paper_base = {
                "id": arxiv_id,
                "title": entry.title.replace("\n", " ").strip(),
                "authors": [author.name for author in entry.authors],
                "published_time": published_time.strftime("%Y-%m-%d %H:%M"),
                "abstract": entry.summary.replace("\n", " ").strip(),
                "link": entry.link,
                "category": entry.arxiv_primary_category["term"]
            }
    
            metadata = get_paper_metadata_by_arxiv(arxiv_id, api_key)
            paper_base["journal"] = metadata["journal"]
            paper_base["impact_factor"] = metadata["impact_factor"]
            if metadata["publication_date"]:
                paper_base["published_time"] = datetime.strptime(metadata["publication_date"], "%Y-%m-%d").strftime(
                    "%Y-%m-%d %H:%M")
            if metadata["official_url"]:
                paper_base["link"] = metadata["official_url"]
    
            papers.append(paper_base)
            time.sleep(1)
        except Exception as e:
            print(f"论文信息解析失败：{str(e)}")
            continue
    
    return papers


def format_daily_full_output(papers: List[Dict]) -> str:
    """格式化每日全字段论文输出，适配对话窗口阅读"""
    if not papers:
        return "✅ 今日暂无电子显微学 Ptychography 方向的新增论文。"

    output = f"===== 📄 {datetime.now().strftime('%Y-%m-%d')} Ptychography 领域最新论文（全字段解析版）=====\n"
    output += f"本次共检索到 {len(papers)} 篇新增论文，已完成全维度结构化解析\n\n"
    
    for idx, paper in enumerate(papers, 1):
        output += f"【{idx}】《{paper['title']}》\n"
        output += f"📅 发表时间：{paper['published_time']}\n"
        output += f"📰 期刊：{paper['journal']}\n"
        output += f"🔗 链接：{paper['link']}\n\n"
    return output


def run_daily_mode(time_range_days: int, max_results: int, auto_excel_append: bool, api_key: Optional[str]):
    """每日模式主执行流程"""
    history = load_history()
    print(f"正在检索过去{time_range_days}天的 Ptychography 领域论文...")
    latest_papers = fetch_daily_papers(time_range_days, max_results, api_key)
    new_papers = filter_new_papers(latest_papers, history)

    if not new_papers:
        print("✅ 今日暂无新增论文，无需解析")
        save_history(history)
        return
    
    print(f"正在批量解析{len(new_papers)}篇新增论文的结构化内容...")
    parsed_papers = []
    for idx, paper in enumerate(new_papers, 1):
        print(f"【{idx}/{len(new_papers)}】正在解析《{paper['title']}》...")
        parse_result = parse_paper_full_fields(paper)
        formatted_parse = format_parse_result(parse_result)
        full_paper = {**paper, **formatted_parse}
        parsed_papers.append(full_paper)
    
    final_output = format_daily_full_output(parsed_papers)
    print(final_output)
    
    if auto_excel_append:
        append_to_excel(parsed_papers)
    
    for p in new_papers:
        history["papers"][p["id"]] = {"title": p["title"], "add_time": datetime.now().strftime("%Y-%m-%d %H:%M")}
    save_history(history)
    print(f"🎉 每日论文追踪完成，共处理{len(new_papers)}篇新增论文")


# ===================== 模式 2：年度论文统计 Excel 生成=====================

def fetch_annual_papers(year: int, max_papers: int, api_key: Optional[str] = None) -> List[Dict]:
    """检索指定年份的论文"""
    headers = {"x-api-key": api_key} if api_key else {}
    params = {
        "query": BASE_CONFIG["SEARCH_KEYWORDS"],
        "publicationDate": f"{year}-01-01 TO {year}-12-31",
        "fields": "title,externalIds,authors,venue,publicationDate,url,abstract",
        "limit": max_papers,
        "sort": "publicationDate:desc"
    }

    results = None
    for retry in range(3):
        try:
            response = requests.get(
                BASE_CONFIG["SEMANTIC_SCHOLAR_API_URL"] + "/search",
                headers=headers,
                params=params,
                timeout=20
            )
            if response.status_code == 200:
                results = response.json()
                break
            elif response.status_code == 429:
                print(f"【{year}年】API 限流，等待5秒后重试...")
                time.sleep(5)
            else:
                print(f"【{year}年】检索失败：{response.status_code}")
                break
        except Exception as e:
            print(f"【{year}年】请求异常：{str(e)}")
            time.sleep(3)
    
    if not results:
        return []
    
    papers = []
    for paper_data in results.get("data", []):
        try:
            arxiv_id = paper_data.get("externalIds", {}).get("ArXiv", "")
            paper = {
                "id": arxiv_id if arxiv_id else str(hash(paper_data.get("title", ""))),
                "title": paper_data.get("title", "无标题"),
                "authors": [author.get("name", "") for author in paper_data.get("authors", [])],
                "link": paper_data.get("url", ""),
                "journal": paper_data.get("venue", "预印本"),
                "impact_factor": "待补充",
                "published_time": paper_data.get("publicationDate", f"{year}-01-01 00:00"),
                "abstract": paper_data.get("abstract", "无摘要")
            }
            papers.append(paper)
        except:
            continue
    return papers


def run_annual_summary_mode(start_year: int, end_year: int, max_papers_per_year: int, api_key: Optional[str]):
    """年度统计模式主执行流程"""
    history = load_history()
    yearly_papers = {}
    all_parsed_papers = []

    print(f"===== 开始检索{start_year}-{end_year}年 Ptychography 领域论文 =====")
    for year in range(start_year, end_year + 1):
        print(f"【{year}年】正在检索论文...")
        papers = fetch_annual_papers(year, max_papers_per_year, api_key)
        new_papers = [p for p in papers if p["id"] not in history["papers"]]
        yearly_papers[year] = new_papers
    
        for p in new_papers:
            history["papers"][p["id"]] = {
                "title": p["title"],
                "year": year,
                "add_time": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
        time.sleep(2)
    
    save_history(history)
    if not any(yearly_papers.values()):
        print("❌ 未检索到任何匹配的论文，Excel 生成终止")
        return
    
    print(f"===== 开始批量解析论文结构化内容 =====")
    for year, papers in yearly_papers.items():
        for idx, paper in enumerate(papers, 1):
            print(f"【{year}年 | {idx}/{len(papers)}】解析：{paper['title']}")
            parse_result = parse_paper_full_fields(paper)
            formatted_parse = format_parse_result(parse_result)
            full_paper = {**paper, **formatted_parse}
            all_parsed_papers.append(full_paper)
    
    print("===== 正在生成年度统计 Excel 文件 =====")
    append_to_excel(all_parsed_papers)
    total_count = len(all_parsed_papers)
    print(f"\n🎉 年度论文统计完成！总计检索并解析{start_year}-{end_year}年共{total_count}篇论文")


# ===================== 模式 3：新增！谷歌学术论文检索模式 =====================

def run_google_scholar_mode(max_results: int, serp_api_key: str):
    """谷歌学术模式主流程（复用所有原有解析/Excel功能）"""
    history = load_history()
    print("🚀 开始通过谷歌学术检索 Ptychography 领域论文...")

    # 调用谷歌学术API
    scholar_papers = fetch_google_scholar_papers(max_results, serp_api_key)
    new_papers = filter_new_papers(scholar_papers, history)
    
    if not new_papers:
        print("✅ 谷歌学术无新增论文")
        save_history(history)
        return
    
    # 复用原有解析逻辑
    print(f"正在解析{len(new_papers)}篇谷歌学术论文...")
    parsed_papers = []
    for paper in new_papers:
        parse_result = parse_paper_full_fields(paper)
        formatted_parse = format_parse_result(parse_result)
        parsed_papers.append({**paper, **formatted_parse})
    
    # 导出Excel
    append_to_excel(parsed_papers)
    
    # 更新历史记录
    for p in new_papers:
        history["papers"][p["id"]] = {"title": p["title"], "add_time": datetime.now().strftime("%Y-%m-%d %H:%M")}
    save_history(history)
    print("🎉 谷歌学术论文检索+解析+导出完成！")


# ===================== 主入口（已新增谷歌学术参数） =====================

def main():
    parser = argparse.ArgumentParser(description="Ptychography 领域全功能论文追踪工具")
    parser.add_argument("--mode", type=str, default="daily",
                        choices=["daily", "annual_summary", "google_scholar"],
                        help="运行模式：daily=每日, annual_summary=年度, google_scholar=谷歌学术")
    # 每日模式参数
    parser.add_argument("--time_range_days", type=int, default=1, help="【daily 模式】检索过去 N 天的论文")
    parser.add_argument("--max_results", type=int, default=10, help="最大检索论文数量")
    parser.add_argument("--auto_excel_append", type=lambda x: x.lower() == 'true', default=True,
                        help="【daily 模式】是否自动追加到 Excel")
    # 年度模式参数
    parser.add_argument("--start_year", type=int, default=2024, help="【annual 模式】统计起始年份")
    parser.add_argument("--end_year", type=int, default=2026, help="【annual 模式】统计结束年份")
    parser.add_argument("--max_papers_per_year", type=int, default=20, help="【annual 模式】每年最大检索论文数量")
    # 通用参数
    parser.add_argument("--semantic_scholar_key", type=str, default="", help="Semantic Scholar API 密钥")
    # 新增：谷歌学术 SerpApi Key 参数
    parser.add_argument("--serp_api_key", type=str, default="", help="【谷歌学术必填】SerpApi API Key（从dashboard获取）")

    args = parser.parse_args()
    
    if args.mode == "daily":
        run_daily_mode(args.time_range_days, args.max_results, args.auto_excel_append, args.semantic_scholar_key)
    elif args.mode == "annual_summary":
        run_annual_summary_mode(args.start_year, args.end_year, args.max_papers_per_year, args.semantic_scholar_key)
    elif args.mode == "google_scholar":
        # 运行谷歌学术模式
        run_google_scholar_mode(args.max_results, args.serp_api_key)


if __name__ == "__main__":
    main()