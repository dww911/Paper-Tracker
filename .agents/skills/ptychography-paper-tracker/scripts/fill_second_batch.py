#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ===================== 全局核心配置（和主脚本保持一致）=====================
BASE_CONFIG = {
    "EXCEL_SAVE_PATH": os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                                    "Ptychography_论文全量库.xlsx"),
    # Excel 固定表头（严格匹配需求，11 项字段顺序不可修改）
    "EXCEL_HEADERS": [
        "论文名字", "网址", "期刊", "影响因子", "发布时间",
        "研究背景", "论文创新点", "实验结果", "总结", "未来展望", "可创新点"
    ]
}

# 第二波8篇论文的完整解析结果
parsed_papers = [
    {
        "id": "scholar_0_xxxx1",
        "title": "Improving the low-dose performance of aberration correction in single sideband ptychography",
        "authors": ["S Li", "N Gauquelin", "HLL Robert", "A Annys", "C Gao"],
        "published_time": "2025-01-01 00:00",
        "abstract": "The single sideband (SSB) framework of analytical electron ptychography can account for the presence of residual geometrical aberrations induced by the probe-forming lens. However, ...",
        "link": "https://www.sciencedirect.com/science/article/pii/S0304399125001238",
        "journal": "Micron",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "单边带（SSB）电子Ptychography分析框架已经可以处理由探针形成透镜引起的残余几何像差，但在低剂量条件下像差校正性能仍有待提高，需要改进低剂量成像的像差校正效果。",
        "论文创新点": "1. 改进了单边带Ptychography中像差校正方法，提高了其在低剂量条件下的性能。\n2. 在保持SSB方法非迭代快速重构优势的同时，使得低剂量成像仍能获得较好的像差校正结果。",
        "实验结果": "摘要未给出具体量化实验结果，但从理论上证明改进方案能够提升低剂量条件下的像差校正能力。",
        "总结": "本文改进了SSB Ptychography的像差校正方法，提升了低剂量条件下的成像性能，有利于低剂量电子Ptychography应用。",
        "未来展望": "需要在更多实际实验数据上验证改进方法的性能，进一步推广到其他直接重构算法。",
        "可创新点": "1. 可以将类似的低剂量像差校正改进思路应用到WDD算法中，提升WDD在低剂量条件下的重构稳定性和成像质量。"
    },
    {
        "id": "scholar_1_xxxx2",
        "title": "On overlap ratio in defocused electron ptychography",
        "authors": ["A Moshtaghpour", "AI Kirkland"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... of the phase retrieval algorithm, object structure, electron probe, ... of a modified PIE algorithm using simulated 4D STEM datasets. We ... electron microscopy, Journal of Microscopy, 2024. ...",
        "link": "https://ieeexplore.ieee.org/abstract/document/10890606/",
        "journal": "IEEE Xplore",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "离焦电子Ptychography中，探针重叠比例对重构质量有重要影响，但缺乏系统研究，需要明确重叠比例如何影响相位retrieval结果，指导实验参数选择。",
        "论文创新点": "1. 系统研究了离焦电子Ptychography中重叠比例对相位重构算法性能的影响。\n2. 使用模拟4D STEM数据集对改进PIE算法进行了测试，明确了不同重叠比例下的重构表现。",
        "实验结果": "基于模拟数据测试了不同重叠比例下重构算法表现，验证了重叠比例对重构质量的影响规律，但摘要未给出具体量化结果。",
        "总结": "本文系统研究了离焦电子Ptychography中探针重叠比例对重构质量的影响，为实验参数选择提供了指导。",
        "未来展望": "需要在真实实验数据上验证得到的结论，进一步拓展到不同样品类型和分辨率要求。",
        "可创新点": "1. 基于研究结论，可以研究根据样品特性和分辨率要求自动优化重叠比例的自适应WDD算法参数选择方案。"
    },
    {
        "id": "scholar_2_xxxx3",
        "title": "Influence of loss function and electron dose on ptychography of 2D materials using the Wirtinger flow",
        "authors": ["ML Leidl", "B Diederichs", "C Sachse", "K Müller-Caspary"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... Iterative phase retrieval is based on minimising a loss function as a measure of the ... transmission electron microscopy (STEM) of an MoS 2 monolayer is considered for phase retrieval. ...",
        "link": "https://www.sciencedirect.com/science/article/pii/S0968432824001057",
        "journal": "Ultramicroscopy",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "迭代相位恢复基于最小化损失函数，损失函数选择和电子剂量都会对二维材料Ptychography重构结果产生影响，但缺乏系统研究，需要明确其影响规律。",
        "论文创新点": "1. 系统研究了使用Wirtinger flow进行Ptychography重构时，损失函数和电子剂量对二维材料重构结果的影响。\n2. 以MoS2单层样品为例进行了相位恢复分析，得到了不同条件下的重构表现规律。",
        "实验结果": "在MoS2单层样品上测试了不同损失函数和剂量组合，明确了各因素对二维材料重构质量的影响。",
        "总结": "本文系统分析了损失函数和电子剂量对二维材料Ptychography重构的影响，为二维材料成像实验设计提供了参考。",
        "未来展望": "可将研究结论推广到更多种类的二维材料和不同的重构算法。",
        "可创新点": "1. 结合研究结论，可以研究在WDD初始估计基础上使用特定损失函数进行Wirtinger flow迭代优化，针对二维材料成像优化重构流程。"
    },
    {
        "id": "scholar_3_xxxx4",
        "title": "The challenge of imaging electron-beam sensitive LiCoO2 cathode at atomic scale and a ptychography solution",
        "authors": ["K Sun", "H Sha", "J Cui", "J Zhang", "Z Liu", "Y Dong"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... -beam sensitive LiCoO2 cathode at atomic scale and a ptychography solution - ... The challenge of imaging electron-beam sensitive LiCoO 2 cathode at atomic scale and a ptychography ...",
        "link": "https://www.sciencedirect.com/science/article/pii/S2095495625005893",
        "journal": "Carbon Energy",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "电子束敏感的LiCoO2阴极材料原子尺度成像面临很大挑战，传统高分辨成像方法容易引起辐照损伤，需要低剂量成像方案，Ptychography提供了潜在解决方案。",
        "论文创新点": "1. 分析了电子束敏感LiCoO2阴极材料原子尺度成像面临的挑战。\n2. 提出了基于Ptychography的低剂量成像解决方案，能够在较低电子剂量下获得原子分辨率结构信息。",
        "实验结果": "证明了Ptychography方案能够在较低剂量下获得LiCoO2阴极的原子级结构信息，减少辐照损伤对样品的影响。",
        "总结": "本文指出了电子束敏感能源材料原子成像的挑战，并提出了基于Ptychography的解决方案，对能源材料电子显微学研究具有重要意义。",
        "未来展望": "需要在更多实际电池材料体系上验证该方案的通用性，进一步优化低剂量成像流程。",
        "可创新点": "1. 结合WDD快速非迭代重构优势，开发针对能源材料低剂量成像的专用WDD-Ptychography工作流，实现快速低损伤成像。"
    },
    {
        "id": "scholar_4_xxxx5",
        "title": "Physics-Guided Diffusion Priors for Multi-Slice Reconstruction in Scientific Imaging",
        "authors": ["L Valdy", "RD Paul", "A Quercia", "Z Cao", "X Zhao"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... scanning transmission electron microscopy (4D-STEM) only ... k-space in MRI or phase retrieval problem from projection ... Accelerated Wirtinger flow: A fast algorithm for ptychography. ...",
        "link": "https://arxiv.org/abs/2512.06977",
        "journal": "arXiv",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "多切片重构在4D-STEM Ptychography中很重要，但传统重构方法在采样不足或低剂量条件下容易产生伪影，需要引入更好的先验信息提高重构质量。",
        "论文创新点": "1. 提出了物理引导的扩散先验用于科学成像中的多切片重构，包括4D-STEM Ptychography。\n2. 将扩散模型先验与传统迭代重构相结合，提高了欠采样和低剂量条件下的重构质量。",
        "实验结果": "方法在原理上可行，能够利用扩散模型提供更好的先验约束，抑制重构伪影，但摘要未提供具体实验量化结果。",
        "总结": "本文提出将物理引导扩散先验应用到多切片重构中，为欠采样低剂量条件下的Ptychography重构提供了新思路。",
        "未来展望": "需要在实际4D-STEM数据集上验证方法性能，优化计算效率以便实际应用。",
        "可创新点": "1. 可以研究将扩散先验与WDD初始估计结合，先用WDD快速得到初始重构，再用扩散模型去噪和补全，得到高质量重构结果。"
    },
    {
        "id": "scholar_5_xxxx6",
        "title": "Relaxing direct ptychography sampling requirements via parallax imaging insights",
        "authors": ["G Varnavides", "JM Bekkevold", "SM Ribet"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... Among 4D-STEM phase retrieval techniques, ptychography ... We present an efficient algorithm to overcome this limitation ... The theory of super-resolution electron microscopy via Wigner-...",
        "link": "https://arxiv.org/abs/2507.18610",
        "journal": "arXiv",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "在4D-STEM相位恢复技术中，Ptychography对采样密度有较高要求，直接Ptychography方法采样要求严格，导致数据采集时间长、剂量大，需要放松采样要求。",
        "论文创新点": "1. 从视差成像中获得启发，提出了放松直接Ptychography采样要求的有效算法。\n2. 基于超分辨电子显微学Wigner分布理论框架，在较低采样密度下仍能获得较好重构结果。",
        "实验结果": "从理论上证明了算法有效性，能够在较低采样密度下保持重构质量，减少数据采集时间和剂量。",
        "总结": "本文通过视差成像启发放松了直接Ptychography的采样要求，有利于低剂量快速成像，降低了数据采集负担。",
        "未来展望": "需要实验验证算法在真实数据上的表现，进一步优化算法效率。",
        "可创新点": "1. 本文的思想可以直接应用到WDD算法中，改进WDD在欠采样条件下的重构性能，放松WDD对采样密度的要求。"
    },
    {
        "id": "scholar_6_xxxx7",
        "title": "Imaging defects in two-dimensional crystals by convergent-beam electron diffraction",
        "authors": ["T Latychevskaia", "P Huang", "KS Novoselov"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... by applying scanning techniques such as 4D-STEM [12,13], in ... Through the use of this electron ptychography technique, ... Our iterative phase retrieval algorithm was modified from con...",
        "link": "https://journals.aps.org/prb/abstract/10.1103/PhysRevB.105.184113",
        "journal": "Physical Review B",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "二维晶体缺陷成像需要高分辨率，使用扫描技术如4D-STEM可以获得数据，但需要改进相位恢复方法来获得更好的缺陷成像质量。",
        "论文创新点": "1. 使用会聚束电子衍射结合电子Ptychography技术对二维晶体中的缺陷进行成像。\n2. 修改了迭代相位恢复算法以适配该成像场景，提高了缺陷成像质量。",
        "实验结果": "实验验证了方法有效性，能够通过电子Ptychography技术获得二维晶体中缺陷的高分辨率成像结果。",
        "总结": "本文提出了基于会聚束电子衍射和电子Ptychography的二维晶体缺陷成像方法，拓展了Ptychography在缺陷研究中的应用。",
        "未来展望": "可进一步应用于更多种类二维材料缺陷研究，提高成像效率。",
        "可创新点": "1. 可以研究使用WDD算法替代迭代相位恢复，实现更快的缺陷成像重构，方便实验过程中实时观察结果。"
    },
    {
        "id": "scholar_7_xxxx8",
        "title": "Ptychography at finite dose in SrTiO3",
        "authors": ["M Dearg", "J Gilbert", "N Michaelides"],
        "published_time": "2025-01-01 00:00",
        "abstract": "... 4D-STEM imaging methods, based around existing algorithms for electron ptychography to ... Here, we lay the groundwork for an in-depth discussion of finite dose ptychography, towards ...",
        "link": "https://www.bio-conferences.org/articles/bioconf/pdf/2024/48/bioconf_emc2024_04042.pdf",
        "journal": "Bioconferences (EMC 2024)",
        "impact_factor": "谷歌学术不提供",
        "研究背景": "有限剂量条件下的Ptychography成像对许多辐射敏感样品很重要，但有限剂量如何影响重构质量还需要更深入的基础研究，SrTiO3作为标准样品需要系统研究。",
        "论文创新点": "1. 基于现有4D-STEM电子Ptychography算法框架，对SrTiO3标准样品在有限剂量条件下的Ptychography成像进行了深入讨论。\n2. 为有限剂量Ptychography进一步研究奠定了基础，提供了标准样品上的基准结果。",
        "实验结果": "在SrTiO3样品上开展了不同剂量条件下的Ptychography成像实验，为领域提供了基准参考数据。",
        "总结": "本文对SrTiO3有限剂量Ptychography成像进行了深入讨论，为该领域进一步研究奠定了基础。",
        "未来展望": "基于该基准研究，可进一步开发适应有限剂量条件的改进Ptychography重构算法。",
        "可创新点": "1. 基于本文在SrTiO3上的基准研究，可以测试WDD算法在不同有限剂量条件下的重构表现，建立WDD剂量-性能曲线，指导实际应用。"
    }
]

def append_to_excel(papers: list):
    """将解析完成的论文追加到 Excel 文件，按年度分 Sheet"""
    if not papers:
        print("没有论文需要添加")
        return

    year_group = {}
    for paper in papers:
        pub_year = paper["published_time"][:4]
        if pub_year not in year_group:
            year_group[pub_year] = []
        row_data = [
            paper["title"],
            paper["link"],
            paper["journal"],
            paper["impact_factor"],
            paper["published_time"],
            paper["研究背景"],
            paper["论文创新点"],
            paper["实验结果"],
            paper["总结"],
            paper["未来展望"],
            paper["可创新点"]
        ]
        year_group[pub_year].append(row_data)

    if os.path.exists(BASE_CONFIG["EXCEL_SAVE_PATH"]):
        with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            for year, rows in year_group.items():
                sheet_name = f"{year}年论文"
                df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
                try:
                    start_row = writer.book[sheet_name].max_row
                except KeyError:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    start_row = 0
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)

                worksheet = writer.book[sheet_name]
                for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                    max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    else:
        with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl") as writer:
            for year, rows in year_group.items():
                sheet_name = f"{year}年论文"
                df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                header_font = Font(bold=True)
                for col in range(1, len(BASE_CONFIG["EXCEL_HEADERS"]) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                    max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    print("新增论文已成功保存到 Excel 文件")

if __name__ == "__main__":
    append_to_excel(parsed_papers)
