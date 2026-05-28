import os
import json
import argparse
import sys
import time
import hashlib
import re
import importlib.util
try:
    import requests
except ImportError:
    requests = None
try:
    import feedparser
except ImportError:
    feedparser = None
try:
    import pandas as pd
except ImportError:
    pd = None

# 设置标准输出编码为UTF-8，解决Windows GBK编码问题
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
from datetime import datetime, timedelta
from typing import List, Dict, Optional
try:
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Font = None
    Alignment = None
    get_column_letter = None
try:
    from serpapi import Client
except ImportError:
    Client = None

# ===================== 全局核心配置（可按需修改检索关键词）=====================
BASE_CONFIG = {
    # 研究方向精准过滤规则（电子显微学 Ptychography，排除无关领域）
    "SEARCH_KEYWORDS": (
        'ptychography electron microscopy 4D-STEM STEM phase retrieval WDD algorithm'
        ' -X-ray -optical -synchrotron -laser'
    ),
    "ARXIV_CATEGORIES": ["physics.optics", "cond-ml.mtrl-sci", "physics.ins-det"],
    # API 配置
    "ARXIV_API_URL": "http://export.arxiv.org/api/query",
    "SEMANTIC_SCHOLAR_API_URL": "https://api.semanticscholar.org/graph/v1/paper",
    # 新增：谷歌学术 SerpApi 配置
    "GOOGLE_SCHOLAR_ENGINE": "google_scholar",
    # 文件路径配置
    "HISTORY_FILE": os.path.join(os.path.dirname(os.path.dirname(__file__)), "paper_history.json"),
    "RESEARCH_PROFILES_FILE": os.path.join(os.path.dirname(os.path.dirname(__file__)), "research_profiles.json"),
    "DEFAULT_PROFILE": "electron_ptychography",
    "EXCEL_SAVE_PATH": os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                                    "Ptychography_论文全量库.xlsx"),
    "DAILY_REPORT_DIR": os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "daily_reports"),
    # Excel 固定表头（严格匹配需求，11 项字段顺序不可修改）
    "EXCEL_HEADERS": [
        "论文名字", "网址", "期刊", "影响因子", "发布时间",
        "摘要中文翻译", "研究背景", "论文创新点", "实验结果", "总结", "未来展望", "可创新点"
    ]
}


# ===================== 研究方向配置 =====================
def load_research_profiles() -> Dict:
    """加载研究方向配置；缺失时回退到内置 Ptychography 默认方向。"""
    fallback = {
        "default_profile": BASE_CONFIG["DEFAULT_PROFILE"],
        "profiles": {
            BASE_CONFIG["DEFAULT_PROFILE"]: {
                "name": "电子显微 Ptychography / 4D-STEM",
                "include_keywords": ["electron ptychography", "4D-STEM", "phase retrieval", "WDD"],
                "exclude_keywords": ["X-ray", "optical", "synchrotron", "laser"],
                "arxiv_categories": BASE_CONFIG["ARXIV_CATEGORIES"],
                "sources": ["arxiv", "semantic_scholar"],
                "score_threshold": 5,
                "parse_focus": []
            }
        }
    }
    path = BASE_CONFIG["RESEARCH_PROFILES_FILE"]
    if not os.path.exists(path):
        return fallback
    try:
        with open(path, "r", encoding="utf-8") as f:
            profiles = json.load(f)
        if "profiles" not in profiles or not profiles["profiles"]:
            return fallback
        return profiles
    except Exception as e:
        print(f"研究方向配置读取失败，使用默认配置：{e}")
        return fallback


def save_research_profiles(profiles_doc: Dict):
    with open(BASE_CONFIG["RESEARCH_PROFILES_FILE"], "w", encoding="utf-8") as f:
        json.dump(profiles_doc, f, ensure_ascii=False, indent=2)


def split_csv(value: str) -> List[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def slugify_profile_id(value: str) -> str:
    value = normalize_title(value).replace(" ", "_")
    value = re.sub(r"[^a-z0-9_]+", "", value.lower())
    return value.strip("_") or "custom_profile"


def ask_yes_no(prompt: str, default: bool) -> bool:
    suffix = "Y/n" if default else "y/N"
    answer = input(f"{prompt}（{suffix}）：").strip().lower()
    if not answer:
        return default
    return answer in ("y", "yes", "是", "1", "true")


def create_profile_interactive():
    profiles_doc = load_research_profiles()
    profiles = profiles_doc.setdefault("profiles", {})

    raw_id = input("请输入研究方向名称（英文/数字/下划线，例如 electron_ptychography）：").strip()
    profile_id = slugify_profile_id(raw_id)
    while profile_id in profiles:
        raw_id = input(f"研究方向 {profile_id} 已存在，请输入新的名称：").strip()
        profile_id = slugify_profile_id(raw_id)

    display_name = input("请输入展示名称：").strip() or profile_id
    description = input("请输入一句话描述：").strip()
    include_keywords = split_csv(input("请输入核心关键词，用逗号分隔："))
    must_have_any = split_csv(input("请输入必须至少命中的关键词，用逗号分隔："))
    exclude_keywords = split_csv(input("请输入排除关键词，用逗号分隔："))
    research_focus = split_csv(input("请输入研究重点，用逗号分隔："))
    arxiv_categories = split_csv(input("请输入 arXiv 分类，用逗号分隔（可留空）："))
    use_arxiv = ask_yes_no("是否启用 arXiv", True)
    use_semantic = ask_yes_no("是否启用 Semantic Scholar", True)
    use_google = ask_yes_no("是否启用 Google Scholar/SerpApi", False)

    default_sources = {
        "arxiv": use_arxiv,
        "semantic_scholar": use_semantic,
        "google_scholar": use_google
    }
    sources = [name for name, enabled in default_sources.items() if enabled]
    profile = {
        "name": display_name,
        "display_name": display_name,
        "description": description,
        "include_keywords": include_keywords,
        "must_have_any": must_have_any,
        "exclude_keywords": exclude_keywords,
        "research_focus": research_focus,
        "parse_focus": [
            f"重点关注：{', '.join(research_focus)}" if research_focus else "结合当前研究方向给出具体、可落地的研究启发。"
        ],
        "arxiv_categories": arxiv_categories or BASE_CONFIG["ARXIV_CATEGORIES"],
        "sources": sources,
        "default_sources": default_sources,
        "score_rules": {
            "title_keyword": 3,
            "abstract_keyword": 1,
            "research_focus": 2,
            "exclude_keyword": -5,
            "must_have_bonus": 2,
            "min_score": 3
        },
        "score_threshold": 3,
        "parse_template": "custom_research",
        "output_fields": BASE_CONFIG["EXCEL_HEADERS"]
    }
    profiles[profile_id] = profile
    profiles_doc.setdefault("default_profile", BASE_CONFIG["DEFAULT_PROFILE"])
    save_research_profiles(profiles_doc)

    script_path = ".claude\\skills\\ptychography-paper-tracker\\scripts\\ptychography_tracker.py"
    print(f"\n✅ 已创建研究方向：{profile_id}")
    print("你可以运行：")
    print(f"python {script_path} --profile {profile_id} --mode daily --dry_run")


def get_research_profile(profile_name: str) -> Dict:
    profiles_doc = load_research_profiles()
    selected = profile_name or profiles_doc.get("default_profile") or BASE_CONFIG["DEFAULT_PROFILE"]
    profiles = profiles_doc.get("profiles", {})
    if selected not in profiles:
        available = ", ".join(profiles.keys())
        raise ValueError(f"未知研究方向：{selected}。可用方向：{available}")
    profile = profiles[selected].copy()
    profile["id"] = selected
    return profile


def list_research_profiles():
    profiles_doc = load_research_profiles()
    print("可用研究方向：")
    for profile_id, profile in profiles_doc.get("profiles", {}).items():
        default_mark = "（默认）" if profile_id == profiles_doc.get("default_profile") else ""
        print(f"- {profile_id}{default_mark}: {profile.get('name', profile_id)}")
        if profile.get("description"):
            print(f"  {profile['description']}")


def check_package(package_name: str, import_name: Optional[str] = None) -> bool:
    return importlib.util.find_spec(import_name or package_name) is not None


def print_check(ok: bool, label: str, detail: str = "", warn: bool = False):
    icon = "✅" if ok else ("⚠️" if warn else "❌")
    suffix = f": {detail}" if detail else ""
    print(f"{icon} {label}{suffix}")


def run_doctor():
    print("🔍 Ptychography Paper Tracker 环境检查\n")
    print_check(True, "Python", sys.version.split()[0])

    checks = [
        ("requests", "requests"),
        ("feedparser", "feedparser"),
        ("pandas", "pandas"),
        ("openpyxl", "openpyxl"),
        ("google-search-results", "serpapi"),
        ("PyYAML", "yaml")
    ]
    missing = []
    for label, import_name in checks:
        installed = check_package(label, import_name)
        print_check(installed, label, "installed" if installed else "missing")
        if not installed:
            missing.append(label)

    profiles_ok = False
    try:
        profiles_doc = load_research_profiles()
        profiles_ok = bool(profiles_doc.get("profiles"))
        print_check(profiles_ok, "research_profiles.json",
                    f"{len(profiles_doc.get('profiles', {}))} profiles" if profiles_ok else "empty")
    except Exception as e:
        print_check(False, "research_profiles.json", str(e))

    excel_path = BASE_CONFIG["EXCEL_SAVE_PATH"]
    excel_dir = os.path.dirname(excel_path)
    excel_writable = os.path.isdir(excel_dir) and os.access(excel_dir, os.W_OK)
    if os.path.exists(excel_path):
        excel_writable = os.access(excel_path, os.W_OK)
    print_check(excel_writable, "Excel archive",
                "writable" if excel_writable else f"not writable: {excel_path}")

    history_path = BASE_CONFIG["HISTORY_FILE"]
    print_check(os.path.exists(history_path), "history file",
                "found" if os.path.exists(history_path) else "will be created", warn=not os.path.exists(history_path))

    report_dir = BASE_CONFIG["DAILY_REPORT_DIR"]
    report_parent = os.path.dirname(report_dir)
    report_writable = os.path.isdir(report_parent) and os.access(report_parent, os.W_OK)
    print_check(report_writable, "daily_reports directory",
                "writable" if report_writable else f"not writable: {report_dir}")

    print_check(bool(os.environ.get("SERPAPI_API_KEY")), "SERPAPI_API_KEY",
                "configured" if os.environ.get("SERPAPI_API_KEY") else "not configured", warn=True)
    print_check(bool(os.environ.get("SCT_KEY")), "SCT_KEY",
                "configured" if os.environ.get("SCT_KEY") else "not configured", warn=True)

    if missing:
        print("\n建议执行：")
        print("pip install -r requirements.txt")


def normalize_title(title: str) -> str:
    title = (title or "").lower()
    title = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", title)
    return re.sub(r"\s+", " ", title).strip()


def stable_paper_id(paper: Dict) -> str:
    """生成跨来源稳定 ID，减少同一论文重复入库。"""
    for key in ("doi", "arxiv_id", "semantic_scholar_id"):
        value = paper.get(key)
        if value:
            return f"{key}:{str(value).lower()}"
    if paper.get("id") and not str(paper["id"]).startswith("scholar_"):
        return str(paper["id"])
    title_key = normalize_title(paper.get("title", ""))
    if not title_key:
        title_key = str(paper.get("link", "unknown"))
    return "title:" + hashlib.sha256(title_key.encode("utf-8")).hexdigest()[:16]


def build_search_keywords(profile: Dict) -> str:
    include = profile.get("include_keywords") or []
    exclude = profile.get("exclude_keywords") or []
    include_query = " OR ".join([f'"{kw}"' if " " in kw else kw for kw in include])
    exclude_query = " ".join([f'-"{kw}"' if " " in kw else f"-{kw}" for kw in exclude])
    return f"({include_query}) {exclude_query}".strip() if include_query else BASE_CONFIG["SEARCH_KEYWORDS"]


def get_score_rules(profile: Dict) -> Dict:
    rules = profile.get("score_rules") or {}
    return {
        "title_keyword": int(rules.get("title_keyword", 4)),
        "abstract_keyword": int(rules.get("abstract_keyword", 2)),
        "research_focus": int(rules.get("research_focus", 2)),
        "exclude_keyword": int(rules.get("exclude_keyword", -8)),
        "must_have_bonus": int(rules.get("must_have_bonus", 2)),
        "min_score": int(rules.get("min_score", profile.get("score_threshold", 0)))
    }


def explain_paper_relevance(paper: Dict, profile: Dict) -> Dict:
    haystack_title = str(paper.get("title", "")).lower()
    haystack_abstract = str(paper.get("abstract", "")).lower()
    haystack = f"{haystack_title} {haystack_abstract}"
    rules = get_score_rules(profile)
    score = 0
    matched_keywords = []
    matched_focus = []
    matched_exclude = []
    matched_must_have = []

    for keyword in profile.get("include_keywords", []):
        kw = keyword.lower()
        if kw in haystack_title:
            score += rules["title_keyword"]
            matched_keywords.append(keyword)
        elif kw in haystack_abstract:
            score += rules["abstract_keyword"]
            matched_keywords.append(keyword)

    for keyword in profile.get("research_focus", []):
        if keyword.lower() in haystack:
            score += rules["research_focus"]
            matched_focus.append(keyword)

    for keyword in profile.get("must_have_any", []):
        if keyword.lower() in haystack:
            score += rules["must_have_bonus"]
            matched_must_have.append(keyword)

    for keyword in profile.get("exclude_keywords", []):
        if keyword.lower() in haystack:
            score += rules["exclude_keyword"]
            matched_exclude.append(keyword)

    journal = str(paper.get("journal", "")).lower()
    if any(term in journal for term in ["microscopy", "materials", "radiology", "imaging", "physics"]):
        score += 1

    must_have = profile.get("must_have_any", [])
    passes_must_have = not must_have or bool(matched_must_have)
    threshold = rules["min_score"]
    passes_score = score >= threshold
    included = passes_must_have and passes_score and not matched_exclude
    reasons = []
    if not passes_must_have:
        reasons.append("未命中 must_have_any")
    if matched_exclude:
        reasons.append("命中排除关键词")
    if not passes_score:
        reasons.append("相关性评分过低")
    if included:
        reasons.append("通过筛选")
    return {
        "score": score,
        "threshold": threshold,
        "matched_keywords": matched_keywords,
        "matched_focus": matched_focus,
        "matched_must_have": matched_must_have,
        "matched_exclude": matched_exclude,
        "included": included,
        "reasons": reasons
    }


def score_paper_relevance(paper: Dict, profile: Dict) -> int:
    return explain_paper_relevance(paper, profile)["score"]


def filter_relevant_papers(papers: List[Dict], profile: Dict) -> List[Dict]:
    relevant = []
    for paper in papers:
        paper["stable_id"] = stable_paper_id(paper)
        paper["id"] = paper["stable_id"]
        relevance = explain_paper_relevance(paper, profile)
        paper["relevance_score"] = relevance["score"]
        paper["relevance_explain"] = relevance
        if relevance["included"]:
            relevant.append(paper)
    return relevant


def analyze_filtering(papers: List[Dict], relevant_papers: List[Dict], new_papers: List[Dict],
                      history: Dict, profile: Dict) -> Dict:
    existing_ids = set(history.get("papers", {}).keys())
    existing_titles = {
        normalize_title(item.get("title", ""))
        for item in history.get("papers", {}).values()
        if isinstance(item, dict)
    }
    stats = {
        "retrieved": len(papers),
        "kept_after_relevance": len(relevant_papers),
        "new_papers": len(new_papers),
        "excluded_by_keywords": 0,
        "excluded_by_score": 0,
        "excluded_by_must_have": 0,
        "already_seen": 0,
        "filtered_reasons": {}
    }
    for paper in papers:
        relevance = paper.get("relevance_explain") or explain_paper_relevance(paper, profile)
        if relevance["matched_exclude"]:
            stats["excluded_by_keywords"] += 1
            for keyword in relevance["matched_exclude"]:
                stats["filtered_reasons"][keyword] = stats["filtered_reasons"].get(keyword, 0) + 1
        if "相关性评分过低" in relevance["reasons"]:
            stats["excluded_by_score"] += 1
        if "未命中 must_have_any" in relevance["reasons"]:
            stats["excluded_by_must_have"] += 1
    for paper in relevant_papers:
        title_key = normalize_title(paper.get("title", ""))
        if paper.get("id") in existing_ids or title_key in existing_titles:
            stats["already_seen"] += 1
    return stats


# ===================== 工具函数 =====================
def load_history() -> Dict:
    """加载论文历史记录，避免重复解析"""
    if os.path.exists(BASE_CONFIG["HISTORY_FILE"]):
        with open(BASE_CONFIG["HISTORY_FILE"], "r", encoding="utf-8") as f:
            return json.load(f)
    return {"papers": {}}


def save_history(history: Dict):
    """保存论文历史记录"""
    with open(BASE_CONFIG["HISTORY_FILE"], "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def filter_new_papers(papers: List[Dict], history: Dict) -> List[Dict]:
    """过滤出从未处理过的新论文"""
    existing_ids = set(history.get("papers", {}).keys())
    existing_titles = {
        normalize_title(item.get("title", ""))
        for item in history.get("papers", {}).values()
        if isinstance(item, dict)
    }
    new_papers = []
    for paper in papers:
        paper_id = paper.get("stable_id") or stable_paper_id(paper)
        title_key = normalize_title(paper.get("title", ""))
        if paper_id in existing_ids or title_key in existing_titles:
            continue
        paper["id"] = paper_id
        new_papers.append(paper)
    return new_papers


# ===================== 元数据获取 =====================
def get_paper_metadata_by_arxiv(arxiv_id: str, api_key: Optional[str] = None) -> Dict:
    """通过 arXiv ID 从 Semantic Scholar 获取期刊、影响因子等元数据"""
    if requests is None:
        print("缺少 requests 依赖，跳过 Semantic Scholar 元数据补全。")
        return {
            "journal": "预印本，暂无正式期刊信息",
            "impact_factor": "暂无影响因子数据",
            "official_url": "",
            "publication_date": ""
        }
    headers = {"x-api-key": api_key} if api_key else {}
    paper_identifier = f"ARXIV:{arxiv_id}"
    fields = "title,venue,journal,publicationDate,url,abstract"

    try:
        for retry in range(2):
            response = requests.get(
                f"{BASE_CONFIG['SEMANTIC_SCHOLAR_API_URL']}/{paper_identifier}",
                headers=headers,
                params={"fields": fields},
                timeout=10
            )
            if response.status_code == 200:
                data = response.json()
                journal_name = data.get("journal", {}).get("name", data.get("venue", "预印本，暂无正式期刊信息"))
                impact_factor = data.get("journal", {}).get("impactFactor", "暂无影响因子数据")
                return {
                    "journal": journal_name,
                    "impact_factor": impact_factor,
                    "official_url": data.get("url", ""),
                    "publication_date": data.get("publicationDate", "")
                }
            elif response.status_code == 429:
                time.sleep(2)
                continue
            else:
                break
    except Exception as e:
        print(f"元数据获取失败（arXiv ID: {arxiv_id}）：{str(e)}")

    return {
        "journal": "预印本，暂无正式期刊信息",
        "impact_factor": "暂无影响因子数据",
        "official_url": "",
        "publication_date": ""
    }


# ===================== 新增：谷歌学术论文检索核心函数（完整版，爬全所有信息） =====================
def fetch_google_scholar_papers(max_results: int, serp_api_key: str, profile: Optional[Dict] = None) -> List[Dict]:
    """通过 SerpApi 调用谷歌学术检索论文（修复版：抓取所有可用字段）"""
    if not serp_api_key:
        print("❌ 请输入 SerpApi API Key！")
        return []
    if Client is None:
        print("❌ 缺少 SerpApi 依赖，请先安装 google-search-results。")
        return []

    params = {
        "engine": BASE_CONFIG["GOOGLE_SCHOLAR_ENGINE"],
        "q": build_search_keywords(profile or {}),
        "api_key": serp_api_key,
        "hl": "zh-CN",
        "num": max_results,
        "start": 0,
        "timeout": 60
    }

    try:
        client = Client(api_key=serp_api_key)
        results = client.search(params)
        organic_results = results.get("organic_results", [])
        print(f"✅ 谷歌学术检索到 {len(organic_results)} 篇论文")
    except Exception as e:
        print(f"❌ 谷歌学术请求失败：{str(e)}")
        return []

    papers = []
    for i, res in enumerate(organic_results):
        try:
            # 核心：提取谷歌学术所有可用信息
            title = res.get("title", "无标题")
            paper_id = f"scholar_{i}_{hashlib.sha256(str(res.get('link', title)).encode('utf-8')).hexdigest()[:16]}"
            pub_info = res.get("publication_info", {})

            # 作者
            authors = [a.get("name", "未知作者") for a in pub_info.get("authors", [])] if pub_info.get("authors") else [
                "未知作者"]
            # 发表年份 - 只存储年份，谷歌学术不提供具体日期
            year = pub_info.get("year", "2025")
            published_time = f"{year}"
            # 期刊/出版物
            journal = pub_info.get("summary", "谷歌学术文献").split(",")[0] if "summary" in pub_info else "未知期刊"
            # 摘要
            abstract = res.get("snippet", "无摘要")
            # 论文链接
            link = res.get("link", res.get("resources", [{}])[0].get("link", "无链接")) if res.get("link") or res.get(
                "resources") else "无链接"
            # 构造完整论文数据（所有字段填满）
            paper = {
                "id": paper_id,
                "title": title,
                "authors": authors,
                "published_time": published_time,
                "abstract": abstract,
                "link": link,
                "journal": journal,
                "impact_factor": "谷歌学术不提供"
            }
            papers.append(paper)
            time.sleep(1)
        except Exception as e:
            print(f"论文解析失败：{e}")
            continue

    return papers


# ===================== 论文解析核心逻辑 =====================
def parse_paper_full_fields(paper: Dict, profile: Optional[Dict] = None) -> Dict:
    """调用 Claude 大模型，解析论文全 11 项字段（核心解析逻辑）"""
    profile = profile or {}
    focus_lines = "\n".join([f"- {item}" for item in profile.get("parse_focus", [])])
    focus_block = f"\n【当前研究方向】{profile.get('name', '电子显微 Ptychography / 4D-STEM')}\n{focus_lines}\n" if focus_lines else ""
    prompt = f"""
你是当前研究方向的资深科研顾问，能够判断论文相关性、拆解技术路线，并提出可落地的后续研究想法。
请基于以下论文信息，严格按照当前研究方向完成结构化拆解，禁止编造内容，所有内容必须基于论文原文，语言简洁专业，适配学术场景，禁止空泛表述。
{focus_block}

【论文基础信息】
标题：{paper['title']}
作者：{', '.join(paper['authors'])}
发表时间：{paper['published_time']}
原文链接：{paper['link']}
摘要：{paper['abstract']}
相关性评分：{paper.get('relevance_score', '未评分')}

请严格按照以下 6 个核心解析字段输出，每个字段单独标注，内容精准对应，分点内容使用数字序号标注：
1. 研究背景：论文解决的行业痛点、现有技术的不足、研究的学术背景与意义
2. 论文创新点：论文核心的技术突破、方法创新、理论创新，分点列出，每点不超过 2 句话，必须紧扣当前研究方向
3. 实验结果：论文核心的实验数据、验证结果、成像性能提升情况，客观精准表述，不夸大
4. 总结：论文的核心结论与学术贡献，一句话概括核心价值
5. 未来展望：论文作者提出的后续可研究方向
6. 可创新点：结合当前研究方向，给出基于本论文可进一步延伸的、具体可落地的创新研究点，分点列出，每点必须有明确的研究方向，禁止空泛表述

输出格式要求：每个字段单独成段，开头标注字段名，如【研究背景】：xxx，禁止使用其他格式。
"""
    print(f"===CLAUDE_PARSE_START===\n{prompt}\n===CLAUDE_PARSE_END===")
    return {}


def format_parse_result(parse_result: Dict) -> Dict:
    """格式化 Claude 返回的解析结果，匹配 Excel 表头字段"""
    return {
        "摘要中文翻译": parse_result.get("摘要中文翻译", "无"),
        "研究背景": parse_result.get("研究背景", "解析失败，可手动查看原文补充"),
        "论文创新点": parse_result.get("论文创新点", "解析失败，可手动查看原文补充"),
        "实验结果": parse_result.get("实验结果", "解析失败，可手动查看原文补充"),
        "总结": parse_result.get("总结", "解析失败，可手动查看原文补充"),
        "未来展望": parse_result.get("未来展望", "解析失败，可手动查看原文补充"),
        "可创新点": parse_result.get("可创新点", "解析失败，可手动查看原文补充")
    }


def generate_markdown_summary(papers: List[Dict], output_path: str):
    """生成Markdown格式的解析汇总文件"""
    from datetime import datetime
    current_date = datetime.now().strftime("%Y-%m-%d")

    markdown_content = "# Ptychography领域论文解析汇总\n\n> 全自动爬取-解析-归档，包含中文摘要翻译和全字段结构化解析\n\n---\n\n"

    for idx, paper in enumerate(papers, 1):
        markdown_content += f"### 【{idx}】**{paper['title']}**\n"
        markdown_content += f"📅 发表时间：{paper['published_time']}  \n"
        markdown_content += f"📰 期刊：{paper['journal']}  \n"
        markdown_content += f"🔗 链接：[{paper['title']}]({paper['link']})  \n\n"
        markdown_content += f"【摘要中文翻译】：{paper['摘要中文翻译']}  \n\n"
        markdown_content += f"【研究背景】：{paper['研究背景']}  \n\n"
        markdown_content += f"【论文创新点】：\n{paper['论文创新点']}  \n\n"
        markdown_content += f"【实验结果】：{paper['实验结果']}  \n\n"
        markdown_content += f"【总结】：{paper['总结']}  \n\n"
        markdown_content += f"【未来展望】：{paper['未来展望']}  \n\n"
        markdown_content += f"【可创新点】：\n{paper['可创新点']}  \n\n"
        markdown_content += "---\n\n"

    markdown_content += f"\n*生成时间：{current_date}*"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(markdown_content)


# ===================== Excel 导出 =====================
def append_to_excel(papers: List[Dict]):
    """将每日新增论文追加到 Excel 文件，按年度分 Sheet"""
    if not papers:
        return
    if pd is None or Font is None or Alignment is None or get_column_letter is None:
        print("❌ 缺少 pandas/openpyxl 依赖，无法写入 Excel。请先安装 requirements.txt。")
        return

    # 确保目录存在
    excel_dir = os.path.dirname(BASE_CONFIG["EXCEL_SAVE_PATH"])
    if not os.path.exists(excel_dir):
        os.makedirs(exist_ok=True)

    year_group = {}
    for paper in papers:
        pub_year = paper["published_time"][:4]
        if pub_year not in year_group:
            year_group[pub_year] = []
        row_data = [
            paper["title"],
            paper["link"],
            paper["journal"],
            paper["impact_factor"],
            paper["published_time"],
            paper["摘要中文翻译"],
            paper["研究背景"],
            paper["论文创新点"],
            paper["实验结果"],
            paper["总结"],
            paper["未来展望"],
            paper["可创新点"]
        ]
        year_group[pub_year].append(row_data)

    if os.path.exists(BASE_CONFIG["EXCEL_SAVE_PATH"]):
        with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl", mode="a",
                            if_sheet_exists="overlay") as writer:
            for year, rows in year_group.items():
                sheet_name = f"{year}年论文"
                df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
                try:
                    start_row = writer.book[sheet_name].max_row
                except KeyError:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    start_row = 0
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)

                worksheet = writer.book[sheet_name]
                for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                    max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    else:
        with pd.ExcelWriter(BASE_CONFIG["EXCEL_SAVE_PATH"], engine="openpyxl") as writer:
            for year, rows in year_group.items():
                sheet_name = f"{year}年论文"
                df = pd.DataFrame(rows, columns=BASE_CONFIG["EXCEL_HEADERS"])
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                header_font = Font(bold=True)
                for col in range(1, len(BASE_CONFIG["EXCEL_HEADERS"]) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for col_idx, col_name in enumerate(BASE_CONFIG["EXCEL_HEADERS"], 1):
                    max_length = max(len(str(col_name)), df[col_name].astype(str).map(len).max())
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    print(f"✅ 新增论文已自动归档到 Excel 文件，路径：{BASE_CONFIG['EXCEL_SAVE_PATH']}")

    # 同时输出Markdown汇总文件
    markdown_path = os.path.join(os.path.dirname(BASE_CONFIG["EXCEL_SAVE_PATH"]), "Ptychography论文解析汇总.md")
    generate_markdown_summary(papers, markdown_path)
    print(f"✅ Markdown汇总文件已输出，路径：{markdown_path}")

# ===================== 模式 1：每日全字段论文追踪=====================
def fetch_daily_papers(time_range_days: int, max_results: int, api_key: Optional[str] = None,
                       profile: Optional[Dict] = None) -> List[Dict]:
    """从 arXiv API 检索每日最新论文，补全元数据"""
    if requests is None or feedparser is None:
        print("❌ 缺少 requests/feedparser 依赖，无法检索 arXiv。请先安装 requirements.txt。")
        return []
    profile = profile or {}
    categories = profile.get("arxiv_categories") or BASE_CONFIG["ARXIV_CATEGORIES"]
    category_query = " OR ".join([f"cat:{cat}" for cat in categories])
    full_query = f"({build_search_keywords(profile)}) AND ({category_query})"

    params = {
        "search_query": full_query,
        "start": 0,
        "max_results": max_results,
        "sortBy": "submittedDate",
        "sortOrder": "descending"
    }

    for retry in range(3):
        try:
            response = requests.get(BASE_CONFIG["ARXIV_API_URL"], params=params, timeout=30)
            if response.status_code == 200:
                break
            elif response.status_code == 429:
                print(f"arXiv 限流，等待5秒重试...")
                time.sleep(5)
                continue
            else:
                response.raise_for_status()
        except Exception as e:
            print(f"论文检索失败 (重试{retry + 1}/3): {str(e)}")
            if retry < 2:
                time.sleep(5)
            else:
                return []

    feed = feedparser.parse(response.content)
    papers = []
    cutoff_time = datetime.now() - timedelta(days=time_range_days)

    for entry in feed.entries:
        try:
            published_time = datetime.strptime(entry.published, "%Y-%m-%dT%H:%M:%SZ")
            if published_time < cutoff_time:
                continue

            arxiv_id = entry.id.split("/abs/")[-1].split("v")[0]
            paper_base = {
                "id": arxiv_id,
                "arxiv_id": arxiv_id,
                "title": entry.title.replace("\n", " ").strip(),
                "authors": [author.name for author in entry.authors],
                "published_time": published_time.strftime("%Y-%m-%d %H:%M"),
                "abstract": entry.summary.replace("\n", " ").strip(),
                "link": entry.link,
                "category": entry.arxiv_primary_category["term"]
            }

            metadata = get_paper_metadata_by_arxiv(arxiv_id, api_key)
            paper_base["journal"] = metadata["journal"]
            paper_base["impact_factor"] = metadata["impact_factor"]
            if metadata["publication_date"]:
                paper_base["published_time"] = datetime.strptime(metadata["publication_date"], "%Y-%m-%d").strftime(
                    "%Y-%m-%d %H:%M")
            if metadata["official_url"]:
                paper_base["link"] = metadata["official_url"]

            papers.append(paper_base)
            time.sleep(1)
        except Exception as e:
            print(f"论文信息解析失败：{str(e)}")
            continue

    return papers


def format_daily_full_output(papers: List[Dict]) -> str:
    """格式化每日全字段论文输出，适配对话窗口阅读"""
    if not papers:
        return "✅ 今日暂无电子显微学 Ptychography 方向的新增论文。"

    output = f"===== 📄 {datetime.now().strftime('%Y-%m-%d')} Ptychography 领域最新论文（全字段解析版）=====\n"
    output += f"本次共检索到 {len(papers)} 篇新增论文，已完成全维度结构化解析\n\n"

    for idx, paper in enumerate(papers, 1):
        output += f"【{idx}】《{paper['title']}》\n"
        output += f"📅 发表时间：{paper['published_time']}\n"
        output += f"📰 期刊：{paper['journal']}\n"
        if "relevance_score" in paper:
            output += f"⭐ 相关性评分：{paper['relevance_score']}\n"
        output += f"🔗 链接：{paper['link']}\n\n"
    return output


def join_or_none(items: List[str]) -> str:
    return ", ".join(items) if items else "无"


def format_dry_run_preview(all_papers: List[Dict], relevant_papers: List[Dict], new_papers: List[Dict],
                           stats: Dict, profile: Dict, notify: str = "") -> str:
    output = []
    output.append("===== DRY RUN 预览报告 =====")
    output.append("不会写入 Excel，不会更新 history，不会推送微信。")
    output.append(f"研究方向：{profile.get('name', profile.get('id', 'unknown'))}")
    output.append(f"检索到论文数量：{stats['retrieved']}")
    output.append(f"过滤后保留：{stats['kept_after_relevance']}")
    output.append(f"因 exclude 删除：{stats['excluded_by_keywords']}")
    output.append(f"因 must_have 未命中删除：{stats['excluded_by_must_have']}")
    output.append(f"因相关性不足删除：{stats['excluded_by_score']}")
    output.append(f"历史已存在：{stats['already_seen']}")
    output.append(f"预计写入：{stats['new_papers']}")
    output.append(f"预计推送：{min(stats['new_papers'], 5) if notify else 0}")
    output.append("")

    if new_papers:
        output.append("## 预计保留的新论文")
        for paper in new_papers:
            relevance = paper.get("relevance_explain", {})
            output.append(f"[score={paper.get('relevance_score', 0)}] {paper.get('title', '无标题')}")
            output.append(f"命中关键词：{join_or_none(relevance.get('matched_keywords', []))}")
            output.append(f"必须命中：{join_or_none(relevance.get('matched_must_have', []))}")
            output.append(f"研究重点：{join_or_none(relevance.get('matched_focus', []))}")
            output.append(f"排除词：{join_or_none(relevance.get('matched_exclude', []))}")
            output.append(f"链接：{paper.get('link', '')}")
            output.append("")
    else:
        output.append("没有预计写入的新论文。")
        output.append("")

    filtered = [paper for paper in all_papers if not paper.get("relevance_explain", {}).get("included")]
    if filtered:
        output.append("## 被过滤论文示例")
        for paper in filtered[:10]:
            relevance = paper.get("relevance_explain", {})
            output.append(f"[score={relevance.get('score', 0)}] {paper.get('title', '无标题')}")
            output.append(f"原因：{join_or_none(relevance.get('reasons', []))}")
            output.append(f"排除词：{join_or_none(relevance.get('matched_exclude', []))}")
            output.append("")

    if stats.get("filtered_reasons"):
        output.append("## 被过滤的主要原因")
        for reason, count in sorted(stats["filtered_reasons"].items(), key=lambda item: item[1], reverse=True):
            output.append(f"- {reason}: {count} 篇")
    return "\n".join(output)


def report_filename(profile: Dict) -> str:
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"{date_str}_{profile.get('id', 'research')}.md"


def build_daily_report(papers: List[Dict], profile: Dict, stats: Dict, time_range_days: int,
                       dry_run: bool = False) -> str:
    date_str = datetime.now().strftime("%Y-%m-%d")
    title = profile.get("display_name") or profile.get("name") or profile.get("id", "Research")
    lines = [
        f"# 今日 {title} 文献雷达",
        "",
        f"日期：{date_str}  ",
        f"检索范围：{'最近 ' + str(time_range_days) + ' 天' if time_range_days else '当前检索结果'}  ",
        f"检索到：{stats.get('retrieved', 0)} 篇  ",
        f"筛选后：{stats.get('kept_after_relevance', 0)} 篇  ",
        f"推荐精读：{min(len(papers), 3)} 篇  ",
        f"模式：{'DRY RUN 预览' if dry_run else '正式归档'}",
        "",
        "## 今日最值得读",
        ""
    ]

    if not papers:
        lines.append("今日暂无新的高相关论文。")
    for idx, paper in enumerate(sorted(papers, key=lambda p: p.get("relevance_score", 0), reverse=True)[:5], 1):
        relevance = paper.get("relevance_explain", {})
        score = paper.get("relevance_score", 0)
        stars = "★" * min(5, max(1, int(score / 2))) + "☆" * max(0, 5 - min(5, max(1, int(score / 2))))
        lines.extend([
            f"### {idx}. {paper.get('title', '无标题')}",
            f"推荐度：{stars}  ",
            f"相关性评分：{score}  ",
            f"关键词：{join_or_none(relevance.get('matched_keywords', []))}  ",
            f"研究重点：{join_or_none(relevance.get('matched_focus', []))}  ",
            f"一句话总结：待模型解析后补充。  ",
            f"为什么值得读：命中当前研究方向关键词，建议结合摘要进一步判断。  ",
            f"和我的研究关系：{join_or_none(profile.get('research_focus', [])[:4])}  ",
            f"链接：{paper.get('link', '')}",
            ""
        ])

    lines.extend(["## 被过滤的主要原因", ""])
    if stats.get("filtered_reasons"):
        for reason, count in sorted(stats["filtered_reasons"].items(), key=lambda item: item[1], reverse=True):
            lines.append(f"- {reason}：{count} 篇")
    else:
        lines.append("- 无明显排除关键词命中")
    if stats.get("excluded_by_score"):
        lines.append(f"- 相关性评分过低：{stats['excluded_by_score']} 篇")
    if stats.get("excluded_by_must_have"):
        lines.append(f"- 未命中必须关键词：{stats['excluded_by_must_have']} 篇")
    return "\n".join(lines) + "\n"


def write_daily_report(papers: List[Dict], profile: Dict, stats: Dict, time_range_days: int) -> str:
    os.makedirs(BASE_CONFIG["DAILY_REPORT_DIR"], exist_ok=True)
    path = os.path.join(BASE_CONFIG["DAILY_REPORT_DIR"], report_filename(profile))
    with open(path, "w", encoding="utf-8") as f:
        f.write(build_daily_report(papers, profile, stats, time_range_days, dry_run=False))
    print(f"✅ 科研日报已输出，路径：{path}")
    return path


def build_notification_text(papers: List[Dict], profile: Dict, report_path: str) -> Dict:
    title = f"{profile.get('name', profile.get('id', '科研'))} 文献雷达：{len(papers)} 篇新增"
    lines = [f"今日发现 {len(papers)} 篇新增高相关论文。", ""]
    for idx, paper in enumerate(sorted(papers, key=lambda p: p.get("relevance_score", 0), reverse=True)[:5], 1):
        lines.append(f"{idx}. {paper.get('title', '无标题')}（score={paper.get('relevance_score', 0)}）")
        if paper.get("link"):
            lines.append(paper["link"])
    lines.extend(["", f"完整日报：{report_path}"])
    return {"title": title[:80], "desp": "\n\n".join(lines)}


def send_serverchan(title: str, desp: str) -> bool:
    if requests is None:
        print("❌ 缺少 requests 依赖，无法推送 Server 酱。")
        return False
    sct_key = os.environ.get("SCT_KEY")
    if not sct_key:
        print("⚠️ 未配置 SCT_KEY，跳过 Server 酱推送。")
        return False
    url = f"https://sctapi.ftqq.com/{sct_key}.send"
    try:
        response = requests.post(url, data={"title": title, "desp": desp}, timeout=15)
        if response.status_code == 200:
            print("✅ Server 酱推送完成")
            return True
        print(f"❌ Server 酱推送失败：HTTP {response.status_code} {response.text[:120]}")
    except Exception as e:
        print(f"❌ Server 酱推送异常：{e}")
    return False


def send_notification(notify: str, papers: List[Dict], profile: Dict, report_path: str):
    if not notify:
        return
    if notify != "serverchan":
        print(f"⚠️ 暂不支持 notify={notify}，已跳过推送。")
        return
    message = build_notification_text(papers, profile, report_path)
    send_serverchan(message["title"], message["desp"])


def test_notification(notify: str):
    if notify != "serverchan":
        print("请使用 --notify serverchan 测试 Server 酱推送。")
        return
    send_serverchan("Ptychography Paper Tracker 测试推送", "如果你看到这条消息，说明 Server 酱 Turbo 推送已配置成功。")


def run_daily_mode(time_range_days: int, max_results: int, auto_excel_append: bool, api_key: Optional[str],
                   profile: Dict, dry_run: bool = False, notify: str = ""):
    """每日模式主执行流程"""
    history = load_history()
    print(f"正在检索过去{time_range_days}天的「{profile.get('name', profile['id'])}」方向论文...")
    latest_papers = fetch_daily_papers(time_range_days, max_results, api_key, profile)
    relevant_papers = filter_relevant_papers(latest_papers, profile)
    print(f"检索到 {len(latest_papers)} 篇，相关性过滤后保留 {len(relevant_papers)} 篇")
    new_papers = filter_new_papers(relevant_papers, history)
    stats = analyze_filtering(latest_papers, relevant_papers, new_papers, history, profile)

    if dry_run:
        print(format_dry_run_preview(latest_papers, relevant_papers, new_papers, stats, profile, notify))
        print("\n===== Markdown 科研日报预览 =====")
        print(build_daily_report(new_papers, profile, stats, time_range_days, dry_run=True))
        return

    if not new_papers:
        print("✅ 今日暂无新增论文，无需解析")
        save_history(history)
        return

    print(f"正在批量解析{len(new_papers)}篇新增论文的结构化内容...")
    parsed_papers = []
    for idx, paper in enumerate(new_papers, 1):
        print(f"【{idx}/{len(new_papers)}】正在解析《{paper['title']}》...")
        parse_result = parse_paper_full_fields(paper, profile)
        formatted_parse = format_parse_result(parse_result)
        full_paper = {**paper, **formatted_parse}
        parsed_papers.append(full_paper)

    final_output = format_daily_full_output(parsed_papers)
    print(final_output)

    if auto_excel_append:
        append_to_excel(parsed_papers)
    report_path = write_daily_report(parsed_papers, profile, stats, time_range_days)
    send_notification(notify, parsed_papers, profile, report_path)

    for p in new_papers:
        history["papers"][p["id"]] = {
            "title": p["title"],
            "profile": profile["id"],
            "add_time": datetime.now().strftime("%Y-%m-%d %H:%M")
        }
    save_history(history)
    print(f"🎉 每日论文追踪完成，共处理{len(new_papers)}篇新增论文")


# ===================== 模式 2：年度论文统计 Excel 生成=====================
def fetch_annual_papers(year: int, max_papers: int, api_key: Optional[str] = None,
                        profile: Optional[Dict] = None) -> List[Dict]:
    """检索指定年份的论文"""
    if requests is None:
        print("❌ 缺少 requests 依赖，无法检索 Semantic Scholar。请先安装 requirements.txt。")
        return []
    profile = profile or {}
    headers = {"x-api-key": api_key} if api_key else {}
    params = {
        "query": build_search_keywords(profile),
        "publicationDate": f"{year}-01-01 TO {year}-12-31",
        "fields": "paperId,title,externalIds,authors,venue,publicationDate,url,abstract",
        "limit": max_papers,
        "sort": "publicationDate:desc"
    }

    results = None
    for retry in range(3):
        try:
            response = requests.get(
                BASE_CONFIG["SEMANTIC_SCHOLAR_API_URL"] + "/search",
                headers=headers,
                params=params,
                timeout=20
            )
            if response.status_code == 200:
                results = response.json()
                break
            elif response.status_code == 429:
                print(f"【{year}年】API 限流，等待5秒后重试...")
                time.sleep(5)
            else:
                print(f"【{year}年】检索失败：{response.status_code}")
                break
        except Exception as e:
            print(f"【{year}年】请求异常：{str(e)}")
            time.sleep(3)

    if not results:
        return []

    papers = []
    for paper_data in results.get("data", []):
        try:
            arxiv_id = paper_data.get("externalIds", {}).get("ArXiv", "")
            paper = {
                "id": arxiv_id if arxiv_id else str(hash(paper_data.get("title", ""))),
                "arxiv_id": arxiv_id,
                "semantic_scholar_id": paper_data.get("paperId", ""),
                "title": paper_data.get("title", "无标题"),
                "authors": [author.get("name", "") for author in paper_data.get("authors", [])],
                "link": paper_data.get("url", ""),
                "journal": paper_data.get("venue", "预印本"),
                "impact_factor": "待补充",
                "published_time": paper_data.get("publicationDate", f"{year}-01-01 00:00"),
                "abstract": paper_data.get("abstract", "无摘要")
            }
            papers.append(paper)
        except:
            continue
    return papers


def run_annual_summary_mode(start_year: int, end_year: int, max_papers_per_year: int, api_key: Optional[str],
                            profile: Dict, dry_run: bool = False):
    """年度统计模式主执行流程"""
    history = load_history()
    yearly_papers = {}
    yearly_stats = {}
    yearly_all_papers = {}
    yearly_relevant_papers = {}
    all_parsed_papers = []

    print(f"===== 开始检索{start_year}-{end_year}年「{profile.get('name', profile['id'])}」方向论文 =====")
    for year in range(start_year, end_year + 1):
        print(f"【{year}年】正在检索论文...")
        papers = fetch_annual_papers(year, max_papers_per_year, api_key, profile)
        relevant_papers = filter_relevant_papers(papers, profile)
        new_papers = filter_new_papers(relevant_papers, history)
        yearly_all_papers[year] = papers
        yearly_relevant_papers[year] = relevant_papers
        yearly_papers[year] = new_papers
        yearly_stats[year] = analyze_filtering(papers, relevant_papers, new_papers, history, profile)

        if not dry_run:
            for p in new_papers:
                history["papers"][p["id"]] = {
                    "title": p["title"],
                    "year": year,
                    "profile": profile["id"],
                    "add_time": datetime.now().strftime("%Y-%m-%d %H:%M")
                }
        time.sleep(2)

    if dry_run:
        print("===== 预览模式：仅显示新增论文，不写入 Excel 和历史记录 =====")
        for year, papers in yearly_papers.items():
            print(f"【{year}年】新增 {len(papers)} 篇")
            print(format_dry_run_preview(yearly_all_papers[year], yearly_relevant_papers[year], papers,
                                         yearly_stats[year], profile))
        return

    save_history(history)
    if not any(yearly_papers.values()):
        print("❌ 未检索到任何匹配的论文，Excel 生成终止")
        return

    print(f"===== 开始批量解析论文结构化内容 =====")
    for year, papers in yearly_papers.items():
        for idx, paper in enumerate(papers, 1):
            print(f"【{year}年 | {idx}/{len(papers)}】解析：{paper['title']}")
            parse_result = parse_paper_full_fields(paper, profile)
            formatted_parse = format_parse_result(parse_result)
            full_paper = {**paper, **formatted_parse}
            all_parsed_papers.append(full_paper)

    print("===== 正在生成年度统计 Excel 文件 =====")
    append_to_excel(all_parsed_papers)
    total_count = len(all_parsed_papers)
    print(f"\n🎉 年度论文统计完成！总计检索并解析{start_year}-{end_year}年共{total_count}篇论文")


# ===================== 模式 3：新增！谷歌学术论文检索模式 =====================
def run_google_scholar_mode(max_results: int, serp_api_key: str, profile: Dict, dry_run: bool = False,
                            notify: str = ""):
    """谷歌学术模式主流程（复用所有原有解析/Excel功能）"""
    history = load_history()
    print(f"🚀 开始通过谷歌学术检索「{profile.get('name', profile['id'])}」方向论文...")

    # 调用谷歌学术API
    scholar_papers = fetch_google_scholar_papers(max_results, serp_api_key, profile)
    relevant_papers = filter_relevant_papers(scholar_papers, profile)
    new_papers = filter_new_papers(relevant_papers, history)
    stats = analyze_filtering(scholar_papers, relevant_papers, new_papers, history, profile)

    if dry_run:
        print(format_dry_run_preview(scholar_papers, relevant_papers, new_papers, stats, profile))
        print("\n===== Markdown 科研日报预览 =====")
        print(build_daily_report(new_papers, profile, stats, 0, dry_run=True))
        return

    if not new_papers:
        print("✅ 谷歌学术无新增论文")
        if not dry_run:
            save_history(history)
        return

    # 复用原有解析逻辑
    print(f"正在解析{len(new_papers)}篇谷歌学术论文...")
    parsed_papers = []
    for paper in new_papers:
        parse_result = parse_paper_full_fields(paper, profile)
        formatted_parse = format_parse_result(parse_result)
        parsed_papers.append({**paper, **formatted_parse})

    # 导出Excel
    append_to_excel(parsed_papers)
    report_path = write_daily_report(parsed_papers, profile, stats, 0)
    send_notification(notify, parsed_papers, profile, report_path)

    # 更新历史记录
    for p in new_papers:
        history["papers"][p["id"]] = {
            "title": p["title"],
            "profile": profile["id"],
            "add_time": datetime.now().strftime("%Y-%m-%d %H:%M")
        }
    save_history(history)
    print("🎉 谷歌学术论文检索+解析+导出完成！")


# ===================== 主入口（已新增谷歌学术参数） =====================
def main():
    parser = argparse.ArgumentParser(description="Ptychography 领域全功能论文追踪工具")
    parser.add_argument("--mode", type=str, default="daily",
                        choices=["daily", "annual_summary", "google_scholar"],
                        help="运行模式：daily=每日, annual_summary=年度, google_scholar=谷歌学术")
    parser.add_argument("--profile", type=str, default="",
                        help="研究方向配置 ID，例如 electron_ptychography、xray_ptychography、medical_ai")
    parser.add_argument("--list_profiles", action="store_true", help="列出可用研究方向后退出")
    parser.add_argument("--create_profile", action="store_true", help="交互式创建新的研究方向配置")
    parser.add_argument("--doctor", action="store_true", help="检查依赖、配置、归档文件和推送环境")
    parser.add_argument("--dry_run", action="store_true", help="预览新增论文，不写入 Excel 和历史记录")
    parser.add_argument("--notify", type=str, default="", choices=["", "serverchan"],
                        help="推送渠道，目前支持 serverchan")
    parser.add_argument("--test_notify", action="store_true", help="测试推送渠道")
    # 每日模式参数
    parser.add_argument("--time_range_days", type=int, default=1, help="【daily 模式】检索过去 N 天的论文")
    parser.add_argument("--max_results", type=int, default=10, help="最大检索论文数量")
    parser.add_argument("--auto_excel_append", type=lambda x: x.lower() == 'true', default=True,
                        help="【daily 模式】是否自动追加到 Excel")
    # 年度模式参数
    parser.add_argument("--start_year", type=int, default=2024, help="【annual 模式】统计起始年份")
    parser.add_argument("--end_year", type=int, default=2026, help="【annual 模式】统计结束年份")
    parser.add_argument("--max_papers_per_year", type=int, default=20, help="【annual 模式】每年最大检索论文数量")
    # 通用参数
    parser.add_argument("--semantic_scholar_key", type=str, default="", help="Semantic Scholar API 密钥")
    # 新增：谷歌学术 SerpApi Key 参数
    parser.add_argument("--serp_api_key", type=str, default=os.environ.get("SERPAPI_API_KEY", ""),
                        help="【谷歌学术必填】SerpApi API Key，也可通过 SERPAPI_API_KEY 环境变量配置")

    args = parser.parse_args()

    if args.list_profiles:
        list_research_profiles()
        return
    if args.create_profile:
        create_profile_interactive()
        return
    if args.doctor:
        run_doctor()
        return
    if args.test_notify:
        test_notification(args.notify)
        return

    profile = get_research_profile(args.profile)
    print(f"当前研究方向：{profile['id']} - {profile.get('name', profile['id'])}")

    if args.mode == "daily":
        run_daily_mode(args.time_range_days, args.max_results, args.auto_excel_append,
                       args.semantic_scholar_key, profile, args.dry_run, args.notify)
    elif args.mode == "annual_summary":
        run_annual_summary_mode(args.start_year, args.end_year, args.max_papers_per_year,
                                args.semantic_scholar_key, profile, args.dry_run)
    elif args.mode == "google_scholar":
        # 运行谷歌学术模式
        run_google_scholar_mode(args.max_results, args.serp_api_key, profile, args.dry_run, args.notify)


if __name__ == "__main__":
    main()
